import customtkinter
from tkinter import ttk
import re
from tkinter import *
import tkinter as tk
import bcrypt
from tkinter import messagebox
import sqlite3
from tkcalendar import DateEntry
import datetime
from datetime import date
from datetime import datetime
import random
from PIL import Image
from winotify import Notification, audio
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from ttkthemes import ThemedStyle
# from distinctipy import distinctipy
from tkinter import filedialog
import openpyxl
from matplotlib.backends.backend_pdf import PdfPages
from fpdf import FPDF


# -------------------- class declaration ----------------------

# ---Class declaration for customtkinter label ---
class Inv_Label:
    def __init__(self, master, row, col, padx, info):
        self.label = customtkinter.CTkLabel(master=master, text=info, text_color='black',
                                            font=customtkinter.CTkFont("SF Pro Display", size=13), )
        self.label.grid(row=row, column=col, padx=padx, pady=3, sticky="w")

    def updateInfo(self, newtxt):
        self.label.configure(text=newtxt)

class Inv_titlelabel:
    def __init__(self, master, row, column, padx, text):
        self.label = customtkinter.CTkLabel(master=master, text=text, text_color='black',
                                            font=customtkinter.CTkFont("SF Pro Display", weight="bold", size=15), )
        self.label.grid(row=row, column=column, sticky=W, padx=padx, pady=10)

class Inv_main_titlelabel:
    def __init__(self, master, row, column, padx, pady, text):
        self.label = customtkinter.CTkLabel(master=master, text=text, text_color='black',
                                            font=customtkinter.CTkFont("SF Pro Display", weight="bold", size=20), )
        self.label.grid(row=row, column=column, sticky=W, padx=padx, pady=pady)


# ---Class declaration for customtkinter Button ---
class Inv_Button:
    def __init__(self, master, row, col, padx, color, info, cmd):
        self.button = customtkinter.CTkButton(master=master, text=info, fg_color=color, text_color="black",
                                              font=customtkinter.CTkFont("SF Pro Display", weight="bold", size=13),
                                              command=cmd, compound='top', corner_radius=200)
        self.button.grid(row=row, column=col, columnspan=2, padx=padx, pady=5)


# ---Class declaration for customtkinter ComboBox ---
class Inv_ComboBox:
    def __init__(self, master, row, col, padx, list):
        self.value = ""
        self.combobox = customtkinter.CTkComboBox(master=master, values=list, width=230, height=30,
                                                  font=customtkinter.CTkFont("SF Pro Display"), border_width=0, )
        self.combobox.grid(row=row, column=col, padx=padx, pady=10, sticky="w")

    def getvalue(self):
        self.value = self.combobox.get()
        return self.value

# ---Class declaration for customtkinter Entry Box ---
class Inv_Entrybox:
    def __init__(self, master, row, col, padx, info):
        self.value = ""
        self.entry = customtkinter.CTkEntry(master, placeholder_text=info, width=230, height=30, border_width=0, # fg_color='lightyellow',
                                            font=customtkinter.CTkFont("SF Pro Display", 13), )
        self.entry.grid(row=row, column=col, sticky=W, padx=padx, pady=3)

    def getvalue(self):
        self.value = self.entry.get()
        return self.value

    def clearField(self):
        self.entry.delete(0, END)

    def insertField(self, info):
        self.entry.insert(0, info)


# ---Class declaration for SQL-Database PRODUCT table ---
class Inv_Product_Database:
    def __init__(self):
        self.sql_filename = 'Inventory Management System.db'

    def insertRecord(self, id, name, qty, desc, min):
        conn = sqlite3.connect(self.sql_filename)
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO PRODUCT (PRODUCT_ID,PRODUCT_NAME,PRODUCT_QUANTITY,PRODUCT_STATUS,PRODUCT_MIN_STOCK) \
        VALUES (?,?,?,?,?)",
            (id, name, qty, desc, min), )
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Product Data record has been inserted!")

    def updateRecord(self, id, name, qty, desc, min, prod_id):
        conn = sqlite3.connect(self.sql_filename)
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE PRODUCT SET PRODUCT_ID = ?, PRODUCT_NAME = ?, PRODUCT_QUANTITY = ?, PRODUCT_STATUS = ?,\
             PRODUCT_MIN_STOCK = ? WHERE PRODUCT_ID = ?",
            (id, name, qty, desc, min, prod_id), )
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Product Record successfully edited!")

    def deleteRecord(self, prod_id):
        conn = sqlite3.connect(self.sql_filename)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM PRODUCT WHERE PRODUCT_ID= ?", (prod_id,))
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Product record entry successfully deleted!!")

    def check_existing_product(self, product_check):
        conn = sqlite3.connect(self.sql_filename)
        cursor = conn.cursor()
        cursor.execute("SELECT PRODUCT_ID FROM PRODUCT WHERE PRODUCT_ID = ?", (product_check,), )
        existing_product = cursor.fetchone()
        conn.commit()
        conn.close()
        return existing_product

    def fetch_product_data(self):
        conn = sqlite3.connect(self.sql_filename)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM PRODUCT")
        product = cursor.fetchall()
        conn.commit()
        conn.close()
        return product

    def fetch_total_quantity_in_hand_data(self):
        conn = sqlite3.connect(self.sql_filename)
        cursor = conn.cursor()
        cursor.execute("SELECT SUM(PRODUCT_QUANTITY) FROM PRODUCT")
        total_quantity_in_hand = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return total_quantity_in_hand

    def fetch_total_items_data(self):
        conn = sqlite3.connect(self.sql_filename)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(PRODUCT_ID) FROM PRODUCT")
        total_items = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return total_items

    def fetch_low_stock_item_data(self):
        conn = sqlite3.connect(self.sql_filename)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(PRODUCT_ID) FROM PRODUCT WHERE PRODUCT_QUANTITY < PRODUCT_MIN_STOCK")
        low_stock_items = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return low_stock_items

    def fetch_bar_chart_data(self):
        conn = sqlite3.connect(self.sql_filename)
        cursor = conn.cursor()
        cursor.execute("SELECT PRODUCT_ID, PRODUCT_QUANTITY, PRODUCT_MIN_STOCK FROM PRODUCT")
        data = cursor.fetchall()
        conn.commit()
        conn.close()
        return data


# ---Class declaration for SQL-Database PURCHASE_ORDER table ---
class Inv_Purchase_Order_dbase:
    def __init__(self):
        self.sql_filename = 'Inventory Management System.db'

    def insertRecord(self, id, product, quantity, status):
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO PURCHASE_ORDER (PURCHASE_ORDER_ID,PURCHASE_ORDER_PRODUCT,PURCHASE_ORDER_PRODUCT_QUANTITY,PURCHASE_ORDER_STATUS) \
        VALUES (?,?,?,?)",
            (id, product, quantity, status,), )
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Purchase Order Data has been inserted")


# ---Class declaration for product tree view table display ---
class Inv_product_tree_display:
    def __init__(self, master, dclick_cmd, buttonrelease_cmd):
        style = ThemedStyle()
        style.set_theme("equilux")
        style.configure(
            "Treeview",
            font=("SF Pro Display", 10),
            rowheight=24,
            foreground="white",
            background="#2a2d2e",
            fieldbackground="#343638",
            bordercolor="#343638",
            borderwidth=0,
        )
        style.map("Treeview", background=[("selected", "#22559b")])
        style.configure(
            "Treeview.Heading", background="#565b5e", foreground="white", relief="flat"
        )
        style.map("Treeview.Heading", background=[("active", "#3484F0")])
        style.configure(
            "DateEntry",
            font=customtkinter.CTkFont(("SF Pro Display"), 12),
            foreground="white",
            fieldbackground="#343638",
            bordercolor="#343638",
            borderwidth=0,
        )

        def sort_treeview_column(tv, col, reverse):
            l = [(tv.set(k, col), k) for k in tv.get_children('')]
            l.sort(reverse=reverse)
            for index, (val, k) in enumerate(l):
                tv.move(k, '', index)
            tv.heading(col, command=lambda: sort_treeview_column(tv, col, not reverse))

        self.product_tree = ttk.Treeview(master=master, height=18)
        self.product_verscrlbar = customtkinter.CTkScrollbar(
            master=master, orientation="vertical", command=self.product_tree.yview
        )

        self.product_tree.configure(yscrollcommand=self.product_verscrlbar.set)
        self.product_tree["columns"] = (
            "PRODUCTID",
            "PRODUCTNAME",
            "PRODUCTQUANTITY",
            "PRODUCTDESCRIPTION",
            "PRODUCTMINSTOCK",
        )
        self.product_tree.column("#0", width=0, stretch=tk.NO)
        self.product_tree.column("PRODUCTID", anchor=tk.CENTER, width=250)
        self.product_tree.column("PRODUCTNAME", anchor=tk.CENTER, width=250)
        self.product_tree.column("PRODUCTQUANTITY", anchor=tk.CENTER, width=250)
        self.product_tree.column("PRODUCTDESCRIPTION", anchor=tk.CENTER, width=250)
        self.product_tree.column("PRODUCTMINSTOCK", anchor=tk.CENTER, width=250)
        self.product_tree.heading("PRODUCTID", text="ID",
                                  command=lambda: sort_treeview_column(self.product_tree, "PRODUCTID", False))
        self.product_tree.heading("PRODUCTNAME", text="Name")
        self.product_tree.heading("PRODUCTQUANTITY", text="Quantity")
        self.product_tree.heading("PRODUCTDESCRIPTION", text="Description")
        self.product_tree.heading("PRODUCTMINSTOCK", text="Min. Stock")

        # Use grid layout to place the treeview and scrollbar properly
        self.product_tree.grid(row=0, column=0, sticky='nsew')
        self.product_verscrlbar.grid(row=0, column=1, sticky='ns')

        # Configure the grid weights to ensure proper resizing
        master.grid_rowconfigure(0, weight=1)
        master.grid_columnconfigure(0, weight=1)

        self.product_tree.bind("<Double-1>", dclick_cmd)
        self.product_tree.bind("<ButtonRelease>", buttonrelease_cmd)

    def add_to_table(self, products):
        self.product_tree.delete(*self.product_tree.get_children())
        for product in products:
            self.product_tree.insert("", END, values=product)

    def selected_item(self):
        selected_item = self.product_tree.focus()
        row = self.product_tree.item(selected_item)["values"]
        return row

    def search_item(self, search, products):
        self.product_tree.delete(*self.product_tree.get_children())
        for product in products:
            if search in str(product).lower():
                self.product_tree.insert("", tk.END, values=product)

# ---Class declaration for outgoing stock tree view table display ---
class Inv_outgoing_stock_tree_display:
    def __init__(self, master, buttonrelease_cmd):
            self.outgoing_stock_tree = ttk.Treeview(master=master, height=20)

            self.outgoing_stock_verscrlbar = customtkinter.CTkScrollbar(
                    master=master,
                    orientation="vertical",
                    command=self.outgoing_stock_tree.yview,
            )
            self.outgoing_stock_verscrlbar.grid(row=0, column=1)

            self.outgoing_stock_tree.configure(yscrollcommand=self.outgoing_stock_verscrlbar.set)
            self.outgoing_stock_tree["columns"] = ("OUTGOINGSTOCKID", "OUTGOINGSTOCKNAME",
                                                       "OUTGOINGSTOCKQUANTITY", "OUTGOINGSTOCKSTATUS"
                                                       )
            self.outgoing_stock_tree.column("#0", width=0, stretch=tk.NO)
            self.outgoing_stock_tree.column("OUTGOINGSTOCKID", anchor=tk.CENTER, width=103)
            self.outgoing_stock_tree.column("OUTGOINGSTOCKNAME", anchor=tk.CENTER, width=103)
            self.outgoing_stock_tree.column("OUTGOINGSTOCKQUANTITY", anchor=tk.CENTER, width=303)
            self.outgoing_stock_tree.column("OUTGOINGSTOCKSTATUS", anchor=tk.CENTER, width=203)

            self.outgoing_stock_tree.heading("OUTGOINGSTOCKID", text="ID")
            self.outgoing_stock_tree.heading("OUTGOINGSTOCKNAME", text="Date Created")
            self.outgoing_stock_tree.heading("OUTGOINGSTOCKQUANTITY", text="Customer Name")
            self.outgoing_stock_tree.heading("OUTGOINGSTOCKSTATUS", text="Status")

            self.outgoing_stock_tree.grid(row=0, column=0)
            self.outgoing_stock_tree.bind("<Double-1>", buttonrelease_cmd)

    def add_to_table(self, outgoing_stock):
            self.outgoing_stock_tree.delete(*self.outgoing_stock_tree.get_children())
            for item in outgoing_stock:
                    self.outgoing_stock_tree.insert("", END, values=item)

    def selected_item(self):
            selected_item = self.outgoing_stock_tree.focus()
            row = self.outgoing_stock_tree.item(selected_item)["values"]
            return row

# ---Class declaration for outgoing product tree view table display ---
class Inv_outgoing_product_tree_display:
    def __init__(self, master):
            self.outgoing_product_tree = ttk.Treeview(master=master, height=20)
            self.outgoing_product_verscrlbar = customtkinter.CTkScrollbar(
                    master=master,
                    orientation="vertical",
                    command=self.outgoing_product_tree.yview,
            )
            self.outgoing_product_verscrlbar.grid(row=0, column=1)

            self.outgoing_product_tree.configure(yscrollcommand=self.outgoing_product_verscrlbar.set)
            self.outgoing_product_tree["columns"] = (
                    "OUTGOING_PRODUCTID",
                    "OUTGOING_PRODUCTNAME",
                    "OUTGOING_QUANTITY",
            )
            self.outgoing_product_tree.column("#0", width=0, stretch=tk.NO)
            self.outgoing_product_tree.column("OUTGOING_PRODUCTID", anchor=tk.CENTER, width=92)
            self.outgoing_product_tree.column("OUTGOING_PRODUCTNAME", anchor=tk.CENTER, width=300)
            self.outgoing_product_tree.column("OUTGOING_QUANTITY", anchor=tk.CENTER, width=120)

            self.outgoing_product_tree.heading("OUTGOING_PRODUCTID", text="Product ID")
            self.outgoing_product_tree.heading("OUTGOING_PRODUCTNAME", text="Product Name")
            self.outgoing_product_tree.heading("OUTGOING_QUANTITY", text="         Total\nOutgoing Quantity")

            self.outgoing_product_tree.grid(row=0, column=0)

    def add_to_table(self, outgoing_product):
                self.outgoing_product_tree.delete(*self.outgoing_product_tree.get_children())
                for item in outgoing_product:
                    self.outgoing_product_tree.insert("", END, values=item)

# ---Class declaration for outgoing stock tree view table display ---
class Inv_outgoing_stock_tree_display:
    def __init__(self, master, buttonrelease_cmd):
            self.outgoing_stock_tree = ttk.Treeview(master=master, height=20)

            self.outgoing_stock_verscrlbar = customtkinter.CTkScrollbar(
                    master=master,
                    orientation="vertical",
                    command=self.outgoing_stock_tree.yview,
            )
            self.outgoing_stock_verscrlbar.grid(row=0, column=1)

            self.outgoing_stock_tree.configure(yscrollcommand=self.outgoing_stock_verscrlbar.set)
            self.outgoing_stock_tree["columns"] = ("OUTGOINGSTOCKID", "OUTGOINGSTOCKNAME",
                                                       "OUTGOINGSTOCKQUANTITY", "OUTGOINGSTOCKSTATUS"
                                                       )
            self.outgoing_stock_tree.column("#0", width=0, stretch=tk.NO)
            self.outgoing_stock_tree.column("OUTGOINGSTOCKID", anchor=tk.CENTER, width=103)
            self.outgoing_stock_tree.column("OUTGOINGSTOCKNAME", anchor=tk.CENTER, width=103)
            self.outgoing_stock_tree.column("OUTGOINGSTOCKQUANTITY", anchor=tk.CENTER, width=303)
            self.outgoing_stock_tree.column("OUTGOINGSTOCKSTATUS", anchor=tk.CENTER, width=203)

            self.outgoing_stock_tree.heading("OUTGOINGSTOCKID", text="ID")
            self.outgoing_stock_tree.heading("OUTGOINGSTOCKNAME", text="Date Created")
            self.outgoing_stock_tree.heading("OUTGOINGSTOCKQUANTITY", text="Customer Name")
            self.outgoing_stock_tree.heading("OUTGOINGSTOCKSTATUS", text="Status")

            self.outgoing_stock_tree.grid(row=0, column=0)
            self.outgoing_stock_tree.bind("<Double-1>", buttonrelease_cmd)

    def add_to_table(self, outgoing_stock):
            self.outgoing_stock_tree.delete(*self.outgoing_stock_tree.get_children())
            for item in outgoing_stock:
                    self.outgoing_stock_tree.insert("", END, values=item)

    def selected_item(self):
            selected_item = self.outgoing_stock_tree.focus()
            row = self.outgoing_stock_tree.item(selected_item)["values"]
            return row

    def search_item(self, search, outgoing_stocks):
        self.outgoing_stock_tree.delete(*self.outgoing_stock_tree.get_children())
        for outgoing_stock in outgoing_stocks:
            if search in str(outgoing_stock).lower():
                self.outgoing_stock_tree.insert("", tk.END, values=outgoing_stock)

# ------ Class declaration for pie chart ------
class Inv_pie_chart:
    def __init__(self, master, row, col):
        fig = plt.Figure()
        fig.set_facecolor('#e5e5e5')
        self.canvas = FigureCanvasTkAgg(fig, master=master)
        self.canvas.get_tk_widget().grid(row=row, column=col, padx=10, pady=10, columnspan=3, sticky="W")

    def display(self, total_items, low_stock_items):
        data = []
        data.append(low_stock_items)
        data.append(total_items - low_stock_items)

        fig = plt.Figure()
        ax = fig.add_subplot(111)
        fig.patch.set_facecolor('#e5e5e5')  # Set the figure background color
        ax.set_facecolor('#e5e5e5')  # Set the axes background color
        ax.axis("equal")

        ax.pie(data, labels=["Low Stock", "In Stock"], autopct="%.2f%%", pctdistance=0.85)
        centre_circle = plt.Circle((0, 0), 0.70, fc='white')
        ax.add_artist(centre_circle)

        self.canvas.figure = fig
        self.canvas.draw()

# ------ Class declaration for bar chart ------
class Inv_bar_chart:
    def __init__(self, master, row, col):
        fig1 = plt.Figure()
        fig1.set_facecolor('#e5e5e5')
        self.canvas1 = FigureCanvasTkAgg(fig1, master=master)
        self.canvas1.get_tk_widget().grid(row=row, column=col, padx=10, pady=10, columnspan=3, sticky="E")

    def display(self, data):
        product_ids = [row[0] for row in data]
        product_quantities = [row[1] for row in data]
        product_min_stocks = [row[2] for row in data]

        fig1 = plt.Figure()
        ax = fig1.add_subplot(111)
        fig1.patch.set_facecolor('#e5e5e5')
        ax.set_facecolor('#e5e5e5')

        colors = ['red' if qty < min_stock else 'blue' for qty, min_stock in
              zip(product_quantities, product_min_stocks)]
        ax.bar(product_ids, product_quantities, color=colors)

        ax.tick_params(axis='x', labelsize=5)
        ax.tick_params(axis='y', labelsize=10)
        ax.set_xlabel('Product Name', fontsize=12)
        ax.set_ylabel('Product Quantity', fontsize=12)
        ax.set_title('Inventory', fontsize=14)
        plt.setp(ax.get_xticklabels(), rotation=45, ha='right', rotation_mode='anchor')

        # Adding legend
        from matplotlib.lines import Line2D
        legend_elements = [Line2D([0], [0], color='red', lw=4, label='Low in Stock'),
                           Line2D([0], [0], color='blue', lw=4, label='Sufficient Stock')]
        ax.legend(handles=legend_elements, loc='upper right')

        self.canvas1.figure = fig1
        self.canvas1.draw()


# ---------------------- end of class declaration -------------------

app = customtkinter.CTk()
app.geometry("1280x720")
app.resizable(False, False)
app.title("Invy")
customtkinter.set_default_color_theme("dark-blue")
customtkinter.set_appearance_mode("Light")

def login_page():
    def fetch_user_activities_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(USER_ACTIVITIES_ID) FROM USER_ACTIVITIES")
        last_user_activities_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_user_activities_id

    def generate_user_activities_id(prefix="UA"):
        last_customer_id = fetch_user_activities_last_id()
        if last_customer_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_customer_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def login():
        name = log_user.get()
        code = log_password.get()
        userbytes = code.encode("utf-8")
        if name == "" or code == "":
            messagebox.showerror("Empty", "Please fill in the empty field")
        else:
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.execute("SELECT USERNAME FROM USER WHERE USERNAME=?", (name,))
            if cursor.fetchone():
                cursor = conn.execute(
                    "SELECT PASSWORD FROM USER WHERE USERNAME=?", (name,)
                )
                stored_password_tuple = cursor.fetchone()
                if stored_password_tuple:
                    cursor2 = conn.execute(
                        "SELECT ACCESSLEVEL FROM USER WHERE USERNAME=?", (name,)
                    )
                    level = cursor2.fetchone()
                    db_password = stored_password_tuple[0]
                    b_stored_password = db_password.encode("utf-8")
                    if bcrypt.checkpw(userbytes, b_stored_password):
                        messagebox.showinfo("Success", "Login Success")
                        if level[0] == 0:
                            login_page_frame.destroy()
                            admin_dashboard(name)
                        if level[0] == 1:
                            login_page_frame.destroy()
                            supervisor_dashboard(name)
                        if level[0] == 2:
                            login_page_frame.destroy()
                            worker_dashboard(name)

                        cursor = conn.cursor()
                        cursor.execute(
                            "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (name,)
                        )
                        fullname = cursor.fetchone()
                        conn.commit()
                        conn.close()

                        conn = sqlite3.connect("Inventory Management System.db")
                        cursor = conn.cursor()
                        cursor.execute(
                            "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID,USER_ACTIVITIES_DATE,USER_ACTIVITIES,USER) \
                        VALUES (?,?,?,?)",
                            (
                                generate_user_activities_id(),
                                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "{} logged in".format(fullname[0]),
                                name,
                            ),
                        )
                        conn.commit()
                        conn.close()


                    else:
                        messagebox.showerror("Invalid", "Invalid Username or Password")
                else:
                    messagebox.showerror("Invalid", "Invalid Username or Password")
            else:
                messagebox.showerror("Invalid", "Invalid Username or Password")

    login_page_frame = customtkinter.CTkFrame(
        master=app, width=1280, height=720, border_width=0
    )
    login_page_frame.place(x=0, y=0)

    image_frame = customtkinter.CTkFrame(master=login_page_frame, width=640, height=720)
    image_frame.place(x=0, y=0)
    logo = customtkinter.CTkImage(Image.open("assets/7078214.png"), size=(150, 150))
    logo_label = customtkinter.CTkLabel(master=image_frame, image=logo, text="")
    logo_label.place(x=70, y=50)

    expense_tracker_label = customtkinter.CTkLabel(
        master=image_frame,
        text="Invy",
        font=customtkinter.CTkFont("SF Pro Display", 80, weight="bold"),
    )
    expense_tracker_label.place(x=70, y=220)
    introduction_label = customtkinter.CTkLabel(
        master=image_frame,
        text="Your streamlined solution for efficient inventory management. Designed to simplify tracking, organizing, and managing your inventory",
        font=customtkinter.CTkFont("SF Pro Display", 16),
        wraplength=450,
        justify="left",
    )
    introduction_label.place(x=80, y=350)

    login_frame = customtkinter.CTkFrame(master=login_page_frame, width=640, height=720)
    login_frame.pack()
    login_frame.place(x=640, y=0)

    log_label = customtkinter.CTkLabel(
        master=login_frame,
        text="Sign in",
        width=212,
        height=80,
        font=customtkinter.CTkFont("SF Pro Display", 64, weight="bold"),
    )
    log_label.place(x=214, y=118.5)

    log_user = customtkinter.CTkEntry(
        master=login_frame,
        placeholder_text="Username",
        width=300,
        height=45,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )
    log_user.place(x=170, y=280.5)
    app.update()
    log_user.focus_set()

    log_password = customtkinter.CTkEntry(
        master=login_frame,
        placeholder_text="Password",
        width=300,
        height=45,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )
    log_password.place(x=170, y=356.5)
    log_password.configure(show="*")

    log_btn_image = customtkinter.CTkImage(
        Image.open("assets/enter.png"), size=(20, 20)
    )
    log_btn = customtkinter.CTkButton(
        master=login_frame,
        text="Login",
        width=300,
        height=45,
        font=customtkinter.CTkFont("SF Pro Display", 15),
        command=login,
        image=log_btn_image,
        compound="right",
        text_color=("black", "white"),
        fg_color="transparent",
        hover=False,
    )
    log_btn.place(x=170, y=436)

    def show_hide_password():
        if log_show_hide.get() == "on":
            log_password.configure(show="")
        else:
            log_password.configure(show="*")

    log_show_hide = customtkinter.StringVar(value="off")
    log_show_hide_password_switch = customtkinter.CTkSwitch(
        master=login_frame,
        text="Show Password",
        command=show_hide_password,
        variable=log_show_hide,
        onvalue="on",
        offvalue="off",
        font=customtkinter.CTkFont("SF Pro Display"),
    )
    log_show_hide_password_switch.place(x=480, y=368.5)

    def switch_theme():
        if log_switch_theme_var.get() == "on":
            customtkinter.set_appearance_mode("Light")
        else:
            customtkinter.set_appearance_mode("Dark")

    log_switch_theme_var = customtkinter.StringVar(value="off")
    log_switch_theme_switch = customtkinter.CTkSwitch(
        master=login_frame,
        text="Switch Theme",
        command=switch_theme,
        variable=log_switch_theme_var,
        onvalue="on",
        offvalue="off",
        font=customtkinter.CTkFont("SF Pro Display"),
    )
    log_switch_theme_switch.place(x=500, y=680)


# =================================== Admin Dashboard ================================================
def admin_dashboard(username):
    product_dbase = Inv_Product_Database()  # declare the product class object.
    purchase_order_dbase = Inv_Purchase_Order_dbase()  # declare the purchase_order class object.

    def fetch_to_be_packed_data():
        total_to_be_packed = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(SALE_ORDER_STATUS) FROM SALE_ORDER WHERE SALE_ORDER_STATUS = ?", ("To be Packed",))
        to_be_packed_data = cursor.fetchone()[0]
        total_to_be_packed += to_be_packed_data
        conn.commit()
        conn.close()
        return total_to_be_packed

    def update_to_be_packed_label():
        new_count = fetch_to_be_packed_data()
        total_to_be_packed_label_1.updateInfo(str(new_count))

    def fetch_to_be_shipped_data():
        total_to_be_shipped = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(SALE_ORDER_STATUS) FROM SALE_ORDER WHERE SALE_ORDER_STATUS = ?",
                       ("To be Shipped",))
        to_be_shipped_data = cursor.fetchone()[0]
        total_to_be_shipped += to_be_shipped_data
        conn.commit()
        conn.close()
        return total_to_be_shipped

    def update_to_be_shipped_label():
        new_count = fetch_to_be_shipped_data()
        total_to_be_shipped_label_1.updateInfo(str(new_count))

    def fetch_to_be_delivered_data():
        total_to_be_delivered = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(SALE_ORDER_STATUS) FROM SALE_ORDER WHERE SALE_ORDER_STATUS = ?",
                       ("To be Delivered",))
        to_be_delivered_data = cursor.fetchone()[0]
        total_to_be_delivered += to_be_delivered_data
        conn.commit()
        conn.close()
        return total_to_be_delivered

    def update_to_be_delivered_label():
        new_count = fetch_to_be_delivered_data()
        total_to_be_delivered_label_1.updateInfo(str(new_count))

    def update_total_quantity_in_hand_label():
        new_count = product_dbase.fetch_total_quantity_in_hand_data()
        total_quantity_in_hand_1.updateInfo(str(new_count))

    def fetch_total_quantity_to_be_received_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT SUM(PURCHASE_ORDER_PRODUCT_QUANTITY) FROM PURCHASE_ORDER WHERE PURCHASE_ORDER_STATUS = ?",
            ("To be Received",))
        total_quantity_to_be_received = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return total_quantity_to_be_received

    def update_total_quantity_to_be_received_label():
        new_count = fetch_total_quantity_to_be_received_data()
        total_quantity_to_be_received_1.updateInfo(str(new_count))

    def update_low_stock_item_label():
        new_count = product_dbase.fetch_low_stock_item_data()
        low_stock_items_1.updateInfo(str(new_count))

    def update_total_items_label():
        new_count = product_dbase.fetch_total_items_data()
        all_items_1.updateInfo(str(new_count))

    def register():
        fullname = reg_user_name.get()
        name = reg_user.get()
        code = reg_password.get()
        level = reg_accesslevel.getvalue()
        if level == "Admin":
            level = 0
        elif level == "Supervisor":
            level = 1
        elif level == "Warehouse Worker":
            level = 2
        bytes = code.encode("utf-8")
        salt = bcrypt.gensalt()
        hashed_password = bcrypt.hashpw(bytes, salt)
        stored_password = hashed_password.decode("utf-8")
        confirm_code = reg_confirm_password.get()
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT USERNAME FROM USER WHERE USERNAME=?", (name,))
        existing_user = cursor.fetchone()
        if fullname == "" or name == "" or code == "" or confirm_code == "":
            messagebox.showerror("Empty", "Please fill in the empty field")
        elif existing_user:
            messagebox.showerror("Invalid", "User already existed")
        elif code != confirm_code:
            messagebox.showerror("Invalid", "Please enter the correct password")
        elif len(code) < 8:
            messagebox.showerror(
                "Invalid", "Make sure your password is at least 8 letters"
            )
        elif re.search("[0-9]", code) is None:
            messagebox.showerror(
                "Invalid", "Make sure your password has a number in it"
            )
        elif re.search("[A-Z]", code) is None:
            messagebox.showerror(
                "Invalid", "Make sure your password has a uppercase letter in it"
            )
        else:
            conn.execute(
                "INSERT INTO USER (USER_FULLNAME,USERNAME,PASSWORD,ACCESSLEVEL) \
            VALUES (?,?,?,?)",
                (fullname, name, stored_password, level),
            )
            conn.commit()
            conn.close()
            messagebox.showinfo("Valid", "Account created")
            add_to_user_table()

    def fetch_user_fullname():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,))
        user = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return user

    def fetch_user_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT USER_FULLNAME, USERNAME, ACCESSLEVEL FROM USER")
        user = cursor.fetchall()
        conn.commit()
        conn.close()
        return user

    def add_to_user_table():
        users = fetch_user_data()
        user_tree.delete(*user_tree.get_children())
        for user in users:
            user_tree.insert("", END, values=user)

    def on_user_double_click(event):
        selected_item = user_tree.selection()[0]
        values = user_tree.item(selected_item, "values")

        change_pw_window = customtkinter.CTk()
        change_pw_window.title("Change Password")
        change_pw_window.resizable(False, False)

        def change_password():
            new_password = new_password_entry.getvalue()
            bytes = new_password.encode("utf-8")
            salt = bcrypt.gensalt()
            hashed_password = bcrypt.hashpw(bytes, salt)
            stored_password = hashed_password.decode("utf-8")
            if not new_password:
                messagebox.showerror("Empty", "Please fill in the empty field")
            elif len(new_password) < 8:
                messagebox.showerror(
                    "Invalid", "Make sure your password is at least 8 letters"
                )
            elif re.search("[0-9]", new_password) is None:
                messagebox.showerror(
                    "Invalid", "Make sure your password has a number in it"
                )
            elif re.search("[A-Z]", new_password) is None:
                messagebox.showerror(
                    "Invalid", "Make sure your password has a uppercase letter in it"
                )
            else:
                conn = sqlite3.connect("Inventory Management System.db")
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE USER SET PASSWORD = ? WHERE USERNAME = ?",
                    (stored_password, values[1]),
                )
                conn.commit()
                conn.close()
                messagebox.showinfo("Success", "New Password updated!")
                change_pw_window.destroy()

        change_pw_title = Inv_titlelabel(change_pw_window, 0, 0, 10, "Change Password")
        targeted_user_label = Inv_Label(change_pw_window, 1, 0, (10, 0), "Targeted User:")
        targeted_user_label_details = Inv_Label(change_pw_window, 1, 1, 5, f"{values[0]}")
        new_password_label = Inv_Label(change_pw_window, 2, 0, (10, 0), "New Password:")
        new_password_entry = Inv_Entrybox(change_pw_window, 2, 1, 5, "Password")
        change_pw_btn = Inv_Button(change_pw_window, 3, 0, (10, 0), "#007FFF", "Change Password", change_password)

        change_pw_window.mainloop()

    def delete_user_database():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        selected_item = user_tree.focus()
        row = user_tree.item(selected_item)["values"]
        cursor.execute("DELETE FROM USER WHERE USERNAME= ?", (row[1],))
        conn.commit()
        conn.close()
        add_to_user_table()

    def delete_user_record():
        selected_item = user_tree.focus()
        if not selected_item:
            messagebox.showerror("Error", "Please select a record to delete")
            return
        confirmation = messagebox.askyesno(
            "Are you sure?", "Are you sure that you want to delete the selected record?"
        )
        if confirmation:
            delete_user_database()
            messagebox.showinfo("Success", "Record successfully deleted")
            return

    def add_to_product_table():
        products = product_dbase.fetch_product_data()
        admin_product_tree.add_to_table(products)

    def clear_product_entry_field():
        product_id_entry.clearField()
        product_name_entry.clearField()
        product_quantity_entry.clearField()
        product_description_entry.clearField()
        product_min_stock_entry.clearField()

    def display_product_record(event):
        selected_item = admin_product_tree.selected_item()
        if selected_item:
            clear_product_entry_field()
            row = selected_item
            product_id_entry.insertField([row[0]])
            product_name_entry.insertField(row[1])
            product_quantity_entry.insertField(row[2])
            product_description_entry.insertField(row[3])
            product_min_stock_entry.insertField([row[4]])

    def add_new_product_details():
        product_id = product_id_entry.getvalue()
        product_name = product_name_entry.getvalue()
        product_quantity = product_quantity_entry.getvalue()
        product_description = product_description_entry.getvalue()
        product_min_stock = product_min_stock_entry.getvalue()

        # Validate that product_id is not empty
        if not product_id:
            messagebox.showerror("Error", "Please enter product id")
            return

        # Validate that all fields are filled
        if not (product_name and product_quantity and product_description):
            messagebox.showerror("Error", "Please enter all fields")
            return

        # Validate that product_quantity is an integer
        if not product_quantity.isdigit() or int(product_quantity) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for product quantity")
            return

        # Validate that product_min_stock is an integer
        if not product_min_stock.isdigit() or int(product_min_stock) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for minimum stock")
            return

        product_quantity = int(product_quantity)
        product_min_stock = int(product_min_stock)

        # Check if product already exists
        if not product_dbase.check_existing_product(product_id):
            try:
                product_dbase.insertRecord(
                    product_id,
                    product_name,
                    product_quantity,
                    product_description,
                    product_min_stock,
                )

                conn = sqlite3.connect("Inventory Management System.db")
                cursor = conn.cursor()

                # Retrieve the user's full name
                cursor.execute(
                    "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
                )
                fullname = cursor.fetchone()

                # Log the user activity
                cursor.execute(
                    "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_user_activities_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "{} added new product: {} with ID: {}".format(fullname[0], product_name, product_id),
                        username,
                    ),
                )
                conn.commit()

                # Log the product movement
                cursor.execute(
                    "INSERT INTO PRODUCT_MOVEMENT (PRODUCT_MOVEMENT_ID, PRODUCT_MOVEMENT_DATE, PRODUCT_MOVEMENT, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_product_movement_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Added product: {} with ID: {} by {} to warehouse".format(product_name, product_id,
                                                                                  fullname[0]),
                        username,
                    ),
                )
                conn.commit()
                conn.close()

            except Exception as e:
                messagebox.showerror("Error", str(e))

        else:
            messagebox.showerror("Warning", "Duplicate product! Please enter new product.")

        add_to_product_table()
        clear_product_entry_field()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()
        low_stock_vs_total_item_pie_chart()
        bar_chart()
        return

    def edit_product_details():
        selected_product_details = admin_product_tree.selected_item()
        if not selected_product_details:
            messagebox.showerror("Error", "Please select a record to edit")
            return

        row = selected_product_details
        new_product_id = product_id_entry.getvalue()
        new_product_name = product_name_entry.getvalue()
        new_product_quantity = product_quantity_entry.getvalue()
        new_product_description = product_description_entry.getvalue()
        new_product_min_stock = product_min_stock_entry.getvalue()

        if not new_product_name:
            messagebox.showerror("Error", "Please enter all fields")
            return

        if not new_product_quantity.isdigit() or int(new_product_quantity) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for product quantity")
            return

        # Validate that product_min_stock is an integer
        if not new_product_min_stock.isdigit() or int(new_product_min_stock) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for minimum stock")
            return

        old_product_id = row[0]
        old_product_name = row[1]
        old_product_quantity = row[2]
        old_product_description = row[3]
        old_product_min_stock = row[4]
        # old_preferred_supplier = row[5]

        try:
            product_dbase.updateRecord(
                new_product_id,
                new_product_name,
                new_product_quantity,
                new_product_description,
                new_product_min_stock,
                new_product_id)

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            # Retrieve the user's full name
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
            )
            fullname = cursor.fetchone()

            # Determine what was edited
            changes = []
            if old_product_id != new_product_id:
                changes.append(f"ID from '{old_product_id}' to '{new_product_id}'")
            if old_product_name != new_product_name:
                changes.append(f"Name from '{old_product_name}' to '{new_product_name}'")
            if old_product_quantity != new_product_quantity:
                changes.append(f"Quantity from '{old_product_quantity}' to '{new_product_quantity}'")
            if old_product_description != new_product_description:
                changes.append(f"Description from '{old_product_description}' to '{new_product_description}'")
            if old_product_min_stock != new_product_min_stock:
                changes.append(f"Min stock from '{old_product_min_stock}' to '{new_product_min_stock}'")
            # if old_preferred_supplier != new_preferred_supplier:
            #    changes.append(f"Preferred supplier from '{old_preferred_supplier}' to '{new_preferred_supplier}'")

            changes_str = "; ".join(changes)

            # Log the user activity
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    f"{fullname[0]} edited product {old_product_name}: {changes_str}",
                    username,
                ),
            )
            conn.commit()
            conn.close()

        except Exception as e:
            messagebox.showerror("Error", str(e))

        add_to_product_table()
        clear_product_entry_field()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()
        low_stock_vs_total_item_pie_chart()
        bar_chart()
        return

    def delete_product_record():
        selected_item = admin_product_tree.selected_item()
        if not selected_item:
            messagebox.showerror("Error", "Please select a record to delete")
            return
        row = selected_item
        product_id = row[0]
        product_name = row[1]

        # Ask for confirmation
        confirm = messagebox.askyesno("Confirm Delete",
                                      f"Are you sure you want to delete the product '{product_name}' with ID '{product_id}'?")
        if not confirm:
            return

        try:
            product_dbase.deleteRecord(product_id)

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            # Retrieve the user's full name
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
            )
            fullname = cursor.fetchone()

            # Log the user activity
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    f"{fullname[0]} deleted product: {product_name} with ID: {product_id}",
                    username,
                ),
            )
            conn.commit()
            conn.close()

        except Exception as e:
            messagebox.showerror("Error", str(e))

        add_to_product_table()
        clear_product_entry_field()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()
        low_stock_vs_total_item_pie_chart()
        bar_chart()
        return

    def search_product(event):
        search_term = search_entry.get().lower()
        products = product_dbase.fetch_product_data()
        admin_product_tree.search_item(search_term, products)

    def fetch_purchase_order_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(PURCHASE_ORDER_ID) FROM PURCHASE_ORDER")
        last_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_id

    def generate_purchase_order_id(prefix="PO"):
        last_id = fetch_purchase_order_last_id()
        if last_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def on_product_double_click(event):
        values = admin_product_tree.selected_item()

        new_window = customtkinter.CTk()
        new_window.title("Product Details")
        new_window.resizable(False, False)

        def add_incoming_stock():
            incoming_stock_id = generate_purchase_order_id("PO")
            incoming_stock_product = values[1]
            incoming_stock_quantity = restock_quantity_entry.getvalue()
            incoming_stock_status = "To be Received"

            if not incoming_stock_quantity:
                messagebox.showerror("Error", "Please enter all fields")
                return

            purchase_order_dbase.insertRecord(
                incoming_stock_id,
                incoming_stock_product,
                incoming_stock_quantity,
                incoming_stock_status)

            # add_to_incoming_stock_table()
            fetch_tbr_quantity()
            update_total_quantity_in_hand_label()
            update_total_quantity_to_be_received_label()
            new_window.destroy()

        def fetch_tbr_quantity():
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT PURCHASE_ORDER_PRODUCT_QUANTITY FROM PURCHASE_ORDER WHERE PURCHASE_ORDER_PRODUCT = ? AND PURCHASE_ORDER_STATUS=?",
                (values[1], "To be Received"),
            )
            incoming_stock_quantity = cursor.fetchall()
            tbr = 0
            for x in incoming_stock_quantity:
                tbr += x[0]
            conn.commit()
            conn.close()
            return tbr

        def fetch_tbs_quantity():
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT SALE_ORDER_PRODUCT_QUANTITY FROM SALE_ORDER_PRODUCT WHERE SALE_ORDER_PRODUCT = ?",
                (values[1],),
            )
            outgoing_stock_quantity = cursor.fetchall()
            tbs = 0
            for x in outgoing_stock_quantity:
                tbs += x[0]
            conn.commit()
            conn.close()
            return tbs

        status_frame = customtkinter.CTkFrame(new_window, border_width=2)

        total_tbs = customtkinter.CTkFrame(
            master=status_frame, width=100, height=100, fg_color="transparent"
        )
        total_tbs_details_1 = customtkinter.CTkLabel(
            total_tbs,
            text=(str(fetch_tbs_quantity()) + " Qty"),
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )
        total_tbs_details_2 = customtkinter.CTkLabel(
            total_tbs,
            text="To be Shipped",
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )
        vertical_separator = customtkinter.CTkFrame(
            status_frame, width=2, height=80, fg_color="#CCCCCC"
        )
        total_tbr = customtkinter.CTkFrame(
            master=status_frame, width=100, height=100, fg_color="transparent"
        )
        total_tbr_details_1 = customtkinter.CTkLabel(
            total_tbr,
            text=(str(fetch_tbr_quantity()) + " Qty"),
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )
        total_tbr_details_2 = customtkinter.CTkLabel(
            total_tbr,
            text="To be Received",
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )
        status_frame.grid(row=0, column=1, sticky=W, padx=10, pady=10)

        status_details_title = Inv_titlelabel(status_frame, 0, 0, 10, "Status")
        total_tbs.grid(row=1, column=0, rowspan=5, padx=15, pady=15)
        total_tbs_details_1.pack(side=TOP)
        total_tbs_details_2.pack(side=BOTTOM)
        vertical_separator.grid(row=1, column=1, rowspan=5, pady=36)
        total_tbr.grid(row=1, column=3, rowspan=5, padx=15, pady=15)
        total_tbr_details_1.pack(side=TOP)
        total_tbr_details_2.pack(side=BOTTOM)

        product_details_frame = customtkinter.CTkFrame(new_window, border_width=2)
        product_details_frame.grid(row=0, column=0, sticky=W, padx=10, pady=10)
        product_details_title = Inv_titlelabel(product_details_frame, 0, 0, 10, "Product Details")

        product_id_label = Inv_Label(product_details_frame, 1, 0, 5, "Product ID:")
        product_id_label_details = Inv_Label(product_details_frame, 1, 1, 5, f"{values[0]}")
        product_name_label = Inv_Label(product_details_frame, 2, 0, 5, "Product Name:")
        product_name_label_details = Inv_Label(product_details_frame, 2, 1, 5, f"{values[1]}")
        product_quantity_label = Inv_Label(product_details_frame, 3, 0, 5, "Product Quantity:")
        product_quantity_label_details = Inv_Label(product_details_frame, 3, 1, 5, f"{values[2]}")
        product_description_label = Inv_Label(product_details_frame, 4, 0, 5, "Product Description:")
        product_description_label_details = Inv_Label(product_details_frame, 4, 1, 5, f"{values[3]}")

        restock_frame = customtkinter.CTkFrame(new_window, border_width=2)
        restock_frame.grid(row=1, column=0, padx=10, pady=10, columnspan=2)

        restock_title = Inv_titlelabel(restock_frame, 0, 0, 10, "Restock")
        restock_quantity_label = Inv_Label(restock_frame, 1, 0, 5, "Restock Quantity:")
        restock_quantity_entry = Inv_Entrybox(restock_frame, 1, 1, 5, "Quantity")
        restock_btn = Inv_Button(restock_frame, 2, 0, 5, "#007FFF", "Restock", add_incoming_stock)

        new_window.mainloop()

    def sort_treeview_column(tv, col, reverse):
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(reverse=reverse)

        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)

        tv.heading(col, command=lambda: sort_treeview_column(tv, col, not reverse))

    def low_stock_vs_total_item_pie_chart():
        low_stock_items = product_dbase.fetch_low_stock_item_data()
        total_items = product_dbase.fetch_total_items_data()
        piechart.display(total_items, low_stock_items)

    def bar_chart():
        barchart_data = product_dbase.fetch_bar_chart_data()
        barchart.display(barchart_data)

    def fetch_user_activities_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(USER_ACTIVITIES_ID) FROM USER_ACTIVITIES")
        last_user_activities_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_user_activities_id

    def generate_user_activities_id(prefix="UA"):
        last_customer_id = fetch_user_activities_last_id()
        if last_customer_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_customer_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def fetch_product_movement_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(PRODUCT_MOVEMENT_ID) FROM PRODUCT_MOVEMENT")
        last_product_movement_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_product_movement_id

    def generate_product_movement_id(prefix="PM"):
        product_movement_id = fetch_product_movement_last_id()
        if product_movement_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(product_movement_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def fetch_user_activities_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM USER_ACTIVITIES")
        data = cursor.fetchall()
        conn.commit()
        conn.close()
        return data

    def add_to_user_activities_table():
        user_activities = fetch_user_activities_data()
        user_activities_tree.delete(*user_activities_tree.get_children())
        for user_activity in user_activities:
            user_activities_tree.insert("", END, values=user_activity)

    def search_user_activities(event):
        search_term = search_user_activities_entry.get().lower()
        user_activities = fetch_user_activities_data()
        user_activities_tree.delete(*user_activities_tree.get_children())
        for user_activity in user_activities:
            if search_term in str(user_activity).lower():
                user_activities_tree.insert("", tk.END, values=user_activity)

    def search_product_movement(event):
        search_term = search_product_movement_entry.get().lower()
        product_movements = fetch_product_movement_data()
        product_movement_tree.delete(*product_movement_tree.get_children())
        for product_movement in product_movements:
            if search_term in str(product_movement).lower():
                product_movement_tree.insert("", tk.END, values=product_movement)

    def fetch_product_movement_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM PRODUCT_MOVEMENT")
        data = cursor.fetchall()
        conn.commit()
        conn.close()
        return data

    def add_to_product_movement_table():
        product_movements = fetch_product_movement_data()
        product_movement_tree.delete(*product_movement_tree.get_children())
        for product_movement in product_movements:
            product_movement_tree.insert("", END, values=product_movement)

    def export_to_excel(treeview):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Treeview Data"

        for col_index, col in enumerate(treeview["columns"], start=1):
            ws.cell(row=1, column=col_index, value=col)

        for row_index, row in enumerate(treeview.get_children(), start=2):
            for col_index, col in enumerate(treeview["columns"], start=1):
                ws.cell(row=row_index, column=col_index, value=treeview.item(row, "values")[col_index - 1])

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx"),
                                                            ("All files", "*.*")])
        if file_path:
            wb.save(file_path)
            print(f"Data exported to {file_path}")

    def export_to_pdf(treeview):
        pdf = FPDF(orientation='L')  # Set the orientation to Landscape
        pdf.add_page()
        pdf.set_font("Arial", size=8)

        # Calculate column widths based on content
        col_widths = []
        for col in treeview["columns"]:
            max_width = pdf.get_string_width(col) + 2  # Start with column header width
            for row in treeview.get_children():
                cell_text = str(treeview.item(row, "values")[treeview["columns"].index(col)])
                cell_width = pdf.get_string_width(cell_text) + 2
                if cell_width > max_width:
                    max_width = cell_width
            col_widths.append(max_width)

        # Add column headers
        for col, width in zip(treeview["columns"], col_widths):
            pdf.cell(width, 10, col, 1)

        pdf.ln()

        # Add rows
        for row in treeview.get_children():
            row_values = treeview.item(row, "values")
            for value, width in zip(row_values, col_widths):
                pdf.cell(width, 10, str(value), 1)
            pdf.ln()

        file_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                 filetypes=[("PDF files", "*.pdf"),
                                                            ("All files", "*.*")])
        if file_path:
            pdf.output(file_path)
            print(f"Data exported to {file_path}")

    def logout():
        confirmation = messagebox.askyesno('Are you sure?', 'Are you sure that you want to logout?')
        if confirmation:
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
            )
            fullname = cursor.fetchone()
            conn.commit()
            conn.close()

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID,USER_ACTIVITIES_DATE,USER_ACTIVITIES,USER) \
            VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "{} logged out".format(fullname[0]),
                    username,
                ),
            )
            conn.commit()
            conn.close()

            admin_dashboard_frame.destroy()
            login_page()

    admin_dashboard_frame = customtkinter.CTkFrame(
        master=app, width=1280, height=720, border_width=0
    )
    admin_dashboard_frame.place(x=0, y=0)

    info_frame = customtkinter.CTkFrame(
        master=admin_dashboard_frame, width=1280, height=720
    )

    info_frame.place(x=0, y=0)
    info_tab = customtkinter.CTkTabview(master=info_frame, width=1280, height=720)
    custom_font = ("SF Pro Display", 15)
    info_tab._segmented_button.configure(font=custom_font, fg_color='grey', text_color='white')
    info_tab.place(x=0, y=0)
    tab_0 = info_tab.add(" Homepage ")
    tab_1 = info_tab.add(" Register User ")
    tab_4 = info_tab.add(" Product ")
    tab_5 = info_tab.add(" User Activities ")
    tab_6 = info_tab.add(" Product Movement ")

    # -----------------------------

    user_table_title1_label = customtkinter.CTkLabel(master=tab_1, text_color="purple",
                                                        text='Double Click on the selected User below to Change Password',
                                                        font=customtkinter.CTkFont("Times", 15, weight='bold'))
    user_table_title1_label.place(x=480, y=25)

    product_table_title1_label = customtkinter.CTkLabel(master=tab_4, text_color="purple",
                                                        text='Double Click on the selected Product below to Restock the Quantity',
                                                        font=customtkinter.CTkFont("Times", 15, weight='bold'))
    product_table_title1_label.place(x=50, y=180)

    # --------------------------------

    welcome_label = customtkinter.CTkLabel(
        master=tab_0,
        text=f"Hello, {fetch_user_fullname()} ",
        font=customtkinter.CTkFont("SF Pro Display", 24, weight="bold"),
    )

    sales_activity_frame = customtkinter.CTkFrame(
        tab_0, corner_radius=10, border_width=2
    )
    sales_activity_title = Inv_titlelabel(sales_activity_frame, 0, 0, 10, "Sales Activity")

    sales_activity_vertical_separator_1 = customtkinter.CTkFrame(
        sales_activity_frame, width=2, height=50, fg_color="#CCCCCC"
    )
    sales_activity_vertical_separator_2 = customtkinter.CTkFrame(
        sales_activity_frame, width=2, height=50, fg_color="#CCCCCC"
    )
    inventory_summary_frame = customtkinter.CTkFrame(
        tab_0, corner_radius=10, border_width=2
    )
    sales_activity_vertical_separator_4 = customtkinter.CTkFrame(
        inventory_summary_frame, width=2, height=50, fg_color="#CCCCCC"
    )
    product_details_frame = customtkinter.CTkFrame(
        tab_0, corner_radius=10, border_width=2
    )
    product_details_title = Inv_titlelabel(product_details_frame,0,0,10,"Product Details")

    sales_activity_vertical_separator_6 = customtkinter.CTkFrame(
        product_details_frame, width=2, height=50, fg_color="#CCCCCC"
    )
    logout_btn = Inv_Button(tab_0, 0, 2, 10, "#ff666d", "Logout", logout)

    welcome_label.grid(row=0, column=0, sticky=W, padx=10, pady=10)
    sales_activity_frame.grid(row=1, column=0, padx=10, pady=10)
    inventory_summary_frame.grid(row=1, column=1, padx=10, pady=10)
    product_details_frame.grid(row=1, column=2, padx=10, pady=10)

    piechart = Inv_pie_chart(tab_0, 2, 0)
    barchart = Inv_bar_chart(tab_0, 2, 1)
    low_stock_vs_total_item_pie_chart()
    bar_chart()

    total_to_be_packed_label_1 = Inv_Label(sales_activity_frame, 1, 0, 20, str(fetch_to_be_packed_data()))
    total_to_be_packed_label_2 = Inv_Label(sales_activity_frame, 2, 0, 20, "To be Packed")
    sales_activity_vertical_separator_1.grid(row=1, column=1, rowspan=2, padx=5, pady=5)
    total_to_be_shipped_label_1 = Inv_Label(sales_activity_frame, 1, 2, 20, str(fetch_to_be_shipped_data()))
    total_to_be_shipped_label_2 = Inv_Label(sales_activity_frame, 2, 2, 20, "To be Shipped")
    sales_activity_vertical_separator_2.grid(row=1, column=3, rowspan=2, padx=5, pady=5)
    total_to_be_delivered_label_1 = Inv_Label(sales_activity_frame, 1, 4, 20, str(fetch_to_be_delivered_data()))
    total_to_be_delivered_label_2 = Inv_Label(sales_activity_frame, 2, 4, 20, "To be Delivered")

    inventory_summary_title = Inv_titlelabel(inventory_summary_frame, 0, 0, 10, "Inventory Summary")
    total_quantity_in_hand_1 = Inv_Label(inventory_summary_frame, 1, 0, 20,
                                         str(product_dbase.fetch_total_quantity_in_hand_data()))
    total_quantity_in_hand_2 = Inv_Label(inventory_summary_frame, 2, 0, 20, "Quantity In Hand")
    sales_activity_vertical_separator_4.grid(row=1, column=1, rowspan=2, padx=5, pady=5)
    total_quantity_to_be_received_1 = Inv_Label(inventory_summary_frame, 1, 2, 20,
                                                str(fetch_total_quantity_to_be_received_data()))
    total_quantity_to_be_received_2 = Inv_Label(inventory_summary_frame, 2, 2, 20, "Quantity To be Received")

    low_stock_items_1 = Inv_Label(product_details_frame, 1, 0, 20, str(product_dbase.fetch_low_stock_item_data()))
    low_stock_items_2 = Inv_Label(product_details_frame, 2, 0, 20, "Low Stock Items")
    sales_activity_vertical_separator_6.grid(row=1, column=1, rowspan=2, padx=5, pady=5)
    all_items_1 = Inv_Label(product_details_frame, 1, 2, 20, str(product_dbase.fetch_total_items_data()))
    all_items_2 = Inv_Label(product_details_frame, 2, 2, 20, "Total Items")

    tab_0.columnconfigure(1, weight=1)

    # --------------------------------------
    user_table_frame = customtkinter.CTkFrame(master=tab_1, width=780, height=645)
    user_table_frame.place(x=450, y=50)

    user_tree = ttk.Treeview(master=user_table_frame, height=25)

    user_verscrlbar = customtkinter.CTkScrollbar(
        master=user_table_frame, orientation="vertical", command=user_tree.yview
    )
    user_verscrlbar.pack(side="right", fill="y")

    user_tree.configure(yscrollcommand=user_verscrlbar.set)
    user_tree["columns"] = (
        "FULLNAME",
        "USERNAME",
        "ACCESSLEVEL",
    )

    user_tree.column("#0", width=0, stretch=tk.NO)
    user_tree.column("FULLNAME", anchor=tk.CENTER, width=260)
    user_tree.column("USERNAME", anchor=tk.CENTER, width=260)
    user_tree.column("ACCESSLEVEL", anchor=tk.CENTER, width=260)

    user_tree.heading("FULLNAME", text="Full Name")
    user_tree.heading("USERNAME", text="Username")
    user_tree.heading("ACCESSLEVEL", text="Access Level")

    user_tree.pack(side="right", fill="both")

    user_tree.bind("<Double-1>", on_user_double_click)

    reg_user_name = customtkinter.CTkEntry(
        master=tab_1,
        placeholder_text="Full Name",
        width=300,
        height=30,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )
    app.update()
    reg_user_name.focus_set()

    reg_user = customtkinter.CTkEntry(
        master=tab_1,
        placeholder_text="Username",
        width=300,
        height=30,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )

    reg_password = customtkinter.CTkEntry(
        master=tab_1,
        placeholder_text="Password",
        width=300,
        height=30,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )
    reg_password.configure(show="*")

    reg_confirm_password = customtkinter.CTkEntry(
        master=tab_1,
        placeholder_text="Confirm Password",
        width=300,
        height=30,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )
    reg_confirm_password.configure(show="*")

    # -----------------------------------------
    def show_hide_password():
        if log_show_hide.get() == "on":
            reg_password.configure(show="")
            reg_confirm_password.configure(show="")
        else:
            reg_password.configure(show="*")
            reg_confirm_password.configure(show="*")

    log_show_hide = customtkinter.StringVar(value="off")
    log_show_hide_password_switch = customtkinter.CTkSwitch(
        master=tab_1,
        text="Show Password",
        command=show_hide_password,
        variable=log_show_hide,
        onvalue="on",
        offvalue="off",
        font=customtkinter.CTkFont("SF Pro Display"),
    )
    log_show_hide_password_switch.place(x=280, y=265)

    #-----------------------------------------

    register_new_user_label = Inv_main_titlelabel(tab_1, 0, 0, 30, 10, "Register New User")
    reg_user_name.grid(row=1, column=0, padx=30, pady=10, sticky="w")
    reg_user.grid(row=2, column=0, padx=30, pady=10, sticky="w")
    reg_password.grid(row=3, column=0, padx=30, pady=10, sticky="w")
    reg_confirm_password.grid(row=4, column=0, padx=30, pady=10, sticky="w")

    reg_accesslevel = Inv_ComboBox(tab_1, 5, 0, 30, ["Admin", "Supervisor", "Warehouse Worker"])
    reg_btn = Inv_Button(tab_1, 6, 0, 30, 'orange', "Register", register)
    dlt_user_btn = Inv_Button(tab_1, 7, 0, 30, 'orange', "Delete", delete_user_record)

    # --------------------------------------------------------------------------------

    product_table_frame = customtkinter.CTkFrame(master=tab_4, width=1280, height=515)
    product_table_frame.place(x=0, y=205)
    admin_product_tree = Inv_product_tree_display(product_table_frame, on_product_double_click, display_product_record)

    search_entry = customtkinter.CTkEntry(tab_4, placeholder_text="Search", width=1050, )
    search_entry.grid(row=1, column=0, padx=10, pady=10)
    search_entry.bind("<KeyRelease>", search_product)

    product_menu_frame = customtkinter.CTkFrame(
        master=tab_4,
        border_width=2
    )
    product_menu_frame.grid(row=0, column=0)

    insert_product_data_label = Inv_main_titlelabel(product_menu_frame, 0, 0, 5, 5, "Insert New Product")
    product_id_entry_label = Inv_Label(product_menu_frame, 1, 0, 5, "Product ID:")
    product_id_entry = Inv_Entrybox(product_menu_frame, 1, 1, 5, "ID")
    product_name_entry_label = Inv_Label(product_menu_frame, 1, 2, 5, "Product Name:")
    product_name_entry = Inv_Entrybox(product_menu_frame, 1, 3, 5, "Item")
    product_quantity_entry_label = Inv_Label(product_menu_frame, 1, 4, 5, "Product Quantity:")
    product_quantity_entry = Inv_Entrybox(product_menu_frame, 1, 5, 5, "Quantity")
    product_description_entry_label = Inv_Label(product_menu_frame, 2, 0, 5, "Product Description:")
    product_description_entry = Inv_Entrybox(product_menu_frame, 2, 1, 5, "Description")
    product_min_stock_entry_label = Inv_Label(product_menu_frame, 2, 2, 5, "Product Min. Stock:")
    product_min_stock_entry = Inv_Entrybox(product_menu_frame, 2, 3, 5, "Min. Stock")

    addproduct_btn = Inv_Button(product_menu_frame, 3, 0, 20, "#007FFF", "Add Product", add_new_product_details)
    editproduct_btn = Inv_Button(product_menu_frame, 3, 2, 20, "#ADD8E6", "Edit Product", edit_product_details)
    deleteproduct_btn = Inv_Button(product_menu_frame, 3, 4, 20, "#ff666d", "Delete Product", delete_product_record)

    tab_4.columnconfigure(0, weight=1)

    user_activities_table_frame = customtkinter.CTkFrame(master=tab_5, width=1280, height=515)
    user_activities_table_frame.place(x=0, y=205)

    user_activities_tree = ttk.Treeview(master=user_activities_table_frame, height=18)

    product_verscrlbar = customtkinter.CTkScrollbar(
        master=user_activities_table_frame, orientation="vertical", command=user_activities_tree.yview
    )
    product_verscrlbar.pack(side="right", fill="y")

    user_activities_tree.configure(yscrollcommand=product_verscrlbar.set)
    user_activities_tree["columns"] = (
        "USER ACTIVITIES ID",
        "USER ACTIVITIES DATE",
        "USER ACTIVITIES",
        "USER ACTIVITIES USERNAME"
    )

    user_activities_tree.column("#0", width=0, stretch=tk.NO)
    user_activities_tree.column("USER ACTIVITIES ID", anchor=tk.CENTER, width=313)
    user_activities_tree.column("USER ACTIVITIES DATE", anchor=tk.CENTER, width=313)
    user_activities_tree.column("USER ACTIVITIES", anchor=tk.CENTER, width=313)
    user_activities_tree.column("USER ACTIVITIES USERNAME", anchor=tk.CENTER, width=313)

    user_activities_tree.heading("USER ACTIVITIES ID", text="ID")
    user_activities_tree.heading("USER ACTIVITIES DATE", text="Date")
    user_activities_tree.heading("USER ACTIVITIES", text="Description")
    user_activities_tree.heading("USER ACTIVITIES USERNAME", text="User")

    user_activities_tree.pack(side="bottom", fill="both")

    user_activities_menu_frame = customtkinter.CTkFrame(
        master=tab_5,
        border_width=2
    )

    user_activities_menu_label = customtkinter.CTkLabel(
        master=user_activities_menu_frame,
        text="Export to Excel Sheet",
        font=customtkinter.CTkFont("SF Pro Display", weight="bold", size=20),
    )

    export_user_activities_btn = customtkinter.CTkButton(
        master=user_activities_menu_frame,
        text="Export Excel",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=lambda: export_to_excel(user_activities_tree),
        compound="top",
        corner_radius=200,
        fg_color="#ADD8E6",
        text_color="black",
    )

    export_user_activities_pdf_btn = customtkinter.CTkButton(
        master=user_activities_menu_frame,
        text="Export PDF",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=lambda: export_to_pdf(user_activities_tree),
        compound="top",
        corner_radius=200,
        fg_color="#ADD8E6",
        text_color="black",
    )

    search_user_activities_entry = customtkinter.CTkEntry(tab_5, placeholder_text="Search", width=1050, )
    search_user_activities_entry.grid(row=1, column=0, padx=10, pady=10)
    search_user_activities_entry.bind("<KeyRelease>", search_user_activities)

    user_activities_menu_frame.grid(row=0, column=0)
    user_activities_menu_label.grid(row=0, column=0, sticky="w", padx=5, pady=5, columnspan=2)
    export_user_activities_btn.grid(row=1, column=0, padx=20, pady=5, columnspan=2)
    export_user_activities_pdf_btn.grid(row=2, column=0, padx=20, pady=5, columnspan=2)

    tab_5.columnconfigure(0, weight=1)

    product_movement_table_frame = customtkinter.CTkFrame(master=tab_6, width=1280, height=515)
    product_movement_table_frame.place(x=0, y=205)

    product_movement_tree = ttk.Treeview(master=product_movement_table_frame, height=18)

    product_verscrlbar = customtkinter.CTkScrollbar(
        master=product_movement_table_frame, orientation="vertical", command=product_movement_tree.yview
    )
    product_verscrlbar.pack(side="right", fill="y")

    product_movement_tree.configure(yscrollcommand=product_verscrlbar.set)
    product_movement_tree["columns"] = (
        "PRODUCT MOVEMENT ID",
        "PRODUCT MOVEMENT DATE",
        "PRODUCT MOVEMENT",
        "USER"
    )

    product_movement_tree.column("#0", width=0, stretch=tk.NO)
    product_movement_tree.column("PRODUCT MOVEMENT ID", anchor=tk.CENTER, width=313)
    product_movement_tree.column("PRODUCT MOVEMENT DATE", anchor=tk.CENTER, width=313)
    product_movement_tree.column("PRODUCT MOVEMENT", anchor=tk.CENTER, width=313)
    product_movement_tree.column("USER", anchor=tk.CENTER, width=313)

    product_movement_tree.heading("PRODUCT MOVEMENT ID", text="ID")
    product_movement_tree.heading("PRODUCT MOVEMENT DATE", text="Date")
    product_movement_tree.heading("PRODUCT MOVEMENT", text="Description")
    product_movement_tree.heading("USER", text="User")

    product_movement_tree.pack(side="bottom", fill="both")

    product_movement_menu_frame = customtkinter.CTkFrame(
        master=tab_6,
        border_width=2
    )

    product_movement_menu_label = customtkinter.CTkLabel(
        master=product_movement_menu_frame,
        text="Export to Excel Sheet",
        font=customtkinter.CTkFont("SF Pro Display", weight="bold", size=20),
    )

    export_product_movement_btn = customtkinter.CTkButton(
        master=product_movement_menu_frame,
        text="Export Excel",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=lambda: export_to_excel(product_movement_tree),
        compound="top",
        corner_radius=200,
        fg_color="#ADD8E6",
        text_color="black",
    )

    export_product_movement_pdf_btn = customtkinter.CTkButton(
        master=product_movement_menu_frame,
        text="Export PDF",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=lambda: export_to_pdf(product_movement_tree),
        compound="top",
        corner_radius=200,
        fg_color="#ADD8E6",
        text_color="black",
    )

    search_product_movement_entry = customtkinter.CTkEntry(tab_6, placeholder_text="Search", width=1050, )
    search_product_movement_entry.grid(row=1, column=0, padx=10, pady=10)
    search_product_movement_entry.bind("<KeyRelease>", search_product_movement)

    product_movement_menu_frame.grid(row=0, column=0)
    product_movement_menu_label.grid(row=0, column=0, sticky="w", padx=5, pady=5, columnspan=2)
    export_product_movement_btn.grid(row=1, column=0, padx=20, pady=5, columnspan=2)
    export_product_movement_pdf_btn.grid(row=2, column=0, padx=20, pady=5, columnspan=2)

    tab_6.columnconfigure(0, weight=1)

    add_to_user_table()
    add_to_product_table()
    add_to_user_activities_table()
    add_to_product_movement_table()


# ====================================== Supervisor Dashboard ================================================
def supervisor_dashboard(username):
    product_dbase = Inv_Product_Database()  # declare the product class object.
    purchase_order_dbase = Inv_Purchase_Order_dbase()  # declare the purchase_order class object.

    def fetch_user_fullname():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,))
        user = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return user

    def fetch_to_be_packed_data():
        total_to_be_packed = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(SALE_ORDER_STATUS) FROM SALE_ORDER WHERE SALE_ORDER_STATUS = ?", ("To be Packed",))
        to_be_packed_data = cursor.fetchone()[0]
        total_to_be_packed += to_be_packed_data
        conn.commit()
        conn.close()
        return total_to_be_packed

    def update_to_be_packed_label():
        new_count = fetch_to_be_packed_data()
        total_to_be_packed_label_1.updateInfo(str(new_count))

    def fetch_to_be_shipped_data():
        total_to_be_shipped = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(SALE_ORDER_STATUS) FROM SALE_ORDER WHERE SALE_ORDER_STATUS = ?",
                       ("To be Shipped",))
        to_be_shipped_data = cursor.fetchone()[0]
        total_to_be_shipped += to_be_shipped_data
        conn.commit()
        conn.close()
        return total_to_be_shipped

    def update_to_be_shipped_label():
        new_count = fetch_to_be_shipped_data()
        total_to_be_shipped_label_1.updateInfo(str(new_count))

    def fetch_to_be_delivered_data():
        total_to_be_delivered = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(SALE_ORDER_STATUS) FROM SALE_ORDER WHERE SALE_ORDER_STATUS = ?",
                       ("To be Delivered",))
        to_be_delivered_data = cursor.fetchone()[0]
        total_to_be_delivered += to_be_delivered_data
        conn.commit()
        conn.close()
        return total_to_be_delivered

    def update_to_be_delivered_label():
        new_count = fetch_to_be_delivered_data()
        total_to_be_delivered_label_1.updateInfo(str(new_count))

    def fetch_total_quantity_in_hand_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT SUM(PRODUCT_QUANTITY) FROM PRODUCT")
        total_quantity_in_hand = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return total_quantity_in_hand

    def update_total_quantity_in_hand_label():
        new_count = fetch_total_quantity_in_hand_data()
        total_quantity_in_hand_1.updateInfo(str(new_count))

    def fetch_total_quantity_to_be_received_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT SUM(PURCHASE_ORDER_PRODUCT_QUANTITY) FROM PURCHASE_ORDER WHERE PURCHASE_ORDER_STATUS = ?",
            ("To be Received",))
        total_quantity_to_be_received = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return total_quantity_to_be_received

    def update_total_quantity_to_be_received_label():
        new_count = fetch_total_quantity_to_be_received_data()
        total_quantity_to_be_received_1.updateInfo(str(new_count))

    def fetch_low_stock_item_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(PRODUCT_ID) FROM PRODUCT WHERE PRODUCT_QUANTITY < PRODUCT_MIN_STOCK")
        low_stock_items = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return low_stock_items

    def update_low_stock_item_label():
        new_count = fetch_low_stock_item_data()
        low_stock_items_1.updateInfo(str(new_count))

    def fetch_total_items_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(PRODUCT_ID) FROM PRODUCT")
        total_items = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return total_items

    def update_total_items_label():
        new_count = fetch_total_items_data()
        all_items_1.updateInfo(str(new_count))

    def fetch_customer_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM CUSTOMER")
        customer = cursor.fetchall()
        conn.commit()
        conn.close()
        return customer

    def clear_customer_entry_field():
        customer_name_entry.clearField()
        customer_email_entry.clearField()
        customer_contact_entry.clearField()

    def display_customer_record(event):
        selected_item = customer_tree.focus()
        if selected_item:
            clear_customer_entry_field()
            row = customer_tree.item(selected_item)["values"]
            customer_name_entry.insertField(row[1])
            customer_email_entry.insertField(row[2])
            customer_contact_entry.insertField(row[3])

    def fetch_customer_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(CUSTOMER_ID) FROM CUSTOMER")
        last_customer_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_customer_id

    def generate_customer_id(prefix="CUST"):
        last_customer_id = fetch_customer_last_id()
        if last_customer_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_customer_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def add_to_customer_table():
        customers = fetch_customer_data()
        customer_tree.delete(*customer_tree.get_children())
        for customer in customers:
            customer_tree.insert("", END, values=customer)

    def search_customer(event):
        search_term = search_customer_entry.get().lower()
        customers = fetch_customer_data()
        customer_tree.delete(*customer_tree.get_children())
        for customer in customers:
            if search_term in str(customer).lower():
                customer_tree.insert("", tk.END, values=customer)

    def check_existing_customer(customer_check):
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT CUSTOMER_NAME FROM CUSTOMER WHERE CUSTOMER_NAME = ?",
            (customer_check,),
        )
        existing_customer = cursor.fetchone()
        conn.commit()
        conn.close()
        return existing_customer

    def is_valid_email(email):
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email)

    def is_valid_contact_number(contact_number):
        return contact_number.isdigit()

    def add_new_customer_details():
        customer_id = generate_customer_id("CUST")
        customer_name = customer_name_entry.getvalue()
        customer_email = customer_email_entry.getvalue()
        customer_contactno = customer_contact_entry.getvalue()

        if not (customer_name and customer_email and customer_contactno):
            messagebox.showerror("Error", "Please enter all fields")
            return

        if not is_valid_email(customer_email):
            messagebox.showerror("Error", "Please enter a valid email address")
            return

        if not is_valid_contact_number(customer_contactno):
            messagebox.showerror("Error", "Contact number must be only digits")
            return

        if not check_existing_customer(customer_name):
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()

            try:
                cursor.execute(
                    "INSERT INTO CUSTOMER (CUSTOMER_ID,CUSTOMER_NAME,CUSTOMER_EMAIL,CUSTOMER_TEL) \
                    VALUES (?,?,?,?)",
                    (customer_id, customer_name, customer_email, customer_contactno),
                )
                conn.commit()
                messagebox.showinfo("Success", "Data has been inserted")

                # Retrieve the user's full name
                cursor.execute(
                    "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
                )
                fullname = cursor.fetchone()

                # Log the user activity
                cursor.execute(
                    "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_user_activities_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "{} added new customer: {} with ID: {}".format(fullname[0], customer_name, customer_id),
                        username,
                    ),
                )
                conn.commit()

            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()

        else:
            messagebox.showerror(
                "Warning", "Duplicate customer! Please enter new customer."
            )

        add_to_customer_table()
        clear_customer_entry_field()
        return

    def edit_customer_details():
        selected_customer_details = customer_tree.focus()
        if not selected_customer_details:
            messagebox.showerror("Error", "Please select a record to edit")
            return

        row = customer_tree.item(selected_customer_details)["values"]
        new_customer_name = customer_name_entry.getvalue()
        new_customer_email = customer_email_entry.getvalue()
        new_customer_contactno = customer_contact_entry.getvalue()

        if not (new_customer_name and new_customer_email and new_customer_contactno):
            messagebox.showerror("Error", "Please enter all fields")
            return

        if not is_valid_email(new_customer_email):
            messagebox.showerror("Error", "Please enter a valid email address")
            return

        if not is_valid_contact_number(new_customer_contactno):
            messagebox.showerror("Error", "Contact number must be only digits")
            return

        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()

        old_customer_name = row[1]
        old_customer_email = row[2]
        old_customer_contactno = row[3]

        # Check if the new customer name is unique
        cursor.execute(
            "SELECT COUNT(*) FROM CUSTOMER WHERE CUSTOMER_NAME = ? AND CUSTOMER_NAME != ?",
            (new_customer_name, old_customer_name)
        )
        existing_customer_count = cursor.fetchone()[0]

        if existing_customer_count > 0:
            messagebox.showerror("Error", "Customer name must be unique")
            conn.close()
            return

        try:
            cursor.execute(
                "UPDATE CUSTOMER SET CUSTOMER_NAME = ?, CUSTOMER_EMAIL = ?, CUSTOMER_TEL = ? WHERE CUSTOMER_NAME = ?",
                (new_customer_name, new_customer_email, new_customer_contactno, old_customer_name),
            )
            conn.commit()

            # Retrieve the user's full name
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?",
                (username,)
            )
            fullname = cursor.fetchone()

            # Determine what was edited
            changes = []
            if old_customer_name != new_customer_name:
                changes.append(f"Name from '{old_customer_name}' to '{new_customer_name}'")
            if old_customer_email != new_customer_email:
                changes.append(f"Email from '{old_customer_email}' to '{new_customer_email}'")
            if old_customer_contactno != new_customer_contactno:
                changes.append(f"Contact number from '{old_customer_contactno}' to '{new_customer_contactno}'")

            changes_str = "; ".join(changes)

            # Log the user activity
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    f"{fullname[0]} edited customer {old_customer_name}: {changes_str}",
                    username,
                ),
            )
            conn.commit()

            messagebox.showinfo("Success", "Record successfully edited")

        except Exception as e:
            messagebox.showerror("Error", str(e))

        finally:
            conn.close()

        add_to_customer_table()
        clear_customer_entry_field()
        return

    def delete_customer_record():
        selected_item = customer_tree.focus()
        if not selected_item:
            messagebox.showerror("Error", "Please select a record to delete")
            return

        row = customer_tree.item(selected_item)["values"]
        customer_name = row[1]
        customer_id = row[0]

        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()

        try:
            cursor.execute("DELETE FROM CUSTOMER WHERE CUSTOMER_NAME= ?", (customer_name,))
            conn.commit()

            # Retrieve the user's full name
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
            )
            fullname = cursor.fetchone()

            # Log the user activity
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "{} deleted customer: {} with ID: {}".format(fullname[0], customer_name, customer_id),
                    username,
                ),
            )
            conn.commit()

            messagebox.showinfo("Success", "Record successfully deleted")

        except Exception as e:
            messagebox.showerror("Error", str(e))

        finally:
            conn.close()

        add_to_customer_table()
        clear_customer_entry_field()
        return

    def fetch_product_to_list():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT PRODUCT_ID FROM PRODUCT")
        products = cursor.fetchall()
        product = []
        for x in products:
            product.append(x[0])
        conn.commit()
        conn.close()
        return product

    def on_customer_double_click(event):

        selected_item = customer_tree.selection()[0]
        values = customer_tree.item(selected_item, "values")

        def add_new_sale_order():
            sale_order_id = generate_sale_order_id("SO")
            sale_order_date = date.today().strftime("%m/%d/%Y")
            customer_name = values[1]

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO SALE_ORDER (SALE_ORDER_ID,SALE_ORDER_DATE,SALE_ORDER_CUSTOMER,SALE_ORDER_STATUS) \
            VALUES (?,?,?,?)",
                (sale_order_id, sale_order_date, customer_name, "To be Packed"),
            )
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Data has been inserted")
            add_to_sale_order_table()
            update_to_be_packed_label()
            update_to_be_shipped_label()
            update_to_be_delivered_label()
            update_total_quantity_in_hand_label()
            update_total_quantity_to_be_received_label()
            update_low_stock_item_label()
            update_total_items_label()
            return

        sales_window = customtkinter.CTk()
        sales_window.title("Product Details")
        sales_window.resizable(False, False)

        customer_id_label = Inv_Label(sales_window, 0, 0, (10, 5), "Customer ID:")
        customer_id_label_details = Inv_Label(sales_window, 0, 1, 5, f"{values[0]}")
        customer_name_label = Inv_Label(sales_window, 1, 0, (10, 5), "Customer Name:")
        customer_name_label_details = Inv_Label(sales_window, 1, 1, 5, f"{values[1]}")
        customer_email_label = Inv_Label(sales_window, 2, 0, (10, 5), "Customer Email:")
        customer_email_label_details = Inv_Label(sales_window, 2, 1, 5, f"{values[2]}")
        customer_contact_label = Inv_Label(sales_window, 3, 0, (10, 5), "Customer Contact No.:")
        customer_contact_label_details = Inv_Label(sales_window, 3, 1, 5, f"{values[3]}")

        sales_btn = Inv_Button(sales_window, 6, 0, 0, "#007FFF", "Create Sales Order", add_new_sale_order)

        sales_window.columnconfigure(0, weight=1)
        sales_window.columnconfigure(1, weight=1)
        sales_window.columnconfigure(2, weight=1)
        sales_window.columnconfigure(3, weight=1)
        sales_window.columnconfigure(4, weight=1)
        sales_window.columnconfigure(5, weight=1)
        sales_window.columnconfigure(6, weight=1)

        sales_window.mainloop()

# ---------------------------------------------------------------------

    def add_to_product_table():
        products = product_dbase.fetch_product_data()
        admin_product_tree.add_to_table(products)

    def clear_product_entry_field():
        product_id_entry.clearField()
        product_name_entry.clearField()
        product_quantity_entry.clearField()
        product_description_entry.clearField()
        product_min_stock_entry.clearField()

    def display_product_record(event):
        selected_item = admin_product_tree.selected_item()
        if selected_item:
            clear_product_entry_field()
            row = selected_item
            product_id_entry.insertField([row[0]])
            product_name_entry.insertField(row[1])
            product_quantity_entry.insertField(row[2])
            product_description_entry.insertField(row[3])
            product_min_stock_entry.insertField([row[4]])

    def fetch_supplier_to_list():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT SUPPLIER_NAME FROM SUPPLIER")
        suppliers = cursor.fetchall()
        supplier = []
        for x in suppliers:
            supplier.append(x[0])
        conn.commit()
        conn.close()
        return supplier

    def update_supplier_to_list():
        workers = fetch_supplier_to_list()
        preferred_supplier_entry.configure(values=workers)

    def add_new_product_details():
        product_id = product_id_entry.getvalue()
        product_name = product_name_entry.getvalue()
        product_quantity = product_quantity_entry.getvalue()
        product_description = product_description_entry.getvalue()
        product_min_stock = product_min_stock_entry.getvalue()

        # Validate that product_id is not empty
        if not product_id:
            messagebox.showerror("Error", "Please enter product id")
            return

        # Validate that all fields are filled
        if not (product_name and product_quantity and product_description):
            messagebox.showerror("Error", "Please enter all fields")
            return

        # Validate that product_quantity is an integer
        if not product_quantity.isdigit() or int(product_quantity) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for product quantity")
            return

        # Validate that product_min_stock is an integer
        if not product_min_stock.isdigit() or int(product_min_stock) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for minimum stock")
            return

        product_quantity = int(product_quantity)
        product_min_stock = int(product_min_stock)

        # Check if product already exists
        if not product_dbase.check_existing_product(product_id):
            try:
                product_dbase.insertRecord(
                    product_id,
                    product_name,
                    product_quantity,
                    product_description,
                    product_min_stock,
                )

                conn = sqlite3.connect("Inventory Management System.db")
                cursor = conn.cursor()

                # Retrieve the user's full name
                cursor.execute(
                    "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
                )
                fullname = cursor.fetchone()

                # Log the user activity
                cursor.execute(
                    "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_user_activities_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "{} added new product: {} with ID: {}".format(fullname[0], product_name, product_id),
                        username,
                    ),
                )
                conn.commit()

                # Log the product movement
                cursor.execute(
                    "INSERT INTO PRODUCT_MOVEMENT (PRODUCT_MOVEMENT_ID, PRODUCT_MOVEMENT_DATE, PRODUCT_MOVEMENT, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_product_movement_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Added product: {} with ID: {} by {} to warehouse".format(product_name, product_id,
                                                                                  fullname[0]),
                        username,
                    ),
                )
                conn.commit()
                conn.close()

            except Exception as e:
                messagebox.showerror("Error", str(e))

        else:
            messagebox.showerror("Warning", "Duplicate product! Please enter new product.")

        add_to_product_table()
        clear_product_entry_field()
        fetch_product_to_list()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()
        low_stock_vs_total_item_pie_chart()
        bar_chart()
        return

    def edit_product_details():
        selected_product_details = admin_product_tree.selected_item()
        if not selected_product_details:
            messagebox.showerror("Error", "Please select a record to edit")
            return

        row = selected_product_details
        new_product_id = product_id_entry.getvalue()
        new_product_name = product_name_entry.getvalue()
        new_product_quantity = product_quantity_entry.getvalue()
        new_product_description = product_description_entry.getvalue()
        new_product_min_stock = product_min_stock_entry.getvalue()

        if not new_product_name:
            messagebox.showerror("Error", "Please enter all fields")
            return

        if not new_product_quantity.isdigit() or int(new_product_quantity) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for product quantity")
            return

        # Validate that product_min_stock is an integer
        if not new_product_min_stock.isdigit() or int(new_product_min_stock) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for minimum stock")
            return

        old_product_id = row[0]
        old_product_name = row[1]
        old_product_quantity = row[2]
        old_product_description = row[3]
        old_product_min_stock = row[4]
        # old_preferred_supplier = row[5]

        try:
            product_dbase.updateRecord(
                new_product_id,
                new_product_name,
                new_product_quantity,
                new_product_description,
                new_product_min_stock,
                new_product_id)

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            # Retrieve the user's full name
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
            )
            fullname = cursor.fetchone()

            # Determine what was edited
            changes = []
            if old_product_id != new_product_id:
                changes.append(f"ID from '{old_product_id}' to '{new_product_id}'")
            if old_product_name != new_product_name:
                changes.append(f"Name from '{old_product_name}' to '{new_product_name}'")
            if old_product_quantity != new_product_quantity:
                changes.append(f"Quantity from '{old_product_quantity}' to '{new_product_quantity}'")
            if old_product_description != new_product_description:
                changes.append(f"Description from '{old_product_description}' to '{new_product_description}'")
            if old_product_min_stock != new_product_min_stock:
                changes.append(f"Min stock from '{old_product_min_stock}' to '{new_product_min_stock}'")
            # if old_preferred_supplier != new_preferred_supplier:
            #    changes.append(f"Preferred supplier from '{old_preferred_supplier}' to '{new_preferred_supplier}'")

            changes_str = "; ".join(changes)

            # Log the user activity
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    f"{fullname[0]} edited product {old_product_name}: {changes_str}",
                    username,
                ),
            )
            conn.commit()
            conn.close()

        except Exception as e:
            messagebox.showerror("Error", str(e))

        add_to_product_table()
        clear_product_entry_field()
        fetch_product_to_list()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()
        low_stock_vs_total_item_pie_chart()
        bar_chart()
        return

    def delete_product_record():
        selected_item = admin_product_tree.selected_item()
        if not selected_item:
            messagebox.showerror("Error", "Please select a record to delete")
            return
        row = selected_item
        product_id = row[0]
        product_name = row[1]

        # Ask for confirmation
        confirm = messagebox.askyesno("Confirm Delete",
                                      f"Are you sure you want to delete the product '{product_name}' with ID '{product_id}'?")
        if not confirm:
            return

        try:
            product_dbase.deleteRecord(product_id)

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            # Retrieve the user's full name
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
            )
            fullname = cursor.fetchone()

            # Log the user activity
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    f"{fullname[0]} deleted product: {product_name} with ID: {product_id}",
                    username,
                ),
            )
            conn.commit()
            conn.close()

        except Exception as e:
            messagebox.showerror("Error", str(e))

        add_to_product_table()
        clear_product_entry_field()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()
        low_stock_vs_total_item_pie_chart()
        bar_chart()
        return

    def search_product(event):
        search_term = search_entry.get().lower()
        products = product_dbase.fetch_product_data()
        admin_product_tree.search_item(search_term, products)

    # -------------------------------------------------
    def on_product_double_click(event):
        values = admin_product_tree.selected_item()

        new_window = customtkinter.CTk()
        new_window.title("Product Details")
        new_window.resizable(False, False)

        def add_incoming_stock():
            incoming_stock_id = generate_purchase_order_id("PO")
            incoming_stock_product = values[1]
            incoming_stock_quantity = restock_quantity_entry.getvalue()
            incoming_stock_status = "To be Received"

            if not incoming_stock_quantity:
                messagebox.showerror("Error", "Please enter all fields")
                return

            try:
                purchase_order_dbase.insertRecord(
                    incoming_stock_id,
                    incoming_stock_product,
                    incoming_stock_quantity,
                    incoming_stock_status)

                conn = sqlite3.connect("Inventory Management System.db")
                cursor = conn.cursor()
                # Retrieve the user's full name
                cursor.execute(
                    "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
                )
                fullname = cursor.fetchone()

                # Log the product movement
                cursor.execute(
                    "INSERT INTO PRODUCT_MOVEMENT (PRODUCT_MOVEMENT_ID, PRODUCT_MOVEMENT_DATE, PRODUCT_MOVEMENT, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_product_movement_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "{} added incoming stock for product: {} with quantity: {}".format(fullname[0],
                                                                                           incoming_stock_product,
                                                                                           incoming_stock_quantity),
                        username,
                    ),
                )
                conn.commit()
                conn.close()

            except Exception as e:
                messagebox.showerror("Error", str(e))

            add_to_incoming_stock_table()
            fetch_tbr_quantity()
            update_total_quantity_in_hand_label()
            update_total_quantity_to_be_received_label()
            new_window.destroy()

        def fetch_tbr_quantity():
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT PURCHASE_ORDER_PRODUCT_QUANTITY FROM PURCHASE_ORDER WHERE PURCHASE_ORDER_PRODUCT = ? AND PURCHASE_ORDER_STATUS=?",
                (values[1], "To be Received"),
            )
            incoming_stock_quantity = cursor.fetchall()
            tbr = 0
            for x in incoming_stock_quantity:
                tbr += x[0]
            conn.commit()
            conn.close()
            return tbr

        def fetch_tbs_quantity():
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT SALE_ORDER_PRODUCT_QUANTITY FROM SALE_ORDER_PRODUCT WHERE SALE_ORDER_PRODUCT = ?",
                (values[1],),
            )
            outgoing_stock_quantity = cursor.fetchall()
            tbs = 0
            for x in outgoing_stock_quantity:
                tbs += x[0]
            conn.commit()
            conn.close()
            return tbs

        # -------------------------------------------
        status_frame = customtkinter.CTkFrame(new_window, border_width=2)
        status_details_title = Inv_titlelabel(status_frame,0,0,10,"Status")

        total_tbs = customtkinter.CTkFrame(
            master=status_frame, width=100, height=100, fg_color="transparent"
        )
        total_tbs_details_1 = customtkinter.CTkLabel(
            total_tbs,
            text=(str(fetch_tbs_quantity()) + " Qty"),
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )
        total_tbs_details_2 = customtkinter.CTkLabel(
            total_tbs,
            text="To be Shipped",
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )
        vertical_separator = customtkinter.CTkFrame(
            status_frame, width=2, height=80, fg_color="#CCCCCC"
        )
        total_tbr = customtkinter.CTkFrame(
            master=status_frame, width=100, height=100, fg_color="transparent"
        )
        total_tbr_details_1 = customtkinter.CTkLabel(
            total_tbr,
            text=(str(fetch_tbr_quantity()) + " Qty"),
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )
        total_tbr_details_2 = customtkinter.CTkLabel(
            total_tbr,
            text="To be Received",
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )
        status_frame.grid(row=0, column=1, sticky=W, padx=10, pady=10)
        total_tbs.grid(row=1, column=0, rowspan=5, padx=15, pady=15)
        total_tbs_details_1.pack(side=TOP)
        total_tbs_details_2.pack(side=BOTTOM)
        vertical_separator.grid(row=1, column=1, rowspan=5, pady=36)
        total_tbr.grid(row=1, column=3, rowspan=5, padx=15, pady=15)
        total_tbr_details_1.pack(side=TOP)
        total_tbr_details_2.pack(side=BOTTOM)

        product_details_frame = customtkinter.CTkFrame(new_window, border_width=2)
        product_details_frame.grid(row=0, column=0, sticky=W, padx=10, pady=10)

        product_details_title = Inv_titlelabel(product_details_frame, 0, 0, 10, "Product Details")

        product_id_label = Inv_Label(product_details_frame, 1, 0, 5, "Product ID:")
        product_id_label_details = Inv_Label(product_details_frame, 1, 1, 5, f"{values[0]}")
        product_name_label = Inv_Label(product_details_frame, 2, 0, 5, "Product Name:")
        product_name_label_details = Inv_Label(product_details_frame, 2, 1, 5, f"{values[1]}")
        product_quantity_label = Inv_Label(product_details_frame, 3, 0, 5, "Product Quantity:")
        product_quantity_label_details = Inv_Label(product_details_frame, 3, 1, 5, f"{values[2]}")
        product_description_label = Inv_Label(product_details_frame, 4, 0, 5, "Product Description:")
        product_description_label_details = Inv_Label(product_details_frame, 4, 1, 5, f"{values[3]}")

        restock_frame = customtkinter.CTkFrame(new_window, border_width=2)
        restock_frame.grid(row=1, column=0, padx=10, pady=10, columnspan=2)

        restock_title = Inv_titlelabel(restock_frame,0,0,10,"Restock")

        restock_quantity_label = Inv_Label(restock_frame, 1, 0, 5, "Restock Quantity:")
        restock_quantity_entry = Inv_Entrybox(restock_frame, 1, 1, 5, "Quantity")
        restock_btn = Inv_Button(restock_frame, 2, 0, 5, "#007FFF", "Restock", add_incoming_stock)

        new_window.mainloop()

    # -----------------------------------------------------------------------------------------------

    def fetch_task_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT TASK_ID,TASK_ASSIGN_DATE, TASK_DESCRIPTION,TASK_ASSIGNED_TO,TASK_STATUS,TASK_DUE_DATE, TASK_FINISH_DATE FROM TASK"
        )
        task = cursor.fetchall()
        conn.commit()
        conn.close()
        return task

    def fetch_worker_to_list():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT USERNAME FROM USER WHERE ACCESSLEVEL = 2")
        workers = cursor.fetchall()
        worker = []
        for x in workers:
            worker.append(x[0])
        conn.commit()
        conn.close()
        return worker

    def days_between(d1, d2):
        d1 = datetime.strptime(d1, "%Y-%m-%d")
        d2 = datetime.strptime(d2, "%Y-%m-%d")
        return (d2 - d1).days

    def check_task_due_date(pname):
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
                "SELECT TASK_DUE_DATE, TASK_ID, TASK_STATUS FROM TASK WHERE TASK_ASSIGNED_TO = ?",
                (pname, ),
        )
        tasks = cursor.fetchall()
        conn.commit()
        conn.close()

        today = str(datetime.today().date())

        due_task = 0
        for task in tasks:
            due_date_str = task[0]
            due_days = days_between(today, due_date_str)

            if due_days < 0:
                due_task += 1
                conn = sqlite3.connect("Inventory Management System.db")
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE TASK SET TASK_STATUS = ? WHERE TASK_ID = ?",
                    ("Due, Not Completed", task[1]),
                )
                conn.commit()
                conn.close()
            elif task[2] != 'Completed':
                conn = sqlite3.connect("Inventory Management System.db")
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE TASK SET TASK_STATUS = ? WHERE TASK_ID = ?",
                    ("Pending", task[1]),
                )
                conn.commit()
                conn.close()

        #check_task_status()
        add_to_task_table()
        return due_task

    def add_to_task_table():
        tasks = fetch_task_data()
        task_tree.delete(*task_tree.get_children())
        for task in tasks:
            task_tree.insert("", END, values=task, tags=task)

    def fetch_task_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(TASK_ID) FROM TASK")
        last_task_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_task_id

    def generate_task_id(prefix="TASK"):
        last_task_id = fetch_task_last_id()
        if last_task_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_task_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def add_new_task_details():
        task_id = generate_task_id("TASK")
        task_assigned_date = datetime.today().date()
        task_name = task_entry.get()
        assigned_worker = assigned_worker_entry.get()
        task_due_date = task_due_date_entry.get_date()

        if not (task_name):
            messagebox.showerror("Error", "Please enter all fields")
            return
        else:
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO TASK (TASK_ID, TASK_ASSIGN_DATE, TASK_DESCRIPTION, TASK_ASSIGNED_TO, TASK_STATUS, TASK_DUE_DATE) \
            VALUES (?,?,?,?,?,?)",
                (task_id, task_assigned_date, task_name, assigned_worker, "Pending", task_due_date),
            )
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Data has been inserted")
        add_to_task_table()
        return

    def edit_task_details():
        selected_task_details = task_tree.focus()
        if not selected_task_details:
            messagebox.showerror("Error", "Please select a record to edit")
            return
        row = task_tree.item(selected_task_details)["values"]
        new_task = task_entry.get()
        new_assigned_worker = assigned_worker_entry.get()
        new_due_date = task_due_date_entry.get_date()
        if not new_task:
            messagebox.showerror("Error", "Please enter all fields")
            return
        else:
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE TASK SET TASK_DESCRIPTION = ?, TASK_ASSIGNED_TO = ?, TASK_DUE_DATE = ? WHERE TASK_ID= ?",
                (new_task, new_assigned_worker, new_due_date, row[0]),
            )
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Record successfully edited")
        add_to_task_table()
        check_task_due_date(new_assigned_worker)

    def delete_task_database():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        selected_item = task_tree.focus()
        row = task_tree.item(selected_item)["values"]
        cursor.execute("DELETE FROM TASK WHERE TASK_ID= ?", (row[0],))
        conn.commit()
        conn.close()
        add_to_task_table()

    def delete_task_record():
        selected_item = task_tree.focus()
        if not selected_item:
            messagebox.showerror("Error", "Please select a record to delete")
            return
        confirmation = messagebox.askyesno(
            "Are you sure?", "Are you sure that you want to delete the selected record?"
        )
        if confirmation:
            delete_task_database()
            messagebox.showinfo("Success", "Record successfully deleted")
            return

    def fetch_purchase_order_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(PURCHASE_ORDER_ID) FROM PURCHASE_ORDER")
        last_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_id

    def generate_purchase_order_id(prefix="PO"):
        last_id = fetch_purchase_order_last_id()
        if last_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def fetch_incoming_stock_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM PURCHASE_ORDER")
        incoming_stock = cursor.fetchall()
        conn.commit()
        conn.close()
        return incoming_stock

    def add_to_incoming_stock_table():
        incoming_stocks = fetch_incoming_stock_data()
        incoming_stock_tree.delete(*incoming_stock_tree.get_children())
        for incoming_stock in incoming_stocks:
            incoming_stock_tree.insert(
                "", END, values=incoming_stock, tags=incoming_stock
            )

    def search_incoming_stock(event):
        search_term = search_incoming_stock_entry.get().lower()
        incoming_stocks = fetch_incoming_stock_data()
        incoming_stock_tree.delete(*incoming_stock_tree.get_children())
        for incoming_stock in incoming_stocks:
            if search_term in str(incoming_stock).lower():
                incoming_stock_tree.insert("", tk.END, values=incoming_stock, tags=incoming_stock)

    def update_purchase_order_status():
        new_status = purchase_order_status_entry.get()
        selected = incoming_stock_tree.focus()
        if not selected:
            messagebox.showerror("Error", "Please select a record to edit")
            return

        row = incoming_stock_tree.item(selected)["values"]
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()

        try:
            cursor.execute(
                "UPDATE PURCHASE_ORDER SET PURCHASE_ORDER_STATUS = ? WHERE PURCHASE_ORDER_ID = ?",
                (new_status, row[0]),
            )

            if new_status == "Received":
                cursor.execute(
                    "UPDATE PRODUCT SET PRODUCT_QUANTITY = PRODUCT_QUANTITY + ? WHERE PRODUCT_NAME = ?",
                    (row[2], row[1]),
                )

                # Retrieve the user's full name
                cursor.execute(
                    "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
                )
                fullname = cursor.fetchone()

                # Log the product movement only when the status is "Received"
                cursor.execute(
                    "INSERT INTO PRODUCT_MOVEMENT (PRODUCT_MOVEMENT_ID, PRODUCT_MOVEMENT_DATE, PRODUCT_MOVEMENT, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_product_movement_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "{} received product: {} (ID: {}) with quantity: {}".format(fullname[0], row[1], row[0],
                                                                                    row[2]),
                        username,
                    ),
                )

            conn.commit()
            messagebox.showinfo("Success", "Record successfully edited")

        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            conn.close()

        add_to_incoming_stock_table()
        add_to_product_table()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()
        low_stock_vs_total_item_pie_chart(canvas)
        bar_chart(canvas1)

    def fetch_sale_order_stock_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM SALE_ORDER")
        sale_order = cursor.fetchall()
        conn.commit()
        conn.close()
        return sale_order

    def add_to_sale_order_table():
        outgoing_stocks = fetch_sale_order_stock_data()
        outgoing_stock_tree.add_to_table(outgoing_stocks)

    def search_outgoing_stock(event):
        search_term = search_outgoing_stock_entry.get().lower()
        outgoing_stocks = fetch_sale_order_stock_data()
        outgoing_stock_tree.search_item(search_term, outgoing_stocks)

    def update_sales_order_status():
        new_status = sales_order_status_entry.get()
        selected = outgoing_stock_tree.selected_item()
        if not selected:
            messagebox.showerror("Error", "Please select a record to edit")
            return
        row = selected
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE SALE_ORDER SET SALE_ORDER_STATUS = ? WHERE SALE_ORDER_ID = ?",
            (new_status, row[0]),
        )
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Record successfully edited")
        add_to_sale_order_table()
        add_to_product_table()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()

    def fetch_sale_order_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(SALE_ORDER_ID) FROM SALE_ORDER")
        last_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_id

    def generate_sale_order_id(prefix="SO"):
        last_id = fetch_sale_order_last_id()
        if last_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def on_sale_order_double_click(event):
        values = outgoing_stock_tree.selected_item()

        sale_order_window = customtkinter.CTk()
        sale_order_window.title("Product Details")
        sale_order_window.resizable(False, False)

        product_selection_frame = customtkinter.CTkFrame(
            master=sale_order_window, border_width=2
        )

        sale_order_product_table_frame = customtkinter.CTkFrame(
            master=sale_order_window
        )

        def fetch_sale_order_status(sale_order_id):
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute("SELECT SALE_ORDER_STATUS FROM SALE_ORDER WHERE SALE_ORDER_ID = ?", (sale_order_id,))
            status = cursor.fetchone()
            conn.close()
            return status[0] if status else None

        def fetch_sale_order_product_data():
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT * FROM SALE_ORDER_PRODUCT WHERE SALE_ORDER_ID = ?", (values[0],)
            )
            sale_order = cursor.fetchall()
            conn.commit()
            conn.close()
            return sale_order

        def add_to_sale_order_product_table():
            sale_order_products = fetch_sale_order_product_data()
            sale_order_product_tree.delete(*sale_order_product_tree.get_children())
            for sale_order_product in sale_order_products:
                sale_order_product_tree.insert("", END, values=sale_order_product)

        def update_product_details(choice):
            selected_product_id = product_selection_entry.get()

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT PRODUCT_NAME,PRODUCT_STATUS FROM PRODUCT WHERE PRODUCT_ID = ?", (selected_product_id,)
            )
            choice = cursor.fetchall()
            conn.commit()
            conn.close()

            product_selection_name_label2.configure(text=choice[0][0])
            product_selection_description_label2.configure(text=choice[0][1])


        def add_sale_order_product():
            sale_order_id = values[0]
            sale_order_product_id = product_selection_entry.get()
            sale_order_product_quantity = sales_quantity_entry.get()

            if not sale_order_product_quantity:
                messagebox.showerror("Error", "Please enter all fields")
                return

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()

            # Check the status of the sale order
            cursor.execute(
                "SELECT SALE_ORDER_STATUS FROM SALE_ORDER WHERE SALE_ORDER_ID = ?", (sale_order_id,)
            )
            sale_order_status = cursor.fetchone()[0]

            if sale_order_status != "To be Packed":
                messagebox.showerror("Error", "Cannot add items. The status is not 'To be Packed'")
                conn.close()
                return

            # Get Product Name
            cursor.execute(
                "SELECT PRODUCT_NAME FROM PRODUCT WHERE PRODUCT_ID = ?", (sale_order_product_id,)
            )
            product_name = cursor.fetchone()[0]

            cursor.execute(
                "SELECT PRODUCT_QUANTITY FROM PRODUCT WHERE PRODUCT_ID = ?", (sale_order_product_id,),
            )
            product_quantity = cursor.fetchone()[0]

            if product_quantity < int(sale_order_product_quantity):
                messagebox.showerror("Error", "Insufficient quantity")
                conn.close()
                return

            try:
                cursor.execute(
                    "INSERT INTO SALE_ORDER_PRODUCT (SALE_ORDER_ID,SALE_ORDER_PRODUCT_ID, SALE_ORDER_PRODUCT, SALE_ORDER_PRODUCT_QUANTITY) \
                    VALUES (?,?,?,?)",
                    (sale_order_id, sale_order_product_id, product_name, sale_order_product_quantity),
                )
                cursor.execute(
                    "UPDATE PRODUCT SET PRODUCT_QUANTITY = PRODUCT_QUANTITY - ? WHERE PRODUCT_ID = ?",
                    (sale_order_product_quantity, sale_order_product_id),
                )

                # Retrieve the user's full name
                cursor.execute(
                    "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
                )
                fullname = cursor.fetchone()

                # Log the product movement
                cursor.execute(
                    "SELECT PRODUCT_ID FROM PRODUCT WHERE PRODUCT_NAME = ?", (product_name,)
                )
                product_id = cursor.fetchone()[0]
                cursor.execute(
                    "INSERT INTO PRODUCT_MOVEMENT (PRODUCT_MOVEMENT_ID, PRODUCT_MOVEMENT_DATE, PRODUCT_MOVEMENT, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_product_movement_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "{} added product: {} (ID: {}) with quantity: {} into sale order (ID: {})".format(fullname[0],
                                                                                                          product_name,
                                                                                                          sale_order_product_id,
                                                                                                          sale_order_product_quantity,
                                                                                                          sale_order_id),
                        username,
                    ),
                )

                conn.commit()
                messagebox.showinfo("Success", "Data has been inserted")

            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()

            add_to_sale_order_product_table()
            add_to_product_table()
            update_low_stock_item_label()
            add_to_outgoing_product_table()
            return

        select_product_label = Inv_titlelabel(product_selection_frame,0,0,10,"Select Product")

        product_selection_label = customtkinter.CTkLabel(
            product_selection_frame,
            text="Product to be purchase:",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        product_selection_entry = customtkinter.CTkComboBox(
            master=product_selection_frame,
            values=fetch_product_to_list(),
            width=230,
            height=30,
            font=customtkinter.CTkFont("SF Pro Display"),
            command=update_product_details,
            border_width=0,
        )
        product_selection_name_label1 = customtkinter.CTkLabel(
            product_selection_frame,
            text="Product Name:",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        product_selection_name_label2 = customtkinter.CTkLabel(
            product_selection_frame,
            text="",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        product_selection_description_label1 = customtkinter.CTkLabel(
            product_selection_frame,
            text="Product Description:",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        product_selection_description_label2 = customtkinter.CTkLabel(
            product_selection_frame,
            text="",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        sales_quantity_label = customtkinter.CTkLabel(
            product_selection_frame,
            text="Purchase Quantity:",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        sales_quantity_entry = customtkinter.CTkEntry(
            master=product_selection_frame,
            placeholder_text="Quantity",
            width=230,
            height=30,
            font=customtkinter.CTkFont("SF Pro Display"),
            border_width=0,
        )
        sales_btn = customtkinter.CTkButton(
            master=product_selection_frame,
            text="Add Item",
            command=add_sale_order_product,
            font=customtkinter.CTkFont("SF Pro Display"),
            corner_radius=200,
            fg_color="#007FFF",
            text_color="black",
        )
        # --------------------------------------------------------

        style1 = ThemedStyle()
        style1.set_theme("equilux")
        style1.configure(
            "Treeview",
            font=("SF Pro Display", 10),
            rowheight=24,
            foreground="white",
            background="#2a2d2e",
            fieldbackground="#343638",
            bordercolor="#343638",
            borderwidth=0,
        )
        style1.map("Treeview", background=[("selected", "#22559b")])
        style1.configure(
            "Treeview.Heading", background="#565b5e", foreground="white", relief="flat"
        )
        style1.map("Treeview.Heading", background=[("active", "#3484F0")])
        style1.configure(
            "DateEntry",
            font=customtkinter.CTkFont("SF Pro Display", 12),
            foreground="white",
            background="#2a2d2e",
            fieldbackground="#343638",
            bordercolor="#343638",
            borderwidth=0,
        )

        sale_order_product_tree = ttk.Treeview(
            master=sale_order_product_table_frame, height=30
        )

        sale_order_product_verscrlbar = customtkinter.CTkScrollbar(
            master=sale_order_product_table_frame,
            orientation="vertical",
            command=sale_order_product_tree.yview,
        )
        sale_order_product_verscrlbar.pack(side="right", fill="y")

        sale_order_product_tree.configure(
            yscrollcommand=sale_order_product_verscrlbar.set
        )
        sale_order_product_tree["columns"] = (
            "SALEORDERID",
            "SALEORDERPRODUCTID",
            "SALEORDERPRODUCT",
            "SALEORDERPRODUCTQUANTITY",
        )

        sale_order_product_tree.column("#0", width=0, stretch=tk.NO)
        sale_order_product_tree.column("SALEORDERID", anchor=tk.CENTER, width=150)
        sale_order_product_tree.column("SALEORDERPRODUCTID", anchor=tk.CENTER, width=150)
        sale_order_product_tree.column("SALEORDERPRODUCT", anchor=tk.CENTER, width=150)
        sale_order_product_tree.column("SALEORDERPRODUCTQUANTITY", anchor=tk.CENTER, width=150)

        sale_order_product_tree.heading("SALEORDERID", text="Sales Order ID")
        sale_order_product_tree.heading("SALEORDERPRODUCTID", text="Item ID")
        sale_order_product_tree.heading("SALEORDERPRODUCT", text="Item")
        sale_order_product_tree.heading("SALEORDERPRODUCTQUANTITY", text="Quantity")

        sale_order_product_tree.pack(side="bottom", fill="both")

        product_selection_frame.grid(row=0, column=0, padx=10, pady=10, sticky=NW)
        product_selection_label.grid(row=1, column=0, padx=10, pady=10)
        product_selection_entry.grid(row=1, column=1, padx=10, pady=10)
        product_selection_name_label1.grid(row=2, column=0, padx=10, pady=10)
        product_selection_name_label2.grid(row=2, column=1, padx=10, pady=10)
        product_selection_description_label1.grid(row=3, column=0, padx=10, pady=10)
        product_selection_description_label2.grid(row=3, column=1, padx=10, pady=10)
        sales_quantity_label.grid(row=4, column=0, padx=10, pady=10)
        sales_quantity_entry.grid(row=4, column=1, padx=10, pady=10)
        sales_btn.grid(row=5, column=0, padx=10, pady=10, columnspan=2)
        sale_order_product_table_frame.grid(row=0, column=1, padx=10, pady=10)

        add_to_sale_order_product_table()
        sale_order_window.mainloop()

    def fetch_outgoing_product_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT SALE_ORDER_PRODUCT_ID, SALE_ORDER_PRODUCT, SUM(SALE_ORDER_PRODUCT_QUANTITY) FROM SALE_ORDER_PRODUCT \
              GROUP BY SALE_ORDER_PRODUCT_ID"
        )
        sale_order = cursor.fetchall()
        conn.commit()
        conn.close()
        return sale_order

    def add_to_outgoing_product_table():
        outgoing_stocks = fetch_outgoing_product_data()
        outgoing_product_tree.add_to_table(outgoing_stocks)

    def low_stock_vs_total_item_pie_chart():
        low_stock_items = product_dbase.fetch_low_stock_item_data()
        total_items = product_dbase.fetch_total_items_data()
        piechart.display(total_items, low_stock_items)

    def bar_chart():
        barchart_data = product_dbase.fetch_bar_chart_data()
        barchart.display(barchart_data)

    def performance_report():
        selected_worker = selected_worker_entry.get()
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()

        # Fetch completed tasks
        cursor.execute("SELECT COUNT(*) FROM TASK WHERE TASK_ASSIGNED_TO = ? AND TASK_STATUS = ?",
                       (selected_worker, "Completed"))
        completed_data = cursor.fetchone()[0]

        # Fetch pending tasks
        cursor.execute("SELECT COUNT(*) FROM TASK WHERE TASK_ASSIGNED_TO = ? AND TASK_STATUS = ?",
                       (selected_worker, "Pending"))
        pending_data = cursor.fetchone()[0]

        # Fetch due, not completed tasks
        cursor.execute("SELECT COUNT(*) FROM TASK WHERE TASK_ASSIGNED_TO = ? AND TASK_STATUS = ?",
                       (selected_worker, "Due, Not Completed"))
        due_not_complete_data = cursor.fetchone()[0]

        # Fetch due, completed tasks
        cursor.execute("SELECT COUNT(*) FROM TASK WHERE TASK_ASSIGNED_TO = ? AND TASK_STATUS = ?",
                       (selected_worker, "Due, Completed"))
        due_complete_data = cursor.fetchone()[0]

        cursor.execute("SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?",
                       (selected_worker,))
        worker_name = cursor.fetchone()[0]

        conn.close()

        # Data for plotting
        labels = ['Completed', 'Pending', 'Due, Not Completed', 'Due, Completed']
        values = [completed_data, pending_data, due_not_complete_data, due_complete_data]

        # Calculate completion rate
        total_tasks = sum(values)
        if total_tasks > 0:
            completion_rate = (completed_data / total_tasks) * 100
        else:
            completion_rate = 0

        # Calculate due rate
        due_rate = ((due_not_complete_data + due_complete_data) / total_tasks) * 100 if total_tasks > 0 else 0

        # Plotting the data as a donut chart
        plt.figure(figsize=(8, 8))
        plt.pie(values, labels=labels, autopct='%1.1f%%', colors=['green', 'orange', 'red', 'pink'], startangle=140,
                wedgeprops={'edgecolor': 'white', 'linewidth': 2}, pctdistance=0.85)
        plt.title(f'Performance Report for {worker_name}')
        plt.gca().add_artist(plt.Circle((0, 0), 0.6, color='white'))  # Add a white circle to create the donut hole

        # Add completion rate text
        plt.text(0, -0.05, f'Completion Rate: {completion_rate:.2f}%', fontsize=12, ha='center')

        # Add due rate text
        plt.text(0, -0.15, f'Due Rate: {due_rate:.2f}%', fontsize=12, ha='center')

        # Add date generated below the chart
        date_generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        plt.text(-0.9, -1.3, f'Date Generated: {date_generated}', fontsize=10, ha='center')

        # Save the plot as a PDF using file dialog
        file_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                 filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if file_path:
            pdf_filename = file_path
            with PdfPages(pdf_filename) as pdf:
                pdf.savefig()
                plt.close()

    def fetch_user_activities_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(USER_ACTIVITIES_ID) FROM USER_ACTIVITIES")
        last_user_activities_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_user_activities_id

    def generate_user_activities_id(prefix="UA"):
        last_customer_id = fetch_user_activities_last_id()
        if last_customer_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_customer_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def fetch_product_movement_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(PRODUCT_MOVEMENT_ID) FROM PRODUCT_MOVEMENT")
        last_product_movement_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_product_movement_id

    def generate_product_movement_id(prefix="PM"):
        product_movement_id = fetch_product_movement_last_id()
        if product_movement_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(product_movement_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def logout():
        confirmation = messagebox.askyesno('Are you sure?', 'Are you sure that you want to logout?')
        if confirmation:
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
            )
            fullname = cursor.fetchone()
            conn.commit()
            conn.close()

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID,USER_ACTIVITIES_DATE,USER_ACTIVITIES,USER) \
            VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "{} logged out".format(fullname[0]),
                    username,
                ),
            )
            conn.commit()
            conn.close()

            supervisor_dashboard_frame.destroy()
            login_page()

    def sort_treeview_column(tv, col, reverse):
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(reverse=reverse)

        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)

        tv.heading(col, command=lambda: sort_treeview_column(tv, col, not reverse))

    supervisor_dashboard_frame = customtkinter.CTkFrame(
        master=app, width=1280, height=720, border_width=0
    )
    supervisor_dashboard_frame.place(x=0, y=0)

    info_frame = customtkinter.CTkFrame(
        master=supervisor_dashboard_frame, width=1280, height=720
    )
    info_frame.place(x=0, y=0)
    info_tab = customtkinter.CTkTabview(master=info_frame, width=1280, height=720)
    custom_font = ("SF Pro Display", 15)
    info_tab._segmented_button.configure(font=custom_font, fg_color='grey', text_color='white')
    info_tab.place(x=0, y=0)
    tab_0 = info_tab.add(" Homepage ")
    tab_1 = info_tab.add(" Customer ")
    #tab_2 = info_tab.add("Supplier")
    tab_3 = info_tab.add(" Product ")
    tab_4 = info_tab.add(" Task")
    tab_5 = info_tab.add(" Incoming ")
    tab_6 = info_tab.add(" Outgoing ")

    # -----------------------------
    outgoing_table_title1_label = customtkinter.CTkLabel(master=tab_6, text_color="purple",
                                                         text='Double Click on the selected Sales Order below to enter Product & Quantity',
                                                         font=customtkinter.CTkFont("Times", 13, weight='bold'))
    outgoing_table_title1_label.place(x=100, y=180)

    outgoing_table_title2_label = customtkinter.CTkLabel(master=tab_6, text_color="darkgreen",
                                                         text='Summary of Total Outgoing Quantity by Product',
                                                         font=customtkinter.CTkFont("Times", 13, weight='bold'))
    outgoing_table_title2_label.place(x=850, y=180)
    customer_table_title1_label = customtkinter.CTkLabel(master=tab_1, text_color="purple",
                                                         text='Double Click on the selected Customer below to create Sales Order',
                                                         font=customtkinter.CTkFont("Times", 15, weight='bold'))
    customer_table_title1_label.place(x=50, y=180)
    product_table_title1_label = customtkinter.CTkLabel(master=tab_3, text_color="purple",
                                                         text='Double Click on the selected Product below to Restock the Quantity',
                                                         font=customtkinter.CTkFont("Times", 15, weight='bold'))
    product_table_title1_label.place(x=50, y=180)
    worker_performance_table_title1_label = customtkinter.CTkLabel(master=tab_4, text_color="purple",
                                                        text='Worker Task Completion Performance',
                                                        font=customtkinter.CTkFont("Times", 15, weight='bold'))
    worker_performance_table_title1_label.place(x=30, y=0)

    # --------------------------------

    style = ThemedStyle()
    style.set_theme("equilux")
    style.configure(
        "Treeview",
        font=("SF Pro Display", 10),
        rowheight=24,
        foreground="white",
        background="#2a2d2e",
        fieldbackground="#343638",
        bordercolor="#343638",
        borderwidth=0,
    )
    style.map("Treeview", background=[("selected", "#22559b")])
    style.configure(
        "Treeview.Heading", background="#565b5e", foreground="white", relief="flat"
    )
    style.map("Treeview.Heading", background=[("active", "#3484F0")])
    style.configure(
        "DateEntry",
        font=customtkinter.CTkFont(("SF Pro Display"), 12),
        foreground="white",
        background="#2a2d2e",
        fieldbackground="#343638",
        bordercolor="#343638",
        borderwidth=0,
    )

    welcome_label = customtkinter.CTkLabel(
        master=tab_0,
        text=f"Hello, {fetch_user_fullname()} ",
        font=customtkinter.CTkFont("SF Pro Display", 24, weight="bold"),
    )

    sales_activity_frame = customtkinter.CTkFrame(
        tab_0, corner_radius=10, border_width=2
    )
    sales_activity_title = Inv_titlelabel(sales_activity_frame, 0, 0, 10, "Sales Activity")

    sales_activity_vertical_separator_1 = customtkinter.CTkFrame(
        sales_activity_frame, width=2, height=50, fg_color="#CCCCCC"
    )

    sales_activity_vertical_separator_2 = customtkinter.CTkFrame(
        sales_activity_frame, width=2, height=50, fg_color="#CCCCCC"
    )

    inventory_summary_frame = customtkinter.CTkFrame(
        tab_0, corner_radius=10, border_width=2
    )

    sales_activity_vertical_separator_4 = customtkinter.CTkFrame(
        inventory_summary_frame, width=2, height=50, fg_color="#CCCCCC"
    )
    product_details_frame = customtkinter.CTkFrame(
        tab_0, corner_radius=10, border_width=2
    )
    sales_activity_vertical_separator_6 = customtkinter.CTkFrame(
        product_details_frame, width=2, height=50, fg_color="#CCCCCC"
    )
    product_details_title = Inv_titlelabel(product_details_frame, 0, 0, 10, "Product Details")

    logout_btn = Inv_Button(tab_0, 0, 2, 10, "#ff666d", "Logout", logout)

    welcome_label.grid(row=0, column=0, sticky=W, padx=10, pady=10)
    sales_activity_frame.grid(row=1, column=0, padx=10, pady=10)
    inventory_summary_frame.grid(row=1, column=1, padx=10, pady=10)
    inventory_summary_title = Inv_titlelabel(inventory_summary_frame, 0, 0, 10, "Inventory Summary")
    product_details_frame.grid(row=1, column=2, padx=10, pady=10)

    piechart = Inv_pie_chart(tab_0, 2, 0)
    barchart = Inv_bar_chart(tab_0, 2, 1)
    low_stock_vs_total_item_pie_chart()
    bar_chart()

    total_to_be_packed_label_1 = Inv_Label(sales_activity_frame, 1, 0, 20, str(fetch_to_be_packed_data()))
    total_to_be_packed_label_2 = Inv_Label(sales_activity_frame, 2, 0, 20, "To be Packed")
    sales_activity_vertical_separator_1.grid(row=1, column=1, rowspan=2, padx=5, pady=5)
    total_to_be_shipped_label_1 = Inv_Label(sales_activity_frame, 1, 2, 20, str(fetch_to_be_shipped_data()))
    total_to_be_shipped_label_2 = Inv_Label(sales_activity_frame, 2, 2, 20, "To be Shipped")
    sales_activity_vertical_separator_2.grid(row=1, column=3, rowspan=2, padx=5, pady=5)
    total_to_be_delivered_label_1 = Inv_Label(sales_activity_frame, 1, 4, 20, str(fetch_to_be_delivered_data()))
    total_to_be_delivered_label_2 = Inv_Label(sales_activity_frame, 2, 4, 20, "To be Delivered")

    total_quantity_in_hand_1 = Inv_Label(inventory_summary_frame, 1, 0, 20,
                                         str(product_dbase.fetch_total_quantity_in_hand_data()))
    total_quantity_in_hand_2 = Inv_Label(inventory_summary_frame, 2, 0, 20, "Quantity In Hand")
    sales_activity_vertical_separator_4.grid(row=1, column=1, rowspan=2, padx=5, pady=5)
    total_quantity_to_be_received_1 = Inv_Label(inventory_summary_frame, 1, 2, 20,
                                                str(fetch_total_quantity_to_be_received_data()))
    total_quantity_to_be_received_2 = Inv_Label(inventory_summary_frame, 2, 2, 20, "Quantity To be Received")

    low_stock_items_1 = Inv_Label(product_details_frame, 1, 0, 20, str(product_dbase.fetch_low_stock_item_data()))
    low_stock_items_2 = Inv_Label(product_details_frame, 2, 0, 20, "Low Stock Items")
    sales_activity_vertical_separator_6.grid(row=1, column=1, rowspan=2, padx=5, pady=5)
    all_items_1 = Inv_Label(product_details_frame, 1, 2, 20, str(product_dbase.fetch_total_items_data()))
    all_items_2 = Inv_Label(product_details_frame, 2, 2, 20, "Total Items")

    tab_0.columnconfigure(1, weight=1)

    # ---------------------------------------------------------

    customer_table_frame = customtkinter.CTkFrame(master=tab_1, width=1280, height=515)
    customer_table_frame.place(x=0, y=205)

    customer_tree = ttk.Treeview(master=customer_table_frame, height=18)

    customer_verscrlbar = customtkinter.CTkScrollbar(
        master=customer_table_frame, orientation="vertical", command=customer_tree.yview
    )
    customer_verscrlbar.pack(side="right", fill="y")

    customer_tree.configure(yscrollcommand=customer_verscrlbar.set)
    customer_tree["columns"] = (
        "CUSTOMERID",
        "CUSTOMERNAME",
        "CUSOTMEREMAILADDRESS",
        "CUSTOMERCONTACTNO",
    )

    customer_tree.column("#0", width=0, stretch=tk.NO)
    customer_tree.column("CUSTOMERID", anchor=tk.CENTER, width=313)
    customer_tree.column("CUSTOMERNAME", anchor=tk.CENTER, width=313)
    customer_tree.column("CUSOTMEREMAILADDRESS", anchor=tk.CENTER, width=313)
    customer_tree.column("CUSTOMERCONTACTNO", anchor=tk.CENTER, width=313)

    customer_tree.heading("CUSTOMERID", text="ID")
    customer_tree.heading("CUSTOMERNAME", text="Name")
    customer_tree.heading("CUSOTMEREMAILADDRESS", text="Email Address")
    customer_tree.heading("CUSTOMERCONTACTNO", text="Contact No.")

    customer_tree.pack(side="right", fill="both")
    customer_tree.bind("<Double-1>", on_customer_double_click)
    customer_tree.bind("<ButtonRelease>", display_customer_record)

    customer_menu_frame = customtkinter.CTkFrame(master=tab_1, border_width=2)
    customer_menu_frame.grid(row=0, column=0)

    insert_customer_data_label = Inv_main_titlelabel(customer_menu_frame,0,0,
                                                     5,5,"Insert New Customer")

    customer_name_entry_label = Inv_Label(customer_menu_frame, 1, 0, 5, "Customer Name:")
    customer_name_entry = Inv_Entrybox(customer_menu_frame, 1, 1, 5, "Name")
    customer_email_entry_label = Inv_Label(customer_menu_frame, 1, 2, 5, "Customer Email Address:")
    customer_email_entry = Inv_Entrybox(customer_menu_frame, 1, 3, 5, "Email Address")
    customer_contact_entry_label = Inv_Label(customer_menu_frame, 1, 4, 5, "Customer Contact No.:")
    customer_contact_entry = Inv_Entrybox(customer_menu_frame, 1, 5, 5, "Contact No.")

    search_customer_entry = customtkinter.CTkEntry(tab_1, placeholder_text="Search", width=1050, )
    search_customer_entry.grid(row=1, column=0, padx=10, pady=10)
    search_customer_entry.bind("<KeyRelease>", search_customer)

    addcustomer_btn = Inv_Button(customer_menu_frame, 2, 0, 5, "#007FFF", "Add", add_new_customer_details)
    editcustomer_btn = Inv_Button(customer_menu_frame, 2, 2, 5, "#ADD8E6", "Edit", edit_customer_details)
    deletecustomer_btn = Inv_Button(customer_menu_frame, 2, 4, 5, "#ff666d", "Delete", delete_customer_record)

    tab_1.columnconfigure(0, weight=2)

    # --------------------------------------------------

    product_table_frame = customtkinter.CTkFrame(master=tab_3, width=1280, height=515)
    product_table_frame.place(x=0, y=205)
    admin_product_tree = Inv_product_tree_display(product_table_frame, on_product_double_click, display_product_record)

    search_entry = customtkinter.CTkEntry(tab_3, placeholder_text="Search", width=1050, )
    search_entry.grid(row=1, column=0, padx=10, pady=10)
    search_entry.bind("<KeyRelease>", search_product)

    product_menu_frame = customtkinter.CTkFrame(
        master=tab_3,
        border_width=2
    )
    product_menu_frame.grid(row=0, column=0)

    insert_product_data_label = Inv_main_titlelabel(product_menu_frame, 0, 0, 5, 5, "Insert New Product")
    product_id_entry_label = Inv_Label(product_menu_frame, 1, 0, 5, "Product ID:")
    product_id_entry = Inv_Entrybox(product_menu_frame, 1, 1, 5, "ID")
    product_name_entry_label = Inv_Label(product_menu_frame, 1, 2, 5, "Product Name:")
    product_name_entry = Inv_Entrybox(product_menu_frame, 1, 3, 5, "Item")
    product_quantity_entry_label = Inv_Label(product_menu_frame, 1, 4, 5, "Product Quantity:")
    product_quantity_entry = Inv_Entrybox(product_menu_frame, 1, 5, 5, "Quantity")
    product_description_entry_label = Inv_Label(product_menu_frame, 2, 0, 5, "Product Description:")
    product_description_entry = Inv_Entrybox(product_menu_frame, 2, 1, 5, "Description")
    product_min_stock_entry_label = Inv_Label(product_menu_frame, 2, 2, 5, "Product Min. Stock:")
    product_min_stock_entry = Inv_Entrybox(product_menu_frame, 2, 3, 5, "Min. Stock")

    addproduct_btn = Inv_Button(product_menu_frame, 3, 0, 20, "#007FFF", "Add Product", add_new_product_details)
    editproduct_btn = Inv_Button(product_menu_frame, 3, 2, 20, "#ADD8E6", "Edit Product", edit_product_details)
    deleteproduct_btn = Inv_Button(product_menu_frame, 3, 4, 20, "#ff666d", "Delete Product", delete_product_record)

    # ---------------------------------------------
    tab_3.columnconfigure(0, weight=2)

    def clear_task_entry_field():
        task_entry.delete(0, END)
        task_due_date_entry.delete(0, END)

    def display_task_record(event):
        selected_item = task_tree.focus()
        if selected_item:
            clear_task_entry_field()
            row = task_tree.item(selected_item)["values"]
            task_entry.insert(0, row[2])
            task_due_date_entry.insert(0, row[5])
            assigned_worker_entry.set(row[3])

    task_table_frame = customtkinter.CTkFrame(master=tab_4, width=1280, height=515)
    task_table_frame.place(x=0, y=250)

    task_tree = ttk.Treeview(master=task_table_frame, height=16)

    task_verscrlbar = customtkinter.CTkScrollbar(
        master=task_table_frame, orientation="vertical", command=task_tree.yview
    )
    task_verscrlbar.pack(side="right", fill="y")

    task_tree.configure(yscrollcommand=task_verscrlbar.set)
    task_tree["columns"] = (
        "TASKID",
        "ASSIGNDATE",
        "TASKDESCRIPTION",
        "ASSIGNEDTO",
        "STATUS",
        "DUEDATE",
        "FINISHDATE"
    )

    task_tree.column("#0", width=0, stretch=tk.NO)
    task_tree.column("TASKID", anchor=tk.CENTER, width=100)
    task_tree.column("ASSIGNDATE", anchor=tk.CENTER, width=130)
    task_tree.column("TASKDESCRIPTION", anchor=tk.CENTER, width=500)
    task_tree.column("ASSIGNEDTO", anchor=tk.CENTER, width=160)
    task_tree.column("STATUS", anchor=tk.CENTER, width=100)
    task_tree.column("DUEDATE", anchor=tk.CENTER, width=130)
    task_tree.column("FINISHDATE", anchor=tk.CENTER, width=130)

    task_tree.heading("TASKID", text="Task ID")
    task_tree.heading("ASSIGNDATE", text="Assigned Date")
    task_tree.heading("TASKDESCRIPTION", text="Task")
    task_tree.heading("ASSIGNEDTO", text="Assigned To")
    task_tree.heading("STATUS", text="Status")
    task_tree.heading("DUEDATE", text="Due Date")
    task_tree.heading("FINISHDATE", text="Completed Date")

    task_tree.pack(side="bottom", fill="both")
    task_tree.bind("<ButtonRelease>", display_task_record)

    task_tree.tag_configure("Pending", background="#FFBA00", foreground="black")
    task_tree.tag_configure("Completed", background="green", foreground="black")
    task_tree.tag_configure("Due, Not Completed", background="#FF7F7F", foreground="black")
    task_tree.tag_configure("Due, Completed", background="#FF7F7F", foreground="black")

    task_menu_frame = customtkinter.CTkFrame(master=tab_4, border_width=2)
    task_menu_frame.grid(row=0, column=1)
    task_menu_frame2 = customtkinter.CTkFrame(master=tab_4, border_width=2)

    insert_task_data_label = Inv_main_titlelabel(task_menu_frame,0,0,
                                                 10,10,"Assign New Task")

    task_label = customtkinter.CTkLabel(
        master=task_menu_frame,
        text="Task:",
        font=customtkinter.CTkFont("SF Pro Display"),
    )
    task_entry = customtkinter.CTkEntry(
        master=task_menu_frame,
        placeholder_text="Task",
        width=230,
        height=30,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )

    assigned_worker_label = customtkinter.CTkLabel(
        master=task_menu_frame,
        text="Assigned To:",
        font=customtkinter.CTkFont("SF Pro Display"),
    )

    assigned_worker_entry = customtkinter.CTkComboBox(
        master=task_menu_frame,
        values=fetch_worker_to_list(),
        width=230,
        height=30,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )

    task_due_date_label = customtkinter.CTkLabel(
        master=task_menu_frame,
        text="Due Date:",
        font=customtkinter.CTkFont("SF Pro Display"),
    )

    task_due_date_entry = DateEntry(
        master=task_menu_frame,
        date=datetime.today(),
        relief=GROOVE,
        font=customtkinter.CTkFont("SF Pro Display", 12),
        width=25,
    )

    addtask_btn = customtkinter.CTkButton(
        master=task_menu_frame,
        text="Add",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=add_new_task_details,
        compound="top",
        corner_radius=200,
        fg_color="#007FFF",
        text_color="black",
    )

    edittask_btn = customtkinter.CTkButton(
        master=task_menu_frame,
        text="Edit",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=edit_task_details,
        compound="top",
        corner_radius=200,
        fg_color="#ADD8E6",
        text_color="black",
    )

    deletetask_btn = customtkinter.CTkButton(
        master=task_menu_frame,
        text="Delete",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=delete_task_record,
        compound="top",
        corner_radius=200,
        fg_color="#ff666d",
        text_color="black",
    )

    generate_performance_report_label = customtkinter.CTkLabel(
        master=task_menu_frame2,
        text="Generate Performance Report",
        font=customtkinter.CTkFont("SF Pro Display", weight="bold", size=20),
    )

    selected_worker_label = customtkinter.CTkLabel(
        master=task_menu_frame2,
        text="Worker:",
        font=customtkinter.CTkFont("SF Pro Display"),
    )

    selected_worker_entry = customtkinter.CTkComboBox(
        master=task_menu_frame2,
        values=fetch_worker_to_list(),
        width=230,
        height=30,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )

    generatereport_btn = customtkinter.CTkButton(
        master=task_menu_frame2,
        text="Generate",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=performance_report,
        compound="top",
        corner_radius=200,
        text_color="black",
    )

    task_label.grid(row=1, column=0, padx=5, pady=5)
    task_entry.grid(row=1, column=1, padx=5, pady=5)
    assigned_worker_label.grid(row=2, column=0, padx=5, pady=5)
    assigned_worker_entry.grid(row=2, column=1, padx=5, pady=5)
    task_due_date_label.grid(row=3, column=0, padx=5, pady=5)
    task_due_date_entry.grid(row=3, column=1, padx=5, pady=5)
    addtask_btn.grid(row=4, column=0, padx=5, pady=5)
    edittask_btn.grid(row=4, column=1, padx=5, pady=5)
    deletetask_btn.grid(row=4, column=2, padx=5, pady=5)

    task_menu_frame2.grid(row=0, column=2)
    generate_performance_report_label.grid(row=0, column=0, sticky=W, padx=10, pady=10, columnspan=2)
    selected_worker_label.grid(row=1, column=0, padx=5, pady=5)
    selected_worker_entry.grid(row=1, column=1, padx=5, pady=5)
    generatereport_btn.grid(row=2, column=0, padx=5, pady=5, columnspan=2)

    def fetch_task_summary_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT TASK_ASSIGNED_TO, avg(julianday(TASK_FINISH_DATE) - julianday(TASK_ASSIGN_DATE)) FROM TASK group by TASK_ASSIGNED_TO"
        )
        task_summary = cursor.fetchall()
        conn.commit()
        conn.close()
        return task_summary

    def add_to_task_summary_table():
        task_summary = fetch_task_summary_data()
        task_summary_tree.delete(*task_summary_tree.get_children())
        for task in task_summary:
            task_summary_tree.insert("", END, values=task)

    task_summary_table_frame = customtkinter.CTkFrame(master=tab_4)
    task_summary_table_frame.grid(row=0, column=0, pady=20)

    task_summary_tree = ttk.Treeview(master=task_summary_table_frame, height=7)

    task_summary_verscrlbar = customtkinter.CTkScrollbar(master=task_summary_table_frame,
                                                         orientation="vertical", command=task_summary_tree.yview, )
    task_summary_verscrlbar.pack(side="right", fill="y")

    task_summary_tree.configure(yscrollcommand=task_summary_verscrlbar.set)
    task_summary_tree["columns"] = ("WORKERID", "AVEFINISHDAYS",)

    task_summary_tree.column("#0", width=0, stretch=tk.NO)
    task_summary_tree.column("WORKERID", anchor=tk.CENTER, width=200)
    task_summary_tree.column("AVEFINISHDAYS", anchor=tk.CENTER, width=120)

    task_summary_tree.heading("WORKERID", text="Worker Username")
    task_summary_tree.heading("AVEFINISHDAYS", text="  Average\nFinish (days)")

    task_summary_tree.pack(side="bottom", fill="both")

    add_to_task_summary_table()

    tab_4.columnconfigure(0, weight=2)
    tab_4.columnconfigure(1, weight=2)
    tab_4.columnconfigure(2, weight=2)

    incoming_stock_table_frame = customtkinter.CTkFrame(
        master=tab_5, width=1006, height=515
    )
    incoming_stock_table_frame.place(x=0, y=205)

    incoming_stock_tree = ttk.Treeview(master=incoming_stock_table_frame, height=18)

    incoming_stock_verscrlbar = customtkinter.CTkScrollbar(
        master=incoming_stock_table_frame,
        orientation="vertical",
        command=incoming_stock_tree.yview,
    )
    incoming_stock_verscrlbar.pack(side="right", fill="y")

    incoming_stock_tree.configure(yscrollcommand=incoming_stock_verscrlbar.set)
    incoming_stock_tree["columns"] = (
        "INCOMINGSTOCKID",
        "INCOMINGSTOCKNAME",
        "INCOMINGSTOCKQUANTITY",
        "INCOMINGSTOCKSTATUS",
    )

    incoming_stock_tree.column("#0", width=0, stretch=tk.NO)
    incoming_stock_tree.column("INCOMINGSTOCKID", anchor=tk.CENTER, width=310)
    incoming_stock_tree.column("INCOMINGSTOCKNAME", anchor=tk.CENTER, width=310)
    incoming_stock_tree.column("INCOMINGSTOCKQUANTITY", anchor=tk.CENTER, width=310)
    incoming_stock_tree.column("INCOMINGSTOCKSTATUS", anchor=tk.CENTER, width=320)

    incoming_stock_tree.heading("INCOMINGSTOCKID", text="ID")
    incoming_stock_tree.heading("INCOMINGSTOCKNAME", text="Name")
    incoming_stock_tree.heading("INCOMINGSTOCKQUANTITY", text="Quantity")
    incoming_stock_tree.heading("INCOMINGSTOCKSTATUS", text="Status")

    incoming_stock_tree.pack(side="bottom", fill="both")

    incoming_stock_tree.tag_configure(
        "To be Received", background="#FFBA00", foreground="black"
    )
    incoming_stock_tree.tag_configure(
        "Received", background="green", foreground="black"
    )

    incoming_stock_menu_frame = customtkinter.CTkFrame(master=tab_5, border_width=2)
    incoming_stock_menu_frame.grid(row=0, column=0, pady=10)

    edit_purchase_order_status_label = Inv_main_titlelabel(incoming_stock_menu_frame,0,0,
                                                           10,10,"Edit Purchase Order Status")
    purchase_order_status_label = customtkinter.CTkLabel(
        master=incoming_stock_menu_frame,
        text="Status:",
        font=customtkinter.CTkFont("SF Pro Display"),
    )
    purchase_order_status_entry = customtkinter.CTkComboBox(
        master=incoming_stock_menu_frame,
        values=["Received"],
        width=230,
        height=30,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )
    editpurchaseorder_btn = customtkinter.CTkButton(
        master=incoming_stock_menu_frame,
        text="Edit Status",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=update_purchase_order_status,
        compound="top",
        corner_radius=200,
        fg_color="#007FFF",
        text_color="black",
    )

    purchase_order_status_label.grid(row=1, column=0, padx=5, pady=5)
    purchase_order_status_entry.grid(row=1, column=1, padx=5, pady=5)
    editpurchaseorder_btn.grid(row=2, column=0, padx=5, pady=5, columnspan=2)

    search_incoming_stock_entry = customtkinter.CTkEntry(tab_5, placeholder_text="Search", width=1050, )
    search_incoming_stock_entry.grid(row=1, column=0, padx=10, pady=10)
    search_incoming_stock_entry.bind("<KeyRelease>", search_incoming_stock)

    tab_5.columnconfigure(0, weight=2)

    outgoing_stock_table_frame = customtkinter.CTkFrame(
        master=tab_6, width=1006, height=515
    )
    outgoing_stock_table_frame.place(x=0, y=205)
    outgoing_stock_tree = Inv_outgoing_stock_tree_display(outgoing_stock_table_frame, on_sale_order_double_click)

    outgoing_stock_menu_frame = customtkinter.CTkFrame(master=tab_6, border_width=2)
    outgoing_stock_menu_frame.grid(row=0, column=0, pady=10)

    edit_sales_order_status_label = Inv_main_titlelabel(outgoing_stock_menu_frame,0,0,
                                                        10,10,"Edit Sales Order Status")
    sales_order_status_label = customtkinter.CTkLabel(
        master=outgoing_stock_menu_frame,
        text="Status:",
        font=customtkinter.CTkFont("SF Pro Display"),
    )
    sales_order_status_entry = customtkinter.CTkComboBox(
        master=outgoing_stock_menu_frame,
        values=["To be Packed", "To be Shipped", "To be Delivered", "Delivered"],
        width=230,
        height=30,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )
    editsalesorder_btn = customtkinter.CTkButton(
        master=outgoing_stock_menu_frame,
        text="Edit Status",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=update_sales_order_status,
        compound="top",
        corner_radius=200,
        fg_color="#007FFF",
        text_color="black",
    )

    sales_order_status_label.grid(row=1, column=0, padx=5, pady=5)
    sales_order_status_entry.grid(row=1, column=1, padx=5, pady=5)
    editsalesorder_btn.grid(row=2, column=0, padx=5, pady=5, columnspan=2)

    search_outgoing_stock_entry = customtkinter.CTkEntry(tab_6, placeholder_text="Search", width=1050, )
    search_outgoing_stock_entry.grid(row=1, column=0, padx=10, pady=10)
    search_outgoing_stock_entry.bind("<KeyRelease>", search_outgoing_stock)

    outgoing_product_table_frame = customtkinter.CTkFrame(master=tab_6, width=506, height=515)
    outgoing_product_table_frame.place(x=740, y=205)

    outgoing_product_tree = Inv_outgoing_product_tree_display(outgoing_product_table_frame)
    add_to_outgoing_product_table()

    tab_6.columnconfigure(0, weight=2)

    add_to_customer_table()
    #add_to_supplier_table()
    add_to_product_table()
    add_to_task_table()
    add_to_incoming_stock_table()
    add_to_sale_order_table()

# =================================== Worker Dashboard ================================================
def worker_dashboard(username):
    product_dbase = Inv_Product_Database()  # declare the product class object.
    purchase_order_dbase = Inv_Purchase_Order_dbase()  # declare the purchase_order class object.

    def fetch_user_fullname():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,))
        user = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return user

    def fetch_to_be_packed_data():
        total_to_be_packed = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(SALE_ORDER_STATUS) FROM SALE_ORDER WHERE SALE_ORDER_STATUS = ?", ("To be Packed",))
        to_be_packed_data = cursor.fetchone()[0]
        total_to_be_packed += to_be_packed_data
        conn.commit()
        conn.close()
        return total_to_be_packed

    def update_to_be_packed_label():
        new_count = fetch_to_be_packed_data()
        total_to_be_packed_label_1.updateInfo(str(new_count))

    def fetch_to_be_shipped_data():
        total_to_be_shipped = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(SALE_ORDER_STATUS) FROM SALE_ORDER WHERE SALE_ORDER_STATUS = ?",
                       ("To be Shipped",))
        to_be_shipped_data = cursor.fetchone()[0]
        total_to_be_shipped += to_be_shipped_data
        conn.commit()
        conn.close()
        return total_to_be_shipped

    def update_to_be_shipped_label():
        new_count = fetch_to_be_shipped_data()
        total_to_be_shipped_label_1.updateInfo(str(new_count))

    def fetch_to_be_delivered_data():
        total_to_be_delivered = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(SALE_ORDER_STATUS) FROM SALE_ORDER WHERE SALE_ORDER_STATUS = ?",
                       ("To be Delivered",))
        to_be_delivered_data = cursor.fetchone()[0]
        total_to_be_delivered += to_be_delivered_data
        conn.commit()
        conn.close()
        return total_to_be_delivered

    def update_to_be_delivered_label():
        new_count = fetch_to_be_delivered_data()
        total_to_be_delivered_label_1.updateInfo(str(new_count))

    def fetch_total_quantity_in_hand_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT SUM(PRODUCT_QUANTITY) FROM PRODUCT")
        total_quantity_in_hand = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return total_quantity_in_hand

    def update_total_quantity_in_hand_label():
        new_count = fetch_total_quantity_in_hand_data()
        total_quantity_in_hand_1.updateInfo(str(new_count))

    def fetch_total_quantity_to_be_received_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT SUM(PURCHASE_ORDER_PRODUCT_QUANTITY) FROM PURCHASE_ORDER WHERE PURCHASE_ORDER_STATUS = ?",
            ("To be Received",))
        total_quantity_to_be_received = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return total_quantity_to_be_received

    def update_total_quantity_to_be_received_label():
        new_count = fetch_total_quantity_to_be_received_data()
        total_quantity_to_be_received_1.updateInfo(str(new_count))

    def fetch_low_stock_item_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(PRODUCT_ID) FROM PRODUCT WHERE PRODUCT_QUANTITY < PRODUCT_MIN_STOCK")
        low_stock_items = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return low_stock_items

    def update_low_stock_item_label():
        new_count = fetch_low_stock_item_data()
        low_stock_items_1.updateInfo(str(new_count))

    def fetch_total_items_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(PRODUCT_ID) FROM PRODUCT")
        total_items = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return total_items

    def update_total_items_label():
        new_count = fetch_total_items_data()
        all_items_1.updateInfo(str(new_count))

    def fetch_customer_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM CUSTOMER")
        customer = cursor.fetchall()
        conn.commit()
        conn.close()
        return customer

    def clear_customer_entry_field():
        customer_name_entry.clearField()
        customer_email_entry.clearField()
        customer_contact_entry.clearField()

    def display_customer_record(event):
        selected_item = customer_tree.focus()
        if selected_item:
            clear_customer_entry_field()
            row = customer_tree.item(selected_item)["values"]
            customer_name_entry.insertField(row[1])
            customer_email_entry.insertField(row[2])
            customer_contact_entry.insertField(row[3])

    def fetch_customer_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(CUSTOMER_ID) FROM CUSTOMER")
        last_customer_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_customer_id

    def generate_customer_id(prefix="CUST"):
        last_customer_id = fetch_customer_last_id()
        if last_customer_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_customer_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def add_to_customer_table():
        customers = fetch_customer_data()
        customer_tree.delete(*customer_tree.get_children())
        for customer in customers:
            customer_tree.insert("", END, values=customer)

    def search_customer(event):
        search_term = search_customer_entry.get().lower()
        customers = fetch_customer_data()
        customer_tree.delete(*customer_tree.get_children())
        for customer in customers:
            if search_term in str(customer).lower():
                customer_tree.insert("", tk.END, values=customer)

    def check_existing_customer(customer_check):
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT CUSTOMER_NAME FROM CUSTOMER WHERE CUSTOMER_NAME = ?",
            (customer_check,),
        )
        existing_customer = cursor.fetchone()
        conn.commit()
        conn.close()
        return existing_customer

    def is_valid_email(email):
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email)

    def is_valid_contact_number(contact_number):
        return contact_number.isdigit()

    def add_new_customer_details():
        customer_id = generate_customer_id("CUST")
        customer_name = customer_name_entry.getvalue()
        customer_email = customer_email_entry.getvalue()
        customer_contactno = customer_contact_entry.getvalue()

        if not (customer_name and customer_email and customer_contactno):
            messagebox.showerror("Error", "Please enter all fields")
            return

        if not is_valid_email(customer_email):
            messagebox.showerror("Error", "Please enter a valid email address")
            return

        if not is_valid_contact_number(customer_contactno):
            messagebox.showerror("Error", "Contact number must be only digits")
            return

        if not check_existing_customer(customer_name):
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()

            try:
                cursor.execute(
                    "INSERT INTO CUSTOMER (CUSTOMER_ID,CUSTOMER_NAME,CUSTOMER_EMAIL,CUSTOMER_TEL) \
                    VALUES (?,?,?,?)",
                    (customer_id, customer_name, customer_email, customer_contactno),
                )
                conn.commit()
                messagebox.showinfo("Success", "Data has been inserted")

                # Retrieve the user's full name
                cursor.execute(
                    "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
                )
                fullname = cursor.fetchone()

                # Log the user activity
                cursor.execute(
                    "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_user_activities_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "{} added new customer: {} with ID: {}".format(fullname[0], customer_name, customer_id),
                        username,
                    ),
                )
                conn.commit()

            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()

        else:
            messagebox.showerror(
                "Warning", "Duplicate customer! Please enter new customer."
            )

        add_to_customer_table()
        clear_customer_entry_field()
        return

    def edit_customer_details():
        selected_customer_details = customer_tree.focus()
        if not selected_customer_details:
            messagebox.showerror("Error", "Please select a record to edit")
            return

        row = customer_tree.item(selected_customer_details)["values"]
        new_customer_name = customer_name_entry.getvalue()
        new_customer_email = customer_email_entry.getvalue()
        new_customer_contactno = customer_contact_entry.getvalue()

        if not (new_customer_name and new_customer_email and new_customer_contactno):
            messagebox.showerror("Error", "Please enter all fields")
            return

        if not is_valid_email(new_customer_email):
            messagebox.showerror("Error", "Please enter a valid email address")
            return

        if not is_valid_contact_number(new_customer_contactno):
            messagebox.showerror("Error", "Contact number must be only digits")
            return

        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()

        old_customer_name = row[1]
        old_customer_email = row[2]
        old_customer_contactno = row[3]

        # Check if the new customer name is unique
        cursor.execute(
            "SELECT COUNT(*) FROM CUSTOMER WHERE CUSTOMER_NAME = ? AND CUSTOMER_NAME != ?",
            (new_customer_name, old_customer_name)
        )
        existing_customer_count = cursor.fetchone()[0]

        if existing_customer_count > 0:
            messagebox.showerror("Error", "Customer name must be unique")
            conn.close()
            return

        try:
            cursor.execute(
                "UPDATE CUSTOMER SET CUSTOMER_NAME = ?, CUSTOMER_EMAIL = ?, CUSTOMER_TEL = ? WHERE CUSTOMER_NAME = ?",
                (new_customer_name, new_customer_email, new_customer_contactno, old_customer_name),
            )
            conn.commit()

            # Retrieve the user's full name
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?",
                (username,)
            )
            fullname = cursor.fetchone()

            # Determine what was edited
            changes = []
            if old_customer_name != new_customer_name:
                changes.append(f"Name from '{old_customer_name}' to '{new_customer_name}'")
            if old_customer_email != new_customer_email:
                changes.append(f"Email from '{old_customer_email}' to '{new_customer_email}'")
            if old_customer_contactno != new_customer_contactno:
                changes.append(f"Contact number from '{old_customer_contactno}' to '{new_customer_contactno}'")

            changes_str = "; ".join(changes)

            # Log the user activity
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    f"{fullname[0]} edited customer {old_customer_name}: {changes_str}",
                    username,
                ),
            )
            conn.commit()

            messagebox.showinfo("Success", "Record successfully edited")

        except Exception as e:
            messagebox.showerror("Error", str(e))

        finally:
            conn.close()

        add_to_customer_table()
        clear_customer_entry_field()
        return

    def delete_customer_record():
        selected_item = customer_tree.focus()
        if not selected_item:
            messagebox.showerror("Error", "Please select a record to delete")
            return

        row = customer_tree.item(selected_item)["values"]
        customer_name = row[1]
        customer_id = row[0]

        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()

        try:
            cursor.execute("DELETE FROM CUSTOMER WHERE CUSTOMER_NAME= ?", (customer_name,))
            conn.commit()

            # Retrieve the user's full name
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
            )
            fullname = cursor.fetchone()

            # Log the user activity
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "{} deleted customer: {} with ID: {}".format(fullname[0], customer_name, customer_id),
                    username,
                ),
            )
            conn.commit()

            messagebox.showinfo("Success", "Record successfully deleted")

        except Exception as e:
            messagebox.showerror("Error", str(e))

        finally:
            conn.close()

        add_to_customer_table()
        clear_customer_entry_field()
        return

    def fetch_product_to_list():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT PRODUCT_NAME FROM PRODUCT")
        products = cursor.fetchall()
        product = []
        for x in products:
            product.append(x[0])
        conn.commit()
        conn.close()
        return product

    def on_customer_double_click(event):

        selected_item = customer_tree.selection()[0]
        values = customer_tree.item(selected_item, "values")

        def add_new_sale_order():
            sale_order_id = generate_sale_order_id("SO")
            sale_order_date = date.today().strftime("%m/%d/%Y")
            customer_name = values[1]

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO SALE_ORDER (SALE_ORDER_ID,SALE_ORDER_DATE,SALE_ORDER_CUSTOMER,SALE_ORDER_STATUS) \
            VALUES (?,?,?,?)",
                (sale_order_id, sale_order_date, customer_name, "To be Packed"),
            )
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Data has been inserted")
            add_to_sale_order_table()
            update_to_be_packed_label()
            update_to_be_shipped_label()
            update_to_be_delivered_label()
            update_total_quantity_in_hand_label()
            update_total_quantity_to_be_received_label()
            update_low_stock_item_label()
            update_total_items_label()
            return

        sales_window = customtkinter.CTk()
        sales_window.title("Product Details")
        sales_window.resizable(False, False)

        customer_id_label = Inv_Label(sales_window, 0, 0, (10, 5), "Customer ID:")
        customer_id_label_details = Inv_Label(sales_window, 0, 1, 5, f"{values[0]}")
        customer_name_label = Inv_Label(sales_window, 1, 0, (10, 5), "Customer Name:")
        customer_name_label_details = Inv_Label(sales_window, 1, 1, 5, f"{values[1]}")
        customer_email_label = Inv_Label(sales_window, 2, 0, (10, 5), "Customer Email:")
        customer_email_label_details = Inv_Label(sales_window, 2, 1, 5, f"{values[2]}")
        customer_contact_label = Inv_Label(sales_window, 3, 0, (10, 5), "Customer Contact No.:")
        customer_contact_label_details = Inv_Label(sales_window, 3, 1, 5, f"{values[3]}")

        sales_btn = Inv_Button(sales_window, 6, 0, 0, "#007FFF", "Create Sales Order", add_new_sale_order)

        sales_window.columnconfigure(0, weight=1)
        sales_window.columnconfigure(1, weight=1)
        sales_window.columnconfigure(2, weight=1)
        sales_window.columnconfigure(3, weight=1)
        sales_window.columnconfigure(4, weight=1)
        sales_window.columnconfigure(5, weight=1)
        sales_window.columnconfigure(6, weight=1)

        sales_window.mainloop()

    # ------------------------------------------------

    def add_to_product_table():
        products = product_dbase.fetch_product_data()
        admin_product_tree.add_to_table(products)

    def clear_product_entry_field():
        product_id_entry.clearField()
        product_name_entry.clearField()
        product_quantity_entry.clearField()
        product_description_entry.clearField()
        product_min_stock_entry.clearField()

    def display_product_record(event):
        selected_item = admin_product_tree.selected_item()
        if selected_item:
            clear_product_entry_field()
            row = selected_item
            product_id_entry.insertField([row[0]])
            product_name_entry.insertField(row[1])
            product_quantity_entry.insertField(row[2])
            product_description_entry.insertField(row[3])
            product_min_stock_entry.insertField([row[4]])

    def add_new_product_details():
        product_id = product_id_entry.getvalue()
        product_name = product_name_entry.getvalue()
        product_quantity = product_quantity_entry.getvalue()
        product_description = product_description_entry.getvalue()
        product_min_stock = product_min_stock_entry.getvalue()

        # Validate that product_id is not empty
        if not product_id:
            messagebox.showerror("Error", "Please enter product id")
            return

        # Validate that all fields are filled
        if not (product_name and product_quantity and product_description):
            messagebox.showerror("Error", "Please enter all fields")
            return

        # Validate that product_quantity is an integer
        if not product_quantity.isdigit() or int(product_quantity) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for product quantity")
            return

        # Validate that product_min_stock is an integer
        if not product_min_stock.isdigit() or int(product_min_stock) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for minimum stock")
            return

        product_quantity = int(product_quantity)
        product_min_stock = int(product_min_stock)

        # Check if product already exists
        if not product_dbase.check_existing_product(product_id):
            try:
                product_dbase.insertRecord(
                    product_id,
                    product_name,
                    product_quantity,
                    product_description,
                    product_min_stock,
                )

                conn = sqlite3.connect("Inventory Management System.db")
                cursor = conn.cursor()

                # Retrieve the user's full name
                cursor.execute(
                    "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
                )
                fullname = cursor.fetchone()

                # Log the user activity
                cursor.execute(
                    "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_user_activities_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "{} added new product: {} with ID: {}".format(fullname[0], product_name, product_id),
                        username,
                    ),
                )
                conn.commit()

                # Log the product movement
                cursor.execute(
                    "INSERT INTO PRODUCT_MOVEMENT (PRODUCT_MOVEMENT_ID, PRODUCT_MOVEMENT_DATE, PRODUCT_MOVEMENT, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_product_movement_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Added product: {} with ID: {} by {} to warehouse".format(product_name, product_id,
                                                                                  fullname[0]),
                        username,
                    ),
                )
                conn.commit()
                conn.close()

            except Exception as e:
                messagebox.showerror("Error", str(e))

        else:
            messagebox.showerror("Warning", "Duplicate product! Please enter new product.")

        add_to_product_table()
        clear_product_entry_field()
        fetch_product_to_list()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()
        low_stock_vs_total_item_pie_chart()
        bar_chart()
        return

    def edit_product_details():
        selected_product_details = admin_product_tree.selected_item()
        if not selected_product_details:
            messagebox.showerror("Error", "Please select a record to edit")
            return

        row = selected_product_details
        new_product_id = product_id_entry.getvalue()
        new_product_name = product_name_entry.getvalue()
        new_product_quantity = product_quantity_entry.getvalue()
        new_product_description = product_description_entry.getvalue()
        new_product_min_stock = product_min_stock_entry.getvalue()

        if not new_product_name:
            messagebox.showerror("Error", "Please enter all fields")
            return

        if not new_product_quantity.isdigit() or int(new_product_quantity) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for product quantity")
            return

        # Validate that product_min_stock is an integer
        if not new_product_min_stock.isdigit() or int(new_product_min_stock) < 0:
            messagebox.showerror("Error", "Please enter a valid non-negative integer for minimum stock")
            return

        old_product_id = row[0]
        old_product_name = row[1]
        old_product_quantity = row[2]
        old_product_description = row[3]
        old_product_min_stock = row[4]
        # old_preferred_supplier = row[5]

        try:
            product_dbase.updateRecord(
                new_product_id,
                new_product_name,
                new_product_quantity,
                new_product_description,
                new_product_min_stock,
                new_product_id)

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            # Retrieve the user's full name
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
            )
            fullname = cursor.fetchone()

            # Determine what was edited
            changes = []
            if old_product_id != new_product_id:
                changes.append(f"ID from '{old_product_id}' to '{new_product_id}'")
            if old_product_name != new_product_name:
                changes.append(f"Name from '{old_product_name}' to '{new_product_name}'")
            if old_product_quantity != new_product_quantity:
                changes.append(f"Quantity from '{old_product_quantity}' to '{new_product_quantity}'")
            if old_product_description != new_product_description:
                changes.append(f"Description from '{old_product_description}' to '{new_product_description}'")
            if old_product_min_stock != new_product_min_stock:
                changes.append(f"Min stock from '{old_product_min_stock}' to '{new_product_min_stock}'")
            # if old_preferred_supplier != new_preferred_supplier:
            #    changes.append(f"Preferred supplier from '{old_preferred_supplier}' to '{new_preferred_supplier}'")

            changes_str = "; ".join(changes)

            # Log the user activity
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    f"{fullname[0]} edited product {old_product_name}: {changes_str}",
                    username,
                ),
            )
            conn.commit()
            conn.close()

        except Exception as e:
            messagebox.showerror("Error", str(e))

        add_to_product_table()
        clear_product_entry_field()
        fetch_product_to_list()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()
        low_stock_vs_total_item_pie_chart()
        bar_chart()
        return

    def delete_product_record():
        selected_item = admin_product_tree.selected_item()
        if not selected_item:
            messagebox.showerror("Error", "Please select a record to delete")
            return
        row = selected_item
        product_id = row[0]
        product_name = row[1]

        # Ask for confirmation
        confirm = messagebox.askyesno("Confirm Delete",
                                      f"Are you sure you want to delete the product '{product_name}' with ID '{product_id}'?")
        if not confirm:
            return

        try:
            product_dbase.deleteRecord(product_id)

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            # Retrieve the user's full name
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
            )
            fullname = cursor.fetchone()

            # Log the user activity
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID, USER_ACTIVITIES_DATE, USER_ACTIVITIES, USER) \
                VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    f"{fullname[0]} deleted product: {product_name} with ID: {product_id}",
                    username,
                ),
            )
            conn.commit()
            conn.close()

        except Exception as e:
            messagebox.showerror("Error", str(e))

        add_to_product_table()
        clear_product_entry_field()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()
        low_stock_vs_total_item_pie_chart()
        bar_chart()
        return

    def search_product(event):
        search_term = search_entry.get().lower()
        products = product_dbase.fetch_product_data()
        admin_product_tree.search_item(search_term, products)

    def on_product_double_click(event):
        values = admin_product_tree.selected_item()

        new_window = customtkinter.CTk()
        new_window.title("Product Details")
        new_window.resizable(False, False)

        def add_incoming_stock():
            incoming_stock_id = generate_purchase_order_id("PO")
            incoming_stock_product = values[1]
            incoming_stock_quantity = restock_quantity_entry.getvalue()
            incoming_stock_status = "To be Received"

            if not incoming_stock_quantity:
                messagebox.showerror("Error", "Please enter all fields")
                return

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()

            try:
                cursor.execute(
                    "INSERT INTO PURCHASE_ORDER (PURCHASE_ORDER_ID, PURCHASE_ORDER_PRODUCT, PURCHASE_ORDER_PRODUCT_QUANTITY, PURCHASE_ORDER_STATUS) \
                    VALUES (?,?,?,?)",
                    (
                        incoming_stock_id,
                        incoming_stock_product,
                        incoming_stock_quantity,
                        incoming_stock_status,
                    ),
                )
                conn.commit()
                messagebox.showinfo("Success", "Data has been inserted")

                # Retrieve the user's full name
                cursor.execute(
                    "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
                )
                fullname = cursor.fetchone()

                # Log the product movement
                cursor.execute(
                    "INSERT INTO PRODUCT_MOVEMENT (PRODUCT_MOVEMENT_ID, PRODUCT_MOVEMENT_DATE, PRODUCT_MOVEMENT, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_product_movement_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "{} added incoming stock for product: {} with quantity: {}".format(fullname[0],
                                                                                           incoming_stock_product,
                                                                                           incoming_stock_quantity),
                        username,
                    ),
                )
                conn.commit()

            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()

            add_to_incoming_stock_table()
            fetch_tbr_quantity()
            update_total_quantity_in_hand_label()
            update_total_quantity_to_be_received_label()
            new_window.destroy()

        def fetch_tbr_quantity():
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT PURCHASE_ORDER_PRODUCT_QUANTITY FROM PURCHASE_ORDER WHERE PURCHASE_ORDER_PRODUCT = ? AND PURCHASE_ORDER_STATUS=?",
                (values[1], "To be Received"),
            )
            incoming_stock_quantity = cursor.fetchall()
            tbr = 0
            for x in incoming_stock_quantity:
                tbr += x[0]
            conn.commit()
            conn.close()
            return tbr

        def fetch_tbs_quantity():
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT SALE_ORDER_PRODUCT_QUANTITY FROM SALE_ORDER_PRODUCT WHERE SALE_ORDER_PRODUCT = ?",
                (values[1],),
            )
            outgoing_stock_quantity = cursor.fetchall()
            tbs = 0
            for x in outgoing_stock_quantity:
                tbs += x[0]
            conn.commit()
            conn.close()
            return tbs

        # ---------------------------------------------------
        product_details_frame = customtkinter.CTkFrame(new_window, border_width=2)
        product_details_frame.grid(row=0, column=0, sticky=W, padx=10, pady=10)

        restock_frame = customtkinter.CTkFrame(new_window, border_width=2)
        restock_frame.grid(row=1, column=0, padx=10, pady=10, columnspan=2)

        status_frame = customtkinter.CTkFrame(new_window, border_width=2)
        status_frame.grid(row=0, column=1, sticky=W, padx=10, pady=10)

        status_details_title = Inv_titlelabel(status_frame,0,0,10,"Status")

        total_tbs = customtkinter.CTkFrame(
            master=status_frame, width=100, height=100, fg_color="transparent"
        )
        total_tbs_details_1 = customtkinter.CTkLabel(
            total_tbs,
            text=(str(fetch_tbs_quantity()) + " Qty"),
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )
        total_tbs_details_2 = customtkinter.CTkLabel(
            total_tbs,
            text="To be Shipped",
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )
        vertical_separator = customtkinter.CTkFrame(
            status_frame, width=2, height=80, fg_color="#CCCCCC"
        )
        total_tbr = customtkinter.CTkFrame(
            master=status_frame, width=100, height=100, fg_color="transparent"
        )
        total_tbr_details_1 = customtkinter.CTkLabel(
            total_tbr,
            text=(str(fetch_tbr_quantity()) + " Qty"),
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )
        total_tbr_details_2 = customtkinter.CTkLabel(
            total_tbr,
            text="To be Received",
            font=customtkinter.CTkFont("SF Pro Display", size=13),
        )

        product_details_title = Inv_titlelabel(product_details_frame, 0, 0, 10, "Product Details")
        product_id_label = Inv_Label(product_details_frame, 1, 0, 5, "Product ID:")
        product_id_label_details = Inv_Label(product_details_frame, 1, 1, 5, f"{values[0]}")
        product_name_label = Inv_Label(product_details_frame, 2, 0, 5, "Product Name:")
        product_name_label_details = Inv_Label(product_details_frame, 2, 1, 5, f"{values[1]}")
        product_quantity_label = Inv_Label(product_details_frame, 3, 0, 5, "Product Quantity:")
        product_quantity_label_details = Inv_Label(product_details_frame, 3, 1, 5, f"{values[2]}")
        product_description_label = Inv_Label(product_details_frame, 4, 0, 5, "Product Description:")
        product_description_label_details = Inv_Label(product_details_frame, 4, 1, 5, f"{values[3]}")

        restock_title = Inv_titlelabel(restock_frame, 0, 0, 10, "Restock")
        restock_quantity_label = Inv_Label(restock_frame, 1, 0, 5, "Restock Quantity:")
        restock_quantity_entry = Inv_Entrybox(restock_frame, 1, 1, 5, "Quantity")
        restock_btn = Inv_Button(restock_frame, 2, 0, 5, "#007FFF", "Restock", add_incoming_stock)

        total_tbs.grid(row=1, column=0, rowspan=5, padx=15, pady=15)
        total_tbs_details_1.pack(side=TOP)
        total_tbs_details_2.pack(side=BOTTOM)
        vertical_separator.grid(row=1, column=1, rowspan=5, pady=36)
        total_tbr.grid(row=1, column=3, rowspan=5, padx=15, pady=15)
        total_tbr_details_1.pack(side=TOP)
        total_tbr_details_2.pack(side=BOTTOM)

        new_window.mainloop()

    def fetch_task_data_worker():
        conn = sqlite3.connect("Inventory Management System.db")
        # sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT TASK_ID, TASK_ASSIGN_DATE, TASK_DESCRIPTION,TASK_STATUS,TASK_DUE_DATE, TASK_FINISH_DATE FROM TASK WHERE TASK_ASSIGNED_TO = ?",
            (username,),
        )
        task = cursor.fetchall()
        conn.commit()
        conn.close()
        return task

    def add_to_task_table():
        tasks = fetch_task_data_worker()
        task_tree.delete(*task_tree.get_children())
        for task in tasks:
            task_tree.insert("", END, values=task, tags=task)

    def check_task_status():
        task_tree.tag_configure("Pending", background="#FFBA00", foreground="black")
        task_tree.tag_configure("Completed", background="green", foreground="black")
        task_tree.tag_configure("Due, Not Completed", background="#FF7F7F", foreground="black")
        task_tree.tag_configure("Due, Completed", background="#FF7F7F", foreground="black")

    def days_between(d1, d2):
        d1 = datetime.strptime(d1, "%Y-%m-%d")
        d2 = datetime.strptime(d2, "%Y-%m-%d")
        return (d2 - d1).days

    def check_task_due_date():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
                "SELECT TASK_DUE_DATE, TASK_ID, TASK_STATUS FROM TASK WHERE TASK_ASSIGNED_TO = ?",
                (username, ),
        )
        tasks = cursor.fetchall()
        conn.commit()
        conn.close()

        today = str(datetime.today().date())

        due_task = 0
        for task in tasks:
            due_date_str = task[0]
            due_days = days_between(today, due_date_str)

            if due_days < 0:
                due_task += 1
                conn = sqlite3.connect("Inventory Management System.db")
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE TASK SET TASK_STATUS = ? WHERE TASK_ID = ?",
                    ("Due, Not Completed", task[1]),
                )
                conn.commit()
                conn.close()
            elif task[2] != 'Completed':
                conn = sqlite3.connect("Inventory Management System.db")
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE TASK SET TASK_STATUS = ? WHERE TASK_ID = ?",
                    ("Pending", task[1]),
                )
                conn.commit()
                conn.close()

        check_task_status()
        add_to_task_table()
        return due_task

    def complete_task():
        selected_task = task_tree.focus()
        if not selected_task:
            messagebox.showerror("Error", "Please select a record to edit")
            return
        row = task_tree.item(selected_task)["values"]
        task_finish_date = str(datetime.today().date())

        if row[3] == "Completed":
            messagebox.showerror("Error", "Task already completed")
            return

        if row[3] == "Pending":
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE TASK SET TASK_STATUS = ?, TASK_FINISH_DATE = ? WHERE TASK_ID = ?",
                ("Completed", task_finish_date, row[0]),
            )
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Record successfully edited")
            check_task_status()
            add_to_task_table()
            update_completed_task_label()
            update_pending_task_label()
            update_overdue_task_label()
            return

        if row[3] == "Due, Not Completed":
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE TASK SET TASK_STATUS = ?, TASK_FINISH_DATE = ? WHERE TASK_ID = ?",
                ("Due, Completed", task_finish_date, row[0]),
            )
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Record successfully edited")
            check_task_status()
            add_to_task_table()
            update_completed_task_label()
            update_pending_task_label()
            update_overdue_task_label()
            return

    def fetch_incoming_stock_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM PURCHASE_ORDER")
        incoming_stock = cursor.fetchall()
        conn.commit()
        conn.close()
        return incoming_stock

    def add_to_incoming_stock_table():
        incoming_stocks = fetch_incoming_stock_data()
        incoming_stock_tree.delete(*incoming_stock_tree.get_children())
        for incoming_stock in incoming_stocks:
            incoming_stock_tree.insert(
                "", END, values=incoming_stock, tags=incoming_stock
            )

    def search_incoming_stock(event):
        search_term = search_incoming_stock_entry.get().lower()
        incoming_stocks = fetch_incoming_stock_data()
        incoming_stock_tree.delete(*incoming_stock_tree.get_children())
        for incoming_stock in incoming_stocks:
            if search_term in str(incoming_stock).lower():
                incoming_stock_tree.insert("", tk.END, values=incoming_stock, tags=incoming_stock)

    def update_purchase_order_status():
        new_status = purchase_order_status_entry.get()
        selected = incoming_stock_tree.focus()
        if not selected:
            messagebox.showerror("Error", "Please select a record to edit")
            return

        row = incoming_stock_tree.item(selected)["values"]
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()

        try:
            cursor.execute(
                "UPDATE PURCHASE_ORDER SET PURCHASE_ORDER_STATUS = ? WHERE PURCHASE_ORDER_ID = ?",
                (new_status, row[0]),
            )

            if new_status == "Received":
                cursor.execute(
                    "UPDATE PRODUCT SET PRODUCT_QUANTITY = PRODUCT_QUANTITY + ? WHERE PRODUCT_NAME = ?",
                    (row[2], row[1]),
                )

                # Retrieve the user's full name
                cursor.execute(
                    "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
                )
                fullname = cursor.fetchone()

                # Log the product movement only when the status is "Received"
                cursor.execute(
                    "INSERT INTO PRODUCT_MOVEMENT (PRODUCT_MOVEMENT_ID, PRODUCT_MOVEMENT_DATE, PRODUCT_MOVEMENT, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_product_movement_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "{} received product: {} (ID: {}) with quantity: {}".format(fullname[0], row[1], row[0],
                                                                                    row[2]),
                        username,
                    ),
                )

            conn.commit()
            messagebox.showinfo("Success", "Record successfully edited")

        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            conn.close()

        add_to_incoming_stock_table()
        add_to_product_table()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()
        low_stock_vs_total_item_pie_chart(canvas)
        bar_chart(canvas1)

    def check_pending_task():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT TASK_STATUS FROM TASK WHERE TASK_ASSIGNED_TO = ?", (username,)
        )
        list_of_task_status = cursor.fetchall()
        conn.commit()
        conn.close()
        pending_task = 0
        for x in list_of_task_status:
            if x[0] == "Pending":
                pending_task += 1
        if pending_task >= 1:
            toast = Notification(
                app_id="IMS",
                title="Pending Task",
                msg=f"There are {pending_task} pending task",
                duration="short",
            )
            toast.set_audio(audio.Default, loop=False)
            toast.show()
        return list_of_task_status

    def fetch_completed_task_data():
        completed_task = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(TASK_STATUS) FROM TASK WHERE (TASK_STATUS = ?) and (TASK_ASSIGNED_TO = ?)", ("Completed",username))
        completed_task_data = cursor.fetchone()[0]
        completed_task += completed_task_data
        conn.commit()
        conn.close()
        return completed_task

    def update_completed_task_label():
        new_count = fetch_completed_task_data()
        completed_tasks_label_1.configure(text=str(new_count))

    def fetch_pending_task_data():
        pending_task = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(TASK_STATUS) FROM TASK WHERE (TASK_STATUS = ?) and (TASK_ASSIGNED_TO = ?)", ("Pending",username))
        pending_task_data = cursor.fetchone()[0]
        pending_task += pending_task_data
        conn.commit()
        conn.close()
        return pending_task

    def update_pending_task_label():
        new_count = fetch_pending_task_data()
        pending_tasks_label_1.configure(text=str(new_count))

    def fetch_overdue_task_data():
        overdue_task = 0
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(TASK_STATUS) FROM TASK WHERE (TASK_STATUS = ?) and (TASK_ASSIGNED_TO = ?)", ("Due, Not Completed",username))
        overdue_task_data1 = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(TASK_STATUS) FROM TASK WHERE (TASK_STATUS = ?) and (TASK_ASSIGNED_TO = ?)", ("Due, Completed",username))
        overdue_task_data2 = cursor.fetchone()[0]
        overdue_task += overdue_task_data2 + overdue_task_data1
        conn.commit()
        conn.close()
        return overdue_task

    def update_overdue_task_label():
        new_count = fetch_overdue_task_data()
        overdue_tasks_label_1.configure(text=str(new_count))

    def fetch_purchase_order_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(PURCHASE_ORDER_ID) FROM PURCHASE_ORDER")
        last_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_id

    def generate_purchase_order_id(prefix="PO"):
        last_id = fetch_purchase_order_last_id()
        if last_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def update_sales_order_status():
        new_status = sales_order_status_entry.get()
        selected = outgoing_stock_tree.focus()
        if not selected:
            messagebox.showerror("Error", "Please select a record to edit")
            return
        row = outgoing_stock_tree.item(selected)["values"]
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE SALE_ORDER SET SALE_ORDER_STATUS = ? WHERE SALE_ORDER_ID = ?",
            (new_status, row[0]),
        )
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Record successfully edited")
        add_to_sale_order_table()
        add_to_product_table()
        update_to_be_packed_label()
        update_to_be_shipped_label()
        update_to_be_delivered_label()
        update_total_quantity_in_hand_label()
        update_total_quantity_to_be_received_label()
        update_low_stock_item_label()
        update_total_items_label()

    def fetch_sale_order_stock_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM SALE_ORDER")
        sale_order = cursor.fetchall()
        conn.commit()
        conn.close()
        return sale_order

    def add_to_sale_order_table():
        outgoing_stocks = fetch_sale_order_stock_data()
        outgoing_stock_tree.add_to_table(outgoing_stocks)

    def search_outgoing_stock(event):
        search_term = search_outgoing_stock_entry.get().lower()
        outgoing_stocks = fetch_sale_order_stock_data()
        outgoing_stock_tree.search_item(search_term, outgoing_stocks)

    def fetch_sale_order_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(SALE_ORDER_ID) FROM SALE_ORDER")
        last_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_id

    def generate_sale_order_id(prefix="SO"):
        last_id = fetch_sale_order_last_id()
        if last_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def on_sale_order_double_click(event):
        values = outgoing_stock_tree.selected_item()

        sale_order_window = customtkinter.CTk()
        sale_order_window.title("Product Details")
        sale_order_window.resizable(False, False)

        product_selection_frame = customtkinter.CTkFrame(
            master=sale_order_window, border_width=2
        )
        sale_order_product_table_frame = customtkinter.CTkFrame(
            master=sale_order_window
        )

        def fetch_sale_order_status(sale_order_id):
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute("SELECT SALE_ORDER_STATUS FROM SALE_ORDER WHERE SALE_ORDER_ID = ?", (sale_order_id,))
            status = cursor.fetchone()
            conn.close()
            return status[0] if status else None

        def fetch_sale_order_product_data():
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT * FROM SALE_ORDER_PRODUCT WHERE SALE_ORDER_ID = ?", (values[0],)
            )
            sale_order = cursor.fetchall()
            conn.commit()
            conn.close()
            return sale_order

        def add_to_sale_order_product_table():
            sale_order_products = fetch_sale_order_product_data()
            sale_order_product_tree.delete(*sale_order_product_tree.get_children())
            for sale_order_product in sale_order_products:
                sale_order_product_tree.insert("", END, values=sale_order_product)

        def update_product_details(choice):
            selected_product_id = product_selection_entry.get()

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT PRODUCT_NAME,PRODUCT_STATUS FROM PRODUCT WHERE PRODUCT_ID = ?", (selected_product_id,)
            )
            choice = cursor.fetchall()
            conn.commit()
            conn.close()

            product_selection_name_label2.configure(text=choice[0][0])
            product_selection_description_label2.configure(text=choice[0][1])


        def add_sale_order_product():
            sale_order_id = values[0]
            sale_order_product_id = product_selection_entry.get()
            sale_order_product_quantity = sales_quantity_entry.get()

            if not sale_order_product_quantity:
                messagebox.showerror("Error", "Please enter all fields")
                return

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()

            # Check the status of the sale order
            cursor.execute(
                "SELECT SALE_ORDER_STATUS FROM SALE_ORDER WHERE SALE_ORDER_ID = ?", (sale_order_id,)
            )
            sale_order_status = cursor.fetchone()[0]

            if sale_order_status != "To be Packed":
                messagebox.showerror("Error", "Cannot add items. The status is not 'To be Packed'")
                conn.close()
                return

            # Get Product Name
            cursor.execute(
                "SELECT PRODUCT_NAME FROM PRODUCT WHERE PRODUCT_ID = ?", (sale_order_product_id,)
            )
            product_name = cursor.fetchone()[0]

            cursor.execute(
                "SELECT PRODUCT_QUANTITY FROM PRODUCT WHERE PRODUCT_ID = ?", (sale_order_product_id,),
            )
            product_quantity = cursor.fetchone()[0]

            if product_quantity < int(sale_order_product_quantity):
                messagebox.showerror("Error", "Insufficient quantity")
                conn.close()
                return

            try:
                cursor.execute(
                    "INSERT INTO SALE_ORDER_PRODUCT (SALE_ORDER_ID,SALE_ORDER_PRODUCT_ID, SALE_ORDER_PRODUCT, SALE_ORDER_PRODUCT_QUANTITY) \
                    VALUES (?,?,?,?)",
                    (sale_order_id, sale_order_product_id, product_name, sale_order_product_quantity),
                )
                cursor.execute(
                    "UPDATE PRODUCT SET PRODUCT_QUANTITY = PRODUCT_QUANTITY - ? WHERE PRODUCT_ID = ?",
                    (sale_order_product_quantity, sale_order_product_id),
                )

                # Retrieve the user's full name
                cursor.execute(
                    "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
                )
                fullname = cursor.fetchone()

                # Log the product movement
                cursor.execute(
                    "SELECT PRODUCT_ID FROM PRODUCT WHERE PRODUCT_NAME = ?", (product_name,)
                )
                product_id = cursor.fetchone()[0]
                cursor.execute(
                    "INSERT INTO PRODUCT_MOVEMENT (PRODUCT_MOVEMENT_ID, PRODUCT_MOVEMENT_DATE, PRODUCT_MOVEMENT, USER) \
                    VALUES (?,?,?,?)",
                    (
                        generate_product_movement_id(),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "{} added product: {} (ID: {}) with quantity: {} into sale order (ID: {})".format(fullname[0],
                                                                                                          product_name,
                                                                                                          sale_order_product_id,
                                                                                                          sale_order_product_quantity,
                                                                                                          sale_order_id),
                        username,
                    ),
                )

                conn.commit()
                messagebox.showinfo("Success", "Data has been inserted")

            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()

            add_to_sale_order_product_table()
            add_to_product_table()
            update_low_stock_item_label()
            add_to_outgoing_product_table()
            return

        select_product_label = Inv_titlelabel(product_selection_frame,0,0,10,"Select Product")

        product_selection_label = customtkinter.CTkLabel(
            product_selection_frame,
            text="Product to be purchase:",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        product_selection_entry = customtkinter.CTkComboBox(
            master=product_selection_frame,
            values=fetch_product_to_list(),
            width=230,
            height=30,
            font=customtkinter.CTkFont("SF Pro Display"),
            command=update_product_details,
            border_width=0,
        )
        product_selection_name_label1 = customtkinter.CTkLabel(
            product_selection_frame,
            text="Product Name:",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        product_selection_name_label2 = customtkinter.CTkLabel(
            product_selection_frame,
            text="",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        product_selection_description_label1 = customtkinter.CTkLabel(
            product_selection_frame,
            text="Product Description:",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        product_selection_description_label2 = customtkinter.CTkLabel(
            product_selection_frame,
            text="",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        sales_quantity_label = customtkinter.CTkLabel(
            product_selection_frame,
            text="Purchase Quantity:",
            font=customtkinter.CTkFont("SF Pro Display"),
        )
        sales_quantity_entry = customtkinter.CTkEntry(
            master=product_selection_frame,
            placeholder_text="Quantity",
            width=230,
            height=30,
            font=customtkinter.CTkFont("SF Pro Display"),
            border_width=0,
        )
        sales_btn = customtkinter.CTkButton(
            master=product_selection_frame,
            text="Add Item",
            command=add_sale_order_product,
            font=customtkinter.CTkFont("SF Pro Display"),
            corner_radius=200,
            fg_color="#007FFF",
            text_color="black",
        )
        # --------------------------------------------------------

        style1 = ThemedStyle()
        style1.set_theme("equilux")
        style1.configure(
            "Treeview",
            font=("SF Pro Display", 10),
            rowheight=24,
            foreground="white",
            background="#2a2d2e",
            fieldbackground="#343638",
            bordercolor="#343638",
            borderwidth=0,
        )
        style1.map("Treeview", background=[("selected", "#22559b")])
        style1.configure(
            "Treeview.Heading", background="#565b5e", foreground="white", relief="flat"
        )
        style1.map("Treeview.Heading", background=[("active", "#3484F0")])
        style1.configure(
            "DateEntry",
            font=customtkinter.CTkFont("SF Pro Display", 12),
            foreground="white",
            background="#2a2d2e",
            fieldbackground="#343638",
            bordercolor="#343638",
            borderwidth=0,
        )

        sale_order_product_tree = ttk.Treeview(
            master=sale_order_product_table_frame, height=30
        )

        sale_order_product_verscrlbar = customtkinter.CTkScrollbar(
            master=sale_order_product_table_frame,
            orientation="vertical",
            command=sale_order_product_tree.yview,
        )
        sale_order_product_verscrlbar.pack(side="right", fill="y")

        sale_order_product_tree.configure(
            yscrollcommand=sale_order_product_verscrlbar.set
        )
        sale_order_product_tree["columns"] = (
            "SALEORDERID",
            "SALEORDERPRODUCTID",
            "SALEORDERPRODUCT",
            "SALEORDERPRODUCTQUANTITY",
        )

        sale_order_product_tree.column("#0", width=0, stretch=tk.NO)
        sale_order_product_tree.column("SALEORDERID", anchor=tk.CENTER, width=150)
        sale_order_product_tree.column("SALEORDERPRODUCTID", anchor=tk.CENTER, width=150)
        sale_order_product_tree.column("SALEORDERPRODUCT", anchor=tk.CENTER, width=150)
        sale_order_product_tree.column("SALEORDERPRODUCTQUANTITY", anchor=tk.CENTER, width=150)

        sale_order_product_tree.heading("SALEORDERID", text="Sales Order ID")
        sale_order_product_tree.heading("SALEORDERPRODUCTID", text="Item ID")
        sale_order_product_tree.heading("SALEORDERPRODUCT", text="Item")
        sale_order_product_tree.heading("SALEORDERPRODUCTQUANTITY", text="Quantity")

        sale_order_product_tree.pack(side="bottom", fill="both")

        product_selection_frame.grid(row=0, column=0, padx=10, pady=10, sticky=NW)
        product_selection_label.grid(row=1, column=0, padx=10, pady=10)
        product_selection_entry.grid(row=1, column=1, padx=10, pady=10)
        product_selection_name_label1.grid(row=2, column=0, padx=10, pady=10)
        product_selection_name_label2.grid(row=2, column=1, padx=10, pady=10)
        product_selection_description_label1.grid(row=3, column=0, padx=10, pady=10)
        product_selection_description_label2.grid(row=3, column=1, padx=10, pady=10)
        sales_quantity_label.grid(row=4, column=0, padx=10, pady=10)
        sales_quantity_entry.grid(row=4, column=1, padx=10, pady=10)
        sales_btn.grid(row=5, column=0, padx=10, pady=10, columnspan=2)
        sale_order_product_table_frame.grid(row=0, column=1, padx=10, pady=10)

        add_to_sale_order_product_table()
        sale_order_window.mainloop()

    def fetch_outgoing_product_data():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT SALE_ORDER_PRODUCT_ID, SALE_ORDER_PRODUCT, SUM(SALE_ORDER_PRODUCT_QUANTITY) FROM SALE_ORDER_PRODUCT \
              GROUP BY SALE_ORDER_PRODUCT_ID"
        )
        sale_order = cursor.fetchall()
        conn.commit()
        conn.close()
        return sale_order

    def add_to_outgoing_product_table():
        outgoing_stocks = fetch_outgoing_product_data()
        outgoing_product_tree.add_to_table(outgoing_stocks)

    def low_stock_vs_total_item_pie_chart():
        low_stock_items = product_dbase.fetch_low_stock_item_data()
        total_items = product_dbase.fetch_total_items_data()
        piechart.display(total_items, low_stock_items)

    def bar_chart():
        barchart_data = product_dbase.fetch_bar_chart_data()
        barchart.display(barchart_data)

    def fetch_user_activities_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(USER_ACTIVITIES_ID) FROM USER_ACTIVITIES")
        last_user_activities_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_user_activities_id

    def generate_user_activities_id(prefix="UA"):
        last_customer_id = fetch_user_activities_last_id()
        if last_customer_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(last_customer_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def fetch_product_movement_last_id():
        conn = sqlite3.connect("Inventory Management System.db")
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(PRODUCT_MOVEMENT_ID) FROM PRODUCT_MOVEMENT")
        last_product_movement_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return last_product_movement_id

    def generate_product_movement_id(prefix="PM"):
        product_movement_id = fetch_product_movement_last_id()
        if product_movement_id is None:
            return f"{prefix}-001"
        else:
            number_part = str(product_movement_id.split("-")[-1])
            new_number = int(number_part) + 1
            return f"{prefix}-{new_number:03d}"

    def logout():
        confirmation = messagebox.askyesno('Are you sure?', 'Are you sure that you want to logout?')
        if confirmation:
            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "SELECT USER_FULLNAME FROM USER WHERE USERNAME = ?", (username,)
            )
            fullname = cursor.fetchone()
            conn.commit()
            conn.close()

            conn = sqlite3.connect("Inventory Management System.db")
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO USER_ACTIVITIES (USER_ACTIVITIES_ID,USER_ACTIVITIES_DATE,USER_ACTIVITIES,USER) \
            VALUES (?,?,?,?)",
                (
                    generate_user_activities_id(),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "{} logged out".format(fullname[0]),
                    username,
                ),
            )
            conn.commit()
            conn.close()

            worker_dashboard_frame.destroy()
            login_page()

    def sort_treeview_column(tv, col, reverse):
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(reverse=reverse)

        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)

        tv.heading(col, command=lambda: sort_treeview_column(tv, col, not reverse))

    worker_dashboard_frame = customtkinter.CTkFrame(
        master=app, width=1280, height=720, border_width=0
    )
    worker_dashboard_frame.place(x=0, y=0)

    info_frame = customtkinter.CTkFrame(
        master=worker_dashboard_frame, width=1280, height=720
    )
    info_frame.place(x=0, y=0)
    info_tab = customtkinter.CTkTabview(master=info_frame, width=1280, height=720)
    custom_font = ("SF Pro Display", 15)
    info_tab._segmented_button.configure(font=custom_font, fg_color='grey', text_color='white')
    info_tab.place(x=0, y=0)
    tab_0 = info_tab.add(" Homepage ")
    tab_1 = info_tab.add(" Customer ")
    #tab_2 = info_tab.add("Supplier")
    tab_3 = info_tab.add(" Product ")
    tab_4 = info_tab.add(" Task ")
    tab_5 = info_tab.add(" Incoming Stock ")
    tab_6 = info_tab.add(" Outgoing Stock ")

# -----------------------------
    outgoing_table_title1_label = customtkinter.CTkLabel(master=tab_6, text_color="purple",
                                                         text='Double Click on the selected Sales Order below to enter Product & Quantity',
                                                         font=customtkinter.CTkFont("Times", 13, weight='bold'))
    outgoing_table_title1_label.place(x=100, y=180)

    outgoing_table_title2_label = customtkinter.CTkLabel(master=tab_6, text_color="darkgreen",
                                                         text='Summary of Total Outgoing Quantity by Product',
                                                         font=customtkinter.CTkFont("Times", 13, weight='bold'))
    outgoing_table_title2_label.place(x=850, y=180)
    customer_table_title1_label = customtkinter.CTkLabel(master=tab_1, text_color="purple",
                                                         text='Double Click on the selected Customer below to create Sales Order',
                                                         font=customtkinter.CTkFont("Times", 15, weight='bold'))
    customer_table_title1_label.place(x=50, y=180)
    product_table_title1_label = customtkinter.CTkLabel(master=tab_3, text_color="purple",
                                                         text='Double Click on the selected Product below to Restock the Quantity',
                                                         font=customtkinter.CTkFont("Times", 15, weight='bold'))
    product_table_title1_label.place(x=50, y=180)

    # --------------------------------

    style = ThemedStyle()
    style.set_theme("equilux")
    style.configure(
        "Treeview",
        font=("SF Pro Display", 10),
        rowheight=24,
        foreground="white",
        background="#2a2d2e",
        fieldbackground="#343638",
        bordercolor="#343638",
        borderwidth=0,
    )
    style.map("Treeview", background=[("selected", "#22559b")])
    style.configure(
        "Treeview.Heading", background="#565b5e", foreground="white", relief="flat"
    )
    style.map("Treeview.Heading", background=[("active", "#3484F0")])
    style.configure(
        "DateEntry",
        font=customtkinter.CTkFont("SF Pro Display", 12),
        foreground="white",
        background="#2a2d2e",
        fieldbackground="#343638",
        bordercolor="#343638",
        borderwidth=0,
    )

    welcome_label = customtkinter.CTkLabel(
        master=tab_0,
        text=f"Hello, {fetch_user_fullname()} ",
        font=customtkinter.CTkFont("SF Pro Display", 24, weight="bold"),
    )

    sales_activity_frame = customtkinter.CTkFrame(
        tab_0, corner_radius=10, border_width=2
    )

    sales_activity_vertical_separator_1 = customtkinter.CTkFrame(
        sales_activity_frame, width=2, height=50, fg_color="#CCCCCC"
    )

    sales_activity_vertical_separator_2 = customtkinter.CTkFrame(
        sales_activity_frame, width=2, height=50, fg_color="#CCCCCC"
    )
    inventory_summary_frame = customtkinter.CTkFrame(
        tab_0, corner_radius=10, border_width=2
    )

    sales_activity_vertical_separator_4 = customtkinter.CTkFrame(
        inventory_summary_frame, width=2, height=50, fg_color="#CCCCCC"
    )
    product_details_frame = customtkinter.CTkFrame(
        tab_0, corner_radius=10, border_width=2
    )
    product_details_title = Inv_titlelabel(product_details_frame,0,0,10,"Product Details")

    sales_activity_vertical_separator_6 = customtkinter.CTkFrame(
        product_details_frame, width=2, height=50, fg_color="#CCCCCC"
    )

    logout_btn = Inv_Button(tab_0, 0, 2, 10, "#ff666d", "Logout", logout)

    welcome_label.grid(row=0, column=0, sticky=W, padx=10, pady=10)
    sales_activity_frame.grid(row=1, column=0, padx=10, pady=10)
    inventory_summary_frame.grid(row=1, column=1, padx=10, pady=10)
    product_details_frame.grid(row=1, column=2, padx=10, pady=10)

    piechart = Inv_pie_chart(tab_0, 2, 0)
    barchart = Inv_bar_chart(tab_0, 2, 1)
    low_stock_vs_total_item_pie_chart()
    bar_chart()

    sales_activity_title = Inv_titlelabel(sales_activity_frame, 0, 0, 10, "Sales Activity")

    total_to_be_packed_label_1 = Inv_Label(sales_activity_frame, 1, 0, 20, str(fetch_to_be_packed_data()))
    total_to_be_packed_label_2 = Inv_Label(sales_activity_frame, 2, 0, 20, "To be Packed")
    sales_activity_vertical_separator_1.grid(row=1, column=1, rowspan=2, padx=5, pady=5)
    total_to_be_shipped_label_1 = Inv_Label(sales_activity_frame, 1, 2, 20, str(fetch_to_be_shipped_data()))
    total_to_be_shipped_label_2 = Inv_Label(sales_activity_frame, 2, 2, 20, "To be Shipped")
    sales_activity_vertical_separator_2.grid(row=1, column=3, rowspan=2, padx=5, pady=5)
    total_to_be_delivered_label_1 = Inv_Label(sales_activity_frame, 1, 4, 20, str(fetch_to_be_delivered_data()))
    total_to_be_delivered_label_2 = Inv_Label(sales_activity_frame, 2, 4, 20, "To be Delivered")

    inventory_summary_title = Inv_titlelabel(inventory_summary_frame, 0, 0, 10, "Inventory Summary")
    total_quantity_in_hand_1 = Inv_Label(inventory_summary_frame, 1, 0, 20,
                                         str(product_dbase.fetch_total_quantity_in_hand_data()))
    total_quantity_in_hand_2 = Inv_Label(inventory_summary_frame, 2, 0, 20, "Quantity In Hand")
    sales_activity_vertical_separator_4.grid(row=1, column=1, rowspan=2, padx=5, pady=5)
    total_quantity_to_be_received_1 = Inv_Label(inventory_summary_frame, 1, 2, 20,
                                                str(fetch_total_quantity_to_be_received_data()))
    total_quantity_to_be_received_2 = Inv_Label(inventory_summary_frame, 2, 2, 20, "Quantity To be Received")

    low_stock_items_1 = Inv_Label(product_details_frame, 1, 0, 20, str(product_dbase.fetch_low_stock_item_data()))
    low_stock_items_2 = Inv_Label(product_details_frame, 2, 0, 20, "Low Stock Items")
    sales_activity_vertical_separator_6.grid(row=1, column=1, rowspan=2, padx=5, pady=5)
    all_items_1 = Inv_Label(product_details_frame, 1, 2, 20, str(product_dbase.fetch_total_items_data()))
    all_items_2 = Inv_Label(product_details_frame, 2, 2, 20, "Total Items")

    tab_0.columnconfigure(1, weight=1)

    customer_table_frame = customtkinter.CTkFrame(master=tab_1, width=1280, height=515)
    customer_table_frame.place(x=0, y=205)

    customer_tree = ttk.Treeview(master=customer_table_frame, height=18)

    customer_verscrlbar = customtkinter.CTkScrollbar(
        master=customer_table_frame, orientation="vertical", command=customer_tree.yview
    )
    customer_verscrlbar.pack(side="right", fill="y")

    customer_tree.configure(yscrollcommand=customer_verscrlbar.set)
    customer_tree["columns"] = (
        "CUSTOMERID",
        "CUSTOMERNAME",
        "CUSOTMEREMAILADDRESS",
        "CUSTOMERCONTACTNO",
    )

    customer_tree.column("#0", width=0, stretch=tk.NO)
    customer_tree.column("CUSTOMERID", anchor=tk.CENTER, width=313)
    customer_tree.column("CUSTOMERNAME", anchor=tk.CENTER, width=313)
    customer_tree.column("CUSOTMEREMAILADDRESS", anchor=tk.CENTER, width=313)
    customer_tree.column("CUSTOMERCONTACTNO", anchor=tk.CENTER, width=313)

    customer_tree.heading("CUSTOMERID", text="ID")
    customer_tree.heading("CUSTOMERNAME", text="Name")
    customer_tree.heading("CUSOTMEREMAILADDRESS", text="Email Address")
    customer_tree.heading("CUSTOMERCONTACTNO", text="Contact No.")

    customer_tree.pack(side="right", fill="both")
    customer_tree.bind("<Double-1>", on_customer_double_click)
    customer_tree.bind("<ButtonRelease>", display_customer_record)

    # -------------------------------------------------------
    customer_menu_frame = customtkinter.CTkFrame(master=tab_1, border_width=2)
    customer_menu_frame.grid(row=0, column=0)
    insert_customer_data_label = Inv_main_titlelabel(customer_menu_frame, 0, 0,
                                                     5, 5, "Insert New Customer")

    customer_name_entry_label = Inv_Label(customer_menu_frame, 1, 0, 5, "Customer Name:")
    customer_name_entry = Inv_Entrybox(customer_menu_frame, 1, 1, 5, "Name")
    customer_email_entry_label = Inv_Label(customer_menu_frame, 1, 2, 5, "Customer Email Address:")
    customer_email_entry = Inv_Entrybox(customer_menu_frame, 1, 3, 5, "Email Address")
    customer_contact_entry_label = Inv_Label(customer_menu_frame, 1, 4, 5, "Customer Contact No.:")
    customer_contact_entry = Inv_Entrybox(customer_menu_frame, 1, 5, 5, "Contact No.")

    addcustomer_btn = Inv_Button(customer_menu_frame, 2, 0, 5, "#007FFF", "Add", add_new_customer_details)
    editcustomer_btn = Inv_Button(customer_menu_frame, 2, 2, 5, "#ADD8E6", "Edit", edit_customer_details)
    deletecustomer_btn = Inv_Button(customer_menu_frame, 2, 4, 5, "#ff666d", "Delete", delete_customer_record)

    search_customer_entry = customtkinter.CTkEntry(tab_1, placeholder_text="Search", width=1050, )
    search_customer_entry.grid(row=1, column=0, padx=10, pady=10)
    search_customer_entry.bind("<KeyRelease>", search_customer)

    tab_1.columnconfigure(0, weight=2)

    # -----------------------------------------------------------------------------------

    product_table_frame = customtkinter.CTkFrame(master=tab_3, width=1280, height=515)
    product_table_frame.place(x=0, y=205)

    admin_product_tree = Inv_product_tree_display(product_table_frame, on_product_double_click, display_product_record)

    search_entry = customtkinter.CTkEntry(tab_3, placeholder_text="Search", width=1050, )
    search_entry.grid(row=1, column=0, padx=10, pady=10)
    search_entry.bind("<KeyRelease>", search_product)

    product_menu_frame = customtkinter.CTkFrame(
        master=tab_3,
        border_width=2
    )
    product_menu_frame.grid(row=0, column=0)

    insert_product_data_label = Inv_main_titlelabel(product_menu_frame, 0, 0, 5, 5, "Insert New Product")
    product_id_entry_label = Inv_Label(product_menu_frame, 1, 0, 5, "Product ID:")
    product_id_entry = Inv_Entrybox(product_menu_frame, 1, 1, 5, "ID")
    product_name_entry_label = Inv_Label(product_menu_frame, 1, 2, 5, "Product Name:")
    product_name_entry = Inv_Entrybox(product_menu_frame, 1, 3, 5, "Item")
    product_quantity_entry_label = Inv_Label(product_menu_frame, 1, 4, 5, "Product Quantity:")
    product_quantity_entry = Inv_Entrybox(product_menu_frame, 1, 5, 5, "Quantity")
    product_description_entry_label = Inv_Label(product_menu_frame, 2, 0, 5, "Product Description:")
    product_description_entry = Inv_Entrybox(product_menu_frame, 2, 1, 5, "Description")
    product_min_stock_entry_label = Inv_Label(product_menu_frame, 2, 2, 5, "Product Min. Stock:")
    product_min_stock_entry = Inv_Entrybox(product_menu_frame, 2, 3, 5, "Min. Stock")

    addproduct_btn = Inv_Button(product_menu_frame, 3, 0, 20, "#007FFF", "Add Product", add_new_product_details)
    editproduct_btn = Inv_Button(product_menu_frame, 3, 2, 20, "#ADD8E6", "Edit Product", edit_product_details)
    deleteproduct_btn = Inv_Button(product_menu_frame, 3, 4, 20, "#ff666d", "Delete Product", delete_product_record)

    product_menu_frame.columnconfigure(1, weight=3)

    tab_3.columnconfigure(0, weight=2)

    task_table_frame = customtkinter.CTkFrame(master=tab_4, width=1280, height=515)
    task_table_frame.place(x=0, y=205)

    task_tree = ttk.Treeview(master=task_table_frame, height=18)

    task_verscrlbar = customtkinter.CTkScrollbar(
        master=task_table_frame, orientation="vertical", command=task_tree.yview
    )
    task_verscrlbar.pack(side="right", fill="y")

    task_tree.configure(yscrollcommand=task_verscrlbar.set)
    task_tree["columns"] = ("TASKID", "TASKASSIGNDATE", "TASKDESCRIPTION", "TASKSTATUS", "DUEDATE", "FINISHDATE")

    task_tree.column("#0", width=0, stretch=tk.NO)
    task_tree.column("TASKID", anchor=tk.CENTER, width=100)
    task_tree.column("TASKASSIGNDATE", anchor=tk.CENTER, width=150)
    task_tree.column("TASKDESCRIPTION", anchor=tk.CENTER, width=400)
    task_tree.column("TASKSTATUS", anchor=tk.CENTER, width=300)
    task_tree.column("DUEDATE", anchor=tk.CENTER, width=150)
    task_tree.column("FINISHDATE", anchor=tk.CENTER, width=150)

    task_tree.heading("TASKID", text="Task ID")
    task_tree.heading("TASKASSIGNDATE", text="Task Assigned Date")
    task_tree.heading("TASKDESCRIPTION", text="Task Description")
    task_tree.heading("TASKSTATUS", text="Status")
    task_tree.heading("DUEDATE", text="Due Date")
    task_tree.heading("FINISHDATE", text="Task Finish Date")

    task_tree.pack(side="bottom", fill="both")
    #task_tree.bind("<ButtonRelease>", display_product_record)

    task_summary_frame = customtkinter.CTkFrame(tab_4, corner_radius=10, border_width=2)
    task_summary_frame.grid(row=0, column=0, padx=10, pady=10)

    task_summary_title = Inv_titlelabel(task_summary_frame,0,0,10,"Task Summary ("+username+")")

    completed_tasks_label_1 = customtkinter.CTkLabel(
        master=task_summary_frame,
        text=str(fetch_completed_task_data()),
        font=customtkinter.CTkFont("SF Pro Display", size=13),
    )

    completed_tasks_label_2 = customtkinter.CTkLabel(
        master=task_summary_frame,
        text="Completed Task",
        font=customtkinter.CTkFont("SF Pro Display", size=13),
    )

    task_summary_vertical_separator_1 = customtkinter.CTkFrame(
        task_summary_frame, width=2, height=50, fg_color="#CCCCCC"
    )

    pending_tasks_label_1 = customtkinter.CTkLabel(
        master=task_summary_frame,
        text=str(fetch_pending_task_data()),
        font=customtkinter.CTkFont("SF Pro Display", size=13),
    )

    pending_tasks_label_2 = customtkinter.CTkLabel(
        master=task_summary_frame,
        text="Pending Task",
        font=customtkinter.CTkFont("SF Pro Display", size=13),
    )

    task_summary_vertical_separator_2 = customtkinter.CTkFrame(
        task_summary_frame, width=2, height=50, fg_color="#CCCCCC"
    )

    overdue_tasks_label_1 = customtkinter.CTkLabel(
        master=task_summary_frame,
        text=str(fetch_overdue_task_data()),
        font=customtkinter.CTkFont("SF Pro Display", size=13),
    )

    overdue_tasks_label_2 = customtkinter.CTkLabel(
        master=task_summary_frame,
        text="Overdue Task",
        font=customtkinter.CTkFont("SF Pro Display", size=13),
    )

    task_summary_vertical_separator_3 = customtkinter.CTkFrame(
        task_summary_frame, width=2, height=50, fg_color="#CCCCCC"
    )

    completed_tasks_label_1.grid(row=1, column=0, padx=20, pady=5)
    completed_tasks_label_2.grid(row=2, column=0, padx=20, pady=5)
    task_summary_vertical_separator_1.grid(row=1, column=1, rowspan=2, padx=5, pady=5)
    pending_tasks_label_1.grid(row=1, column=2, padx=20, pady=5)
    pending_tasks_label_2.grid(row=2, column=2, padx=20, pady=5)
    task_summary_vertical_separator_2.grid(row=1, column=3, rowspan=2, padx=5, pady=5)
    overdue_tasks_label_1.grid(row=1, column=4, padx=20, pady=5)
    overdue_tasks_label_2.grid(row=2, column=4, padx=20, pady=5)
    # task_summary_vertical_separator_3.grid(row=1, column=5, rowspan=2, padx=5, pady=5)

    completetask_btn = customtkinter.CTkButton(
        master=tab_4,
        text="Complete Task",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=complete_task,
        compound="top",
        corner_radius=200,
        text_color="black",
        width=150,
        height=50,
    )

    completetask_btn.grid(row=1, column=0, padx=5, pady=5)

    tab_4.columnconfigure(0, weight=2)

    incoming_stock_table_frame = customtkinter.CTkFrame(
        master=tab_5, width=1006, height=515
    )
    incoming_stock_table_frame.place(x=0, y=205)

    incoming_stock_tree = ttk.Treeview(master=incoming_stock_table_frame, height=18)

    incoming_stock_verscrlbar = customtkinter.CTkScrollbar(
        master=incoming_stock_table_frame,
        orientation="vertical",
        command=incoming_stock_tree.yview,
    )
    incoming_stock_verscrlbar.pack(side="right", fill="y")

    incoming_stock_tree.configure(yscrollcommand=incoming_stock_verscrlbar.set)
    incoming_stock_tree["columns"] = (
        "INCOMINGSTOCKID",
        "INCOMINGSTOCKNAME",
        "INCOMINGSTOCKQUANTITY",
        "INCOMINGSTOCKSTATUS",
    )

    incoming_stock_tree.column("#0", width=0, stretch=tk.NO)
    incoming_stock_tree.column("INCOMINGSTOCKID", anchor=tk.CENTER, width=310)
    incoming_stock_tree.column("INCOMINGSTOCKNAME", anchor=tk.CENTER, width=310)
    incoming_stock_tree.column("INCOMINGSTOCKQUANTITY", anchor=tk.CENTER, width=310)
    incoming_stock_tree.column("INCOMINGSTOCKSTATUS", anchor=tk.CENTER, width=320)

    incoming_stock_tree.heading("INCOMINGSTOCKID", text="ID")
    incoming_stock_tree.heading("INCOMINGSTOCKNAME", text="Name")
    incoming_stock_tree.heading("INCOMINGSTOCKQUANTITY", text="Quantity")
    incoming_stock_tree.heading("INCOMINGSTOCKSTATUS", text="Status")

    incoming_stock_tree.pack(side="bottom", fill="both")

    incoming_stock_menu_frame = customtkinter.CTkFrame(master=tab_5, border_width=2)
    incoming_stock_menu_frame.grid(row=0, column=0, pady=10)

    incoming_stock_tree.tag_configure(
        "To be Received", background="#FFBA00", foreground="black"
    )
    incoming_stock_tree.tag_configure(
        "Received", background="green", foreground="black"
    )

    edit_purchase_order_status_label = Inv_main_titlelabel(incoming_stock_menu_frame, 0, 0,
                                                           10, 10, "Edit Purchase Order Status")
    purchase_order_status_label = customtkinter.CTkLabel(
        master=incoming_stock_menu_frame,
        text="Status:",
        font=customtkinter.CTkFont("SF Pro Display"),
    )
    purchase_order_status_entry = customtkinter.CTkComboBox(
        master=incoming_stock_menu_frame,
        values=["Received"],
        width=230,
        height=30,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )
    editpurchaseorder_btn = customtkinter.CTkButton(
        master=incoming_stock_menu_frame,
        text="Edit Status",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=update_purchase_order_status,
        compound="top",
        corner_radius=200,
        fg_color="#007FFF",
        text_color="black",
    )

    purchase_order_status_label.grid(row=1, column=0, padx=5, pady=5)
    purchase_order_status_entry.grid(row=1, column=1, padx=5, pady=5)
    editpurchaseorder_btn.grid(row=2, column=0, padx=5, pady=5, columnspan=2)

    search_incoming_stock_entry = customtkinter.CTkEntry(tab_5, placeholder_text="Search", width=1050, )
    search_incoming_stock_entry.grid(row=1, column=0, padx=10, pady=10)
    search_incoming_stock_entry.bind("<KeyRelease>", search_incoming_stock)

    tab_5.columnconfigure(0, weight=1)

    outgoing_stock_table_frame = customtkinter.CTkFrame(
        master=tab_6, width=1006, height=515
    )
    outgoing_stock_table_frame.place(x=0, y=205)
    outgoing_stock_tree = Inv_outgoing_stock_tree_display(outgoing_stock_table_frame, on_sale_order_double_click)

    outgoing_stock_menu_frame = customtkinter.CTkFrame(master=tab_6, border_width=2)
    outgoing_stock_menu_frame.grid(row=0, column=0, pady=10)

    edit_sales_order_status_label = Inv_main_titlelabel(outgoing_stock_menu_frame, 0, 0,
                                                        10, 10, "Edit Sales Order Status")
    sales_order_status_label = customtkinter.CTkLabel(
        master=outgoing_stock_menu_frame,
        text="Status:",
        font=customtkinter.CTkFont("SF Pro Display"),
    )
    sales_order_status_entry = customtkinter.CTkComboBox(
        master=outgoing_stock_menu_frame,
        values=["To be Packed", "To be Shipped", "To be Delivered", "Delivered"],
        width=230,
        height=30,
        font=customtkinter.CTkFont("SF Pro Display"),
        border_width=0,
    )
    editsalesorder_btn = customtkinter.CTkButton(
        master=outgoing_stock_menu_frame,
        text="Edit Status",
        font=customtkinter.CTkFont("SF Pro Display"),
        command=update_sales_order_status,
        compound="top",
        corner_radius=200,
        fg_color="#007FFF",
        text_color="black",
    )

    sales_order_status_label.grid(row=1, column=0, padx=5, pady=5)
    sales_order_status_entry.grid(row=1, column=1, padx=5, pady=5)
    editsalesorder_btn.grid(row=2, column=0, padx=5, pady=5, columnspan=2)

    search_outgoing_stock_entry = customtkinter.CTkEntry(tab_6, placeholder_text="Search", width=1050, )
    search_outgoing_stock_entry.grid(row=1, column=0, padx=10, pady=10)
    search_outgoing_stock_entry.bind("<KeyRelease>", search_outgoing_stock)

    outgoing_product_table_frame = customtkinter.CTkFrame(master=tab_6, width=506, height=515)
    outgoing_product_table_frame.place(x=740, y=205)
    # create object
    outgoing_product_tree = Inv_outgoing_product_tree_display(outgoing_product_table_frame)

    add_to_outgoing_product_table()

    tab_6.columnconfigure(0, weight=2)

    add_to_customer_table()
    #add_to_supplier_table()
    add_to_product_table()
    add_to_task_table()
    check_task_status()
    check_pending_task()
    check_task_due_date()
    add_to_incoming_stock_table()
    add_to_sale_order_table()

# -----------------------------------------------

#login_page()
admin_dashboard("admin")
#supervisor_dashboard("supervisor")
#worker_dashboard("worker1")
app.mainloop()